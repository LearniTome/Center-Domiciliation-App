import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Optional
import logging
import os
import subprocess
import sys
import threading
from contextlib import nullcontext
from .societe_form import SocieteForm
from .associe_form import AssocieForm
from .contrat_form import ContratForm
from .collaborateur_form import CollaborateurForm
from ..utils.utils import ThemeManager, WidgetFactory, WindowManager, PathManager, ensure_excel_db
from ..utils import constants as _const
from pathlib import Path

logger = logging.getLogger(__name__)

class MainForm(ttk.Frame):
    def __init__(self, parent, values_dict=None, startup_profiler=None):
        super().__init__(parent)
        self.parent = parent
        self.values = values_dict or {}
        self.startup_profiler = startup_profiler

        # Initialize navigation button attributes with explicit types so
        # linters/type checkers know these exist and accept Button assignments.
        # They are attached by the outer application toolbar (main.py)
        # to keep a single unified control row.
        self.prev_btn: Optional[ttk.Button] = None
        self.next_btn: Optional[ttk.Button] = None
        self.save_btn: Optional[ttk.Button] = None
        self.finish_btn: Optional[ttk.Button] = None
        self.generate_btn: Optional[ttk.Button] = None
        self.config_btn: Optional[ttk.Button] = None
        self.quit_btn: Optional[ttk.Button] = None
        self._dashboard_window: Optional[tk.Toplevel] = None
        self._last_collaborateur_name: str = ''
        self.toolbar_style_names = {
            'secondary': 'Toolbar.Secondary.TButton',
            'success': 'Toolbar.Success.TButton',
            'cancel': 'Toolbar.Cancel.TButton',
            'copy': 'Toolbar.Copy.TButton',
        }
        self.toolbar_button_widths = {
            'tools': 14,
            'dashboard': 18,
            'new': 12,
            'save': 14,
            'finish': 12,
            'prev': 12,
            'next': 12,
            'quit': 12,
        }
        self.next_default_style: str = self.toolbar_style_names['secondary']
        self.next_finish_style: str = self.toolbar_style_names['success']
        self.next_default_width: int = self.toolbar_button_widths['next']
        self.next_finish_width: int = self.toolbar_button_widths['finish']

        # Allow this main frame to expand inside its parent
        try:
            self.pack(fill="both", expand=True)
        except Exception:
            pass

        # Initialize theme manager
        self.theme_manager = ThemeManager(self.winfo_toplevel())
        self.style = self.theme_manager.style

        # Create scrollable container
        with self._profile_scope("MainForm.setup_scrollable_container"):
            self.setup_scrollable_container()

        # Create forms
        with self._profile_scope("MainForm.setup_forms"):
            self.setup_forms()

    def _profile_scope(self, name: str):
        if self.startup_profiler is None:
            return nullcontext()
        try:
            return self.startup_profiler.span(name)
        except Exception:
            return nullcontext()

    def setup_scrollable_container(self):
        """Configure the scrollable container for all forms"""
        # Create main container
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)

        # Create canvas and scrollbar
        self.canvas = tk.Canvas(container, background=self.theme_manager.colors['bg'])
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)

        # Create main container for forms
        self.forms_container = ttk.Frame(self.canvas)
        # Make sure forms container will expand within the canvas window
        self.forms_container.grid_columnconfigure(0, weight=1)
        # Allow multiple rows to stretch if necessary
        # Pages are placed at row=1 below any header rows; set row 0..2 to stretch
        for r in range(0, 3):
            self.forms_container.grid_rowconfigure(r, weight=1)

        # Create the canvas window with auto-width
        self.canvas_window = self.canvas.create_window(
            (0, 0), window=self.forms_container, anchor="nw"
        )

        # Configure scroll region when forms are updated
        def _on_frame_configure(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        def _on_canvas_configure(event):
            # Update the width to expand to the canvas width
            self.canvas.itemconfig(self.canvas_window, width=event.width)

        self.forms_container.bind("<Configure>", _on_frame_configure)
        self.canvas.bind("<Configure>", _on_canvas_configure)

        # Configure mouse wheel scrolling
        def _on_mousewheel(event):
            if event.state & 0x4: # Check if Control key is pressed
                # Zoom with Ctrl + Mouse wheel
                return
            # Ignore wheel events coming from modal/top-level dialogs.
            try:
                source_widget = event.widget
                source_top = source_widget.winfo_toplevel() if source_widget is not None else None
                main_top = self.winfo_toplevel()
                if source_top is not None and source_top is not main_top:
                    return
            except Exception:
                pass

            try:
                delta = int(getattr(event, "delta", 0))
            except Exception:
                delta = 0
            if delta == 0:
                return
            steps = max(1, abs(delta) // 120)
            direction = -1 if delta > 0 else 1
            self.canvas.yview_scroll(direction * steps, "units")
            return "break"

        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Grid layout for better control — make canvas expand
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Footer container for navigation (fixed, not inside the scroll area)
        # Placed directly inside this frame so it stays visible when content scrolls
        # Use Card.TFrame style if available so footer is visually separated
        try:
            self.footer_container = ttk.Frame(self, style='Card.TFrame')
        except Exception:
            self.footer_container = ttk.Frame(self)
        # Keep footer fixed height and add padding so content doesn't overlap
        self.footer_container.pack(fill="x", padx=10, pady=(8, 10))
        # Internal container to host navigation widgets consistently
        self.footer_inner = ttk.Frame(self.footer_container)
        self.footer_inner.pack(fill="x")

    def setup_forms(self):
        """Create one page per logical section and prepare navigation.

        Previously the form showed all sections on a single screen. We now
        create one frame (page) per section and provide Next/Previous
        navigation as well as Save and Finish actions.
        """
        # Container for pages
        self.pages = []
        self.current_page = 0

        # Create pages
        with self._profile_scope("MainForm.create_societe_page"):
            self.create_societe_page()
        with self._profile_scope("MainForm.create_associe_page"):
            self.create_associe_page()
        with self._profile_scope("MainForm.create_contrat_page"):
            self.create_contrat_page()
        self._bind_cross_form_synchronization()

        # Show first page
        with self._profile_scope("MainForm.show_page.initial"):
            self.show_page(0)

        # Setup navigation controls
        with self._profile_scope("MainForm.setup_navigation"):
            self.setup_navigation()

    def create_section_header(self, parent, text, icon, row, column, columnspan=1):
        """Crée un en-tête de section stylisé"""
        header_frame = ttk.Frame(parent, style='Section.TFrame')
        # Ensure the header frame expands horizontally within its parent
        try:
            header_frame.grid(row=row, column=column, columnspan=columnspan,
                              sticky="ew", pady=(0, 10), padx=5)
            header_frame.grid_columnconfigure(0, weight=1)
        except Exception:
            # Fallback to pack if grid placement isn't supported
            header_frame.pack(fill="x", pady=(0, 10), padx=5)

        label = ttk.Label(
            header_frame,
            text=f"{icon} {text}",
            style='SectionHeader.TLabel'
        )
        label.pack(fill="x", expand=True, padx=6, pady=4)

        return header_frame

    def create_societe_page(self):
        # Create a dedicated page container and put the header inside it so
        # the header shows/hides together with the page.
        # Use a plain frame for the page so the header's styled frame provides
        # the visual section border/strip; avoids nested Section.TFrame borders
        page = ttk.Frame(self.forms_container)
        page.grid(row=1, column=0, sticky="nsew", padx=5, pady=(0, 10))
        # Make the page expand so its inner form can stretch
        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=1)
        page.grid_columnconfigure(0, weight=1)
        header = self.create_section_header(page, "Informations de la Société", "📝", 0, 0)
        with self._profile_scope("SocieteForm.__init__"):
            self.societe_form = SocieteForm(page, self.theme_manager, self.values.get('societe', {}))
        self.societe_form.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.pages.append(('societe', page, self.societe_form))

    def create_associe_page(self):
        page = ttk.Frame(self.forms_container)
        page.grid(row=1, column=0, sticky="nsew", padx=5, pady=(0, 10))
        # Expand the page so the associe list/canvas can grow
        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=1)
        page.grid_columnconfigure(0, weight=1)
        header = self.create_section_header(page, "Informations des Associés", "👥", 0, 0)
        # AssocieForm expects a ThemeManager instance
        with self._profile_scope("AssocieForm.__init__"):
            self.associe_form = AssocieForm(
                page,
                self.theme_manager,
                get_societe_totals=self._get_societe_totals_for_associes,
            )
        self.associe_form.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        try:
            self.associe_form.bind_societe_totals_vars(
                getattr(self.societe_form, 'capital_var', None),
                getattr(self.societe_form, 'parts_social_var', None),
            )
        except Exception:
            pass
        # Defensive: remove any accidental scrollbar widgets that might have been
        # created inside the page by older code or third-party widgets. We want
        # a single scrollbar (the one on the main form canvas) to control
        # vertical scrolling for the whole page.
        for child in list(page.winfo_children()):
            try:
                if isinstance(child, ttk.Scrollbar):
                    child.destroy()
            except Exception:
                pass
        # Ensure initial associates based on legal form (SARL AU: 1, SARL: 2, others: 1)
        try:
            legal_form = ''
            try:
                legal_form = self.societe_form.forme_jur_var.get()
            except Exception:
                legal_form = ''
            with self._profile_scope("AssocieForm.apply_legal_form_constraints"):
                self.associe_form.apply_legal_form_constraints(
                    legal_form,
                    auto_adjust=True,
                    add_if_empty=True,
                )
        except Exception:
            # conservative: ignore errors here to avoid breaking startup
            pass
        self.pages.append(('associes', page, self.associe_form))

    def _get_societe_totals_for_associes(self):
        """Return company capital/parts totals for associate distribution."""
        try:
            if hasattr(self, 'societe_form') and self.societe_form is not None:
                cap = self.societe_form.capital_var.get() if hasattr(self.societe_form, 'capital_var') else ''
                parts = self.societe_form.parts_social_var.get() if hasattr(self.societe_form, 'parts_social_var') else ''
                return cap, parts
        except Exception:
            pass
        return '', ''

    def _map_contract_type_from_legal_form(self, legal_form: str) -> str:
        form = str(legal_form or '').strip().lower()
        if form == 'personne physique':
            return 'Personne Physique'
        if form == 'association':
            return 'Association'
        if form == 'fondation':
            return 'Fondation'
        return 'Personne Morale'

    def _sync_legal_form_dependents(self):
        """Synchronize legal-form dependent fields across sections."""
        try:
            legal_form = str(getattr(self.societe_form, 'forme_jur_var', None).get() or '').strip()
        except Exception:
            legal_form = ''
        if not legal_form:
            return

        try:
            if hasattr(self, 'contrat_form') and hasattr(self.contrat_form, 'type_contrat_domiciliation_var'):
                self.contrat_form.type_contrat_domiciliation_var.set(
                    self._map_contract_type_from_legal_form(legal_form)
                )
        except Exception:
            pass

        try:
            if hasattr(self, 'associe_form') and self.associe_form is not None:
                self.associe_form.apply_legal_form_constraints(
                    legal_form,
                    auto_adjust=True,
                    add_if_empty=False,
                )
        except Exception:
            pass

    def _on_societe_legal_form_changed(self, *_args):
        self._sync_legal_form_dependents()

    def _bind_cross_form_synchronization(self):
        """Bind one-way synchronization from Société legal form to other sections."""
        try:
            var = getattr(self.societe_form, 'forme_jur_var', None)
            if var is None:
                return
            try:
                self._societe_legal_form_trace_id = var.trace_add('write', self._on_societe_legal_form_changed)
            except Exception:
                var.trace('w', self._on_societe_legal_form_changed)
            self._sync_legal_form_dependents()
        except Exception:
            pass

    def create_contrat_page(self):
        page = ttk.Frame(self.forms_container)
        page.grid(row=1, column=0, sticky="nsew", padx=5, pady=(0, 10))
        # Expand the page so its inner form can stretch
        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=1)
        page.grid_columnconfigure(0, weight=1)
        header = self.create_section_header(page, "Informations du Contrat", "📋", 0, 0)
        with self._profile_scope("ContratForm.__init__"):
            self.contrat_form = ContratForm(
                page,
                self.theme_manager,
                self.values.get('contrat', {}),
                on_add_collaborateur=self._open_collaborateur_from_contrat,
                get_collaborateur_name=self._get_current_collaborateur_name,
            )
        self.contrat_form.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.pages.append(('contrat', page, self.contrat_form))

    def _open_collaborateur_from_contrat(self):
        self._open_collaborateur_dialog()

    def _get_current_collaborateur_name(self) -> str:
        try:
            return str(self._last_collaborateur_name or '').strip()
        except Exception:
            return ''

    def _open_collaborateur_dialog(self, values: Optional[dict] = None, edit_context: Optional[dict] = None):
        dialog = tk.Toplevel(self)
        dialog.title("Collaborateur")
        try:
            dialog.configure(bg=self.theme_manager.colors.get('bg', '#1f1f1f'))
        except Exception:
            pass
        try:
            dialog.transient(self.winfo_toplevel())
            dialog.grab_set()
        except Exception:
            pass

        body = ttk.Frame(dialog, style='Card.TFrame')
        body.pack(fill="both", expand=True, padx=12, pady=10)

        form = CollaborateurForm(body, self.theme_manager, values or {})
        form.pack(fill="both", expand=True)

        btn_row = ttk.Frame(dialog)
        btn_row.pack(fill="x", padx=12, pady=(0, 10))

        def _on_save():
            vals = form.get_values()
            if not any(str(v or '').strip() for v in vals.values()):
                messagebox.showwarning("Collaborateur", "Veuillez saisir au moins le nom du collaborateur.")
                return
            saved = self._save_collaborateur_to_db(vals, edit_context=edit_context)
            if saved:
                try:
                    if hasattr(self, 'contrat_form') and self.contrat_form is not None:
                        self.contrat_form.refresh_collaborateur_names()
                        nom = str(vals.get('nom') or '').strip()
                        if nom:
                            self.contrat_form.collaborateur_nom_var.set(nom)
                except Exception:
                    pass
                dialog.destroy()

        def _on_close():
            dialog.destroy()

        WidgetFactory.create_button(
            btn_row,
            text="💾 Enregistrer",
            command=_on_save,
            style='Success.TButton',
        ).pack(side='right', padx=6)
        WidgetFactory.create_button(
            btn_row,
            text="Annuler",
            command=_on_close,
            style='Secondary.TButton',
        ).pack(side='right')

        try:
            WindowManager.center_window(dialog)
        except Exception:
            pass

    def _save_collaborateur_to_db(self, values: dict, edit_context: Optional[dict] = None) -> Optional[Path]:
        try:
            db_path = Path(PathManager.DATABASE_DIR) / _const.DB_FILENAME
            from ..utils.utils import ensure_excel_db, write_records_to_db, normalize_excel_storage
            ensure_excel_db(db_path, _const.excel_sheets)
            updated = False
            if edit_context:
                try:
                    import pandas as _pd
                    from ..utils.utils import normalize_canonical_dataframe_for_storage

                    df = _pd.read_excel(db_path, sheet_name='Collaborateurs', dtype=str).fillna('')
                    alias_map = getattr(_const, 'collaborateur_header_aliases', {}) or {}
                    if alias_map:
                        for old_col, new_col in alias_map.items():
                            if old_col not in df.columns:
                                continue
                            if new_col not in df.columns:
                                df[new_col] = df[old_col]
                            else:
                                try:
                                    old_vals = df[old_col].fillna('').astype(str).str.strip()
                                    new_vals = df[new_col].fillna('').astype(str).str.strip()
                                    mask = (new_vals == '') & (old_vals != '')
                                    df.loc[mask, new_col] = df.loc[mask, old_col]
                                except Exception:
                                    pass
                            try:
                                df.drop(columns=[old_col], inplace=True)
                            except Exception:
                                pass

                    collab_id = str(edit_context.get('id_collaborateur') or '').strip()
                    sid = str(edit_context.get('id_societe') or '').strip()
                    mask = None
                    if collab_id and 'id_collaborateur' in df.columns:
                        mask = df['id_collaborateur'].astype(str).str.strip() == collab_id
                    elif sid and 'id_societe' in df.columns:
                        mask = df['id_societe'].astype(str).str.strip() == sid
                        name = str(values.get('nom') or '').strip()
                        if name and 'collaborateur_nom' in df.columns:
                            mask = mask & (df['collaborateur_nom'].astype(str).str.strip().str.lower() == name.lower())

                    if mask is not None and not df.empty and mask.any():
                        mapping = {
                            'type': 'collaborateur_type',
                            'code': 'collaborateur_code',
                            'nom': 'collaborateur_nom',
                            'ice': 'collaborateur_ice',
                            'tp': 'collaborateur_tp',
                            'rc': 'collaborateur_rc',
                            'if': 'collaborateur_if',
                            'tel_fixe': 'collaborateur_tel_fixe',
                            'tel_mobile': 'collaborateur_tel_mobile',
                            'adresse': 'collaborateur_adresse',
                            'email': 'collaborateur_email',
                        }
                        for k, col in mapping.items():
                            if col in df.columns:
                                df.loc[mask, col] = values.get(k)
                        if sid and 'id_societe' in df.columns:
                            df.loc[mask, 'id_societe'] = sid

                        df = normalize_canonical_dataframe_for_storage(
                            df.reindex(columns=_const.collaborateur_headers, fill_value='')
                        )
                        try:
                            with _pd.ExcelWriter(db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                df.to_excel(writer, sheet_name='Collaborateurs', index=False)
                        except TypeError:
                            from openpyxl import load_workbook
                            wb = load_workbook(db_path)
                            if 'Collaborateurs' in wb.sheetnames:
                                try:
                                    wb.remove(wb['Collaborateurs'])
                                except Exception:
                                    pass
                            wb.save(db_path)
                            with _pd.ExcelWriter(db_path, engine='openpyxl', mode='a') as writer:
                                df.to_excel(writer, sheet_name='Collaborateurs', index=False)
                        normalize_excel_storage(db_path)
                        updated = True
                except Exception:
                    updated = False

            if not updated:
                write_records_to_db(db_path, {}, [], {}, values)
                normalize_excel_storage(db_path)
            try:
                self._last_collaborateur_name = str(values.get('nom') or '').strip()
            except Exception:
                self._last_collaborateur_name = ''
            try:
                if edit_context and getattr(self, '_dashboard_window', None):
                    self._dashboard_window._load_data()
                    self._dashboard_window._show_page(getattr(self._dashboard_window, '_current_page', 'societe'))
            except Exception:
                pass
            try:
                if edit_context:
                    self._dashboard_edit_context = None
            except Exception:
                pass
            messagebox.showinfo("Collaborateur", "Collaborateur enregistré avec succès.")
            return db_path
        except Exception as e:
            try:
                from ..utils.utils import ErrorHandler
                ErrorHandler.handle_error(e, "Erreur lors de l'enregistrement du collaborateur.")
            except Exception:
                messagebox.showerror("Erreur", f"Impossible d'enregistrer le collaborateur: {e}")
            return None

    def create_collapsible_section(self, title, form_creator):
        """Create a collapsible section with the given title and form"""
        section = ttk.Frame(self.forms_container)
        section.pack(fill="x", padx=20, pady=10)

        # Header with title and toggle button
        header = ttk.Frame(section)
        header.pack(fill="x", pady=5)

        title_label = ttk.Label(
            header,
            text=title,
            font=('Segoe UI', 12, 'bold'),
            foreground=self.theme_manager.colors['accent']
        )
        title_label.pack(side="left")

        # Container for the form
        content_frame = ttk.Frame(section)
        content_frame.pack(fill="x", pady=5)

        # Create the form
        form = form_creator(content_frame)
        form.pack(fill="x", expand=True)

        # Toggle button
        def toggle_section():
            if content_frame.winfo_viewable():
                content_frame.pack_forget()
                toggle_btn.configure(text="🔽")
            else:
                content_frame.pack(fill="x", pady=5)
                toggle_btn.configure(text="🔼")

        toggle_btn = WidgetFactory.create_button(
            header,
            text="🔼",
            command=toggle_section,
            style='Secondary.TButton',
            tooltip="Afficher/Masquer la section"
        )
        toggle_btn.pack(side="right")

        return section

    def setup_navigation(self):
        """Create navigation controls for the wizard pages.

        Adds Previous, Next, Save and Finish buttons. Previous/Next navigate
        between created pages. Save stores the current page's values into
        `self.values`. Finish will save then emit an event `<<FormsFinished>>`.
        """
        # Navigation buttons are created at the app-level toolbar (in main.py)
        # to keep a single, unified control row. Here we only ensure nav button
        # states are initialized and keyboard shortcuts are bound.
        try:
            # Ensure attributes exist so update_nav_buttons can safely introspect
            if not hasattr(self, 'prev_btn'):
                self.prev_btn = None
            if not hasattr(self, 'next_btn'):
                self.next_btn = None
            if not hasattr(self, 'save_btn'):
                self.save_btn = None
            if not hasattr(self, 'finish_btn'):
                self.finish_btn = None
            if not hasattr(self, 'quit_btn'):
                self.quit_btn = None
        except Exception:
            pass

        # Initialize nav button states if possible
        try:
            self.update_nav_buttons()
        except Exception:
            pass

        # Keyboard shortcuts on the top-level window
        try:
            top = self.winfo_toplevel()
            top.bind('<Left>', lambda e: self.prev_page())
            top.bind('<Right>', lambda e: self.next_page())
            top.bind('<Control-s>', lambda e: self.save_current())
        except Exception:
            pass

    def show_dashboard(self, start_fullscreen: bool = False):
        """Switch to dashboard view."""
        from .dashboard_view import DashboardView
        existing = getattr(self, '_dashboard_window', None)
        try:
            if existing is not None and existing.winfo_exists():
                try:
                    existing.deiconify()
                except Exception:
                    pass
                existing.lift()
                existing.focus_force()
                if start_fullscreen:
                    try:
                        existing._enter_fullscreen()
                    except Exception:
                        pass
                return existing
        except Exception:
            self._dashboard_window = None

        dashboard = DashboardView(
            self.winfo_toplevel(),
            action_handler=self.handle_dashboard_action,
            start_fullscreen=start_fullscreen,
        )
        self._dashboard_window = dashboard
        return dashboard

    def configure_main_toolbar_styles(self):
        """Create main-toolbar style aliases once, based on the app theme."""
        toolbar_font = ('Segoe UI', 10, 'bold')
        toolbar_padding = (12, 7)

        def _clone_style(new_style: str, base_style: str):
            cfg = {'font': toolbar_font, 'padding': toolbar_padding}
            for opt in ('background', 'foreground', 'relief', 'borderwidth'):
                try:
                    val = self.style.lookup(base_style, opt)
                except Exception:
                    val = None
                if val not in (None, ''):
                    cfg[opt] = val
            self.style.configure(new_style, **cfg)

            state_map = {}
            for opt in ('background', 'foreground'):
                try:
                    mapped = self.style.map(base_style, query_opt=opt)
                except Exception:
                    mapped = []
                if mapped:
                    state_map[opt] = mapped
            if state_map:
                self.style.map(new_style, **state_map)

        try:
            _clone_style(self.toolbar_style_names['secondary'], 'Secondary.TButton')
            _clone_style(self.toolbar_style_names['success'], 'Success.TButton')
            _clone_style(self.toolbar_style_names['cancel'], 'Cancel.TButton')
            _clone_style(self.toolbar_style_names['copy'], 'Copy.TButton')
        except Exception:
            pass

    def _get_main_toolbar_clear_command(self):
        top = self.winfo_toplevel()
        clear_fn = getattr(top, 'clear_form', None)
        if callable(clear_fn):
            return clear_fn
        return self.reset

    def _get_main_toolbar_quit_command(self):
        return self.return_to_dashboard

    def return_to_dashboard(self, start_fullscreen: bool = True):
        """Hide the generator and show the dashboard instead of quitting."""
        try:
            dashboard = self.show_dashboard(start_fullscreen=start_fullscreen)
            try:
                if dashboard is not None:
                    dashboard._load_data()
                    dashboard._show_page(getattr(dashboard, '_current_page', 'societe'))
            except Exception:
                pass
        except Exception:
            pass
        try:
            self.winfo_toplevel().withdraw()
        except Exception:
            pass

    def get_main_toolbar_button_specs(self):
        """Return button definitions for the main application toolbar."""
        secondary = self.toolbar_style_names['secondary']
        success = self.toolbar_style_names['success']
        cancel = self.toolbar_style_names['cancel']
        copy = self.toolbar_style_names['copy']
        widths = self.toolbar_button_widths

        left = [
            {
                'key': 'dashboard',
                'text': '📊 Tableau de bord',
                'command': lambda: self.show_dashboard(start_fullscreen=True),
                'style': secondary,
                'width': widths['dashboard'],
            },
        ]
        right = [
            {
                'key': 'quit',
                'text': '❌ Quitter',
                'command': self._get_main_toolbar_quit_command(),
                'style': cancel,
                'width': widths['quit'],
            },
            {
                'key': 'next',
                'text': 'Suivant ▶',
                'command': self.next_page,
                'style': self.next_default_style,
                'width': self.next_default_width,
            },
            {
                'key': 'prev',
                'text': '◀ Précédent',
                'command': self.prev_page,
                'style': secondary,
                'width': widths['prev'],
            },
            {
                'key': 'finish',
                'text': 'Terminer',
                'command': self._trigger_generate_documents,
                'style': self.next_finish_style,
                'width': self.next_finish_width,
                'hidden': True,
            },
            {
                'key': 'save',
                'text': '💾 Sauvegarder',
                'command': self.save_current,
                'style': secondary,
                'width': widths['save'],
            },
            {
                'key': 'new',
                'text': '🆕 Nouvelle',
                'command': self._get_main_toolbar_clear_command(),
                'style': copy,
                'width': widths['new'],
            },
        ]
        return {'left': left, 'right': right}

    def register_main_toolbar_button(self, key: str, button: ttk.Button):
        """Attach main-toolbar widget references used by MainForm state updates."""
        mapping = {
            'tools': 'config_btn',
            'prev': 'prev_btn',
            'next': 'next_btn',
            'save': 'save_btn',
            'finish': 'finish_btn',
            'quit': 'quit_btn',
        }
        attr_name = mapping.get(key)
        if attr_name:
            setattr(self, attr_name, button)

    def open_configuration(self, parent_window=None):
        """Open a lightweight tools launcher."""
        try:
            owner = parent_window or self.winfo_toplevel()
            top = tk.Toplevel(owner)
            top.transient(owner)
            top.title("Outils")
            top.geometry("470x340")
            top.resizable(False, False)

            try:
                top.grab_set()
            except Exception:
                pass

            main_frame = ttk.Frame(top, padding=14)
            main_frame.pack(fill="both", expand=True)

            ttk.Label(
                main_frame,
                text="🧰 Outils",
                font=("Segoe UI", 12, "bold"),
            ).pack(anchor="w", pady=(0, 8))
            ttk.Label(
                main_frame,
                text="Choisissez le module à ouvrir :",
            ).pack(anchor="w", pady=(0, 10))

            actions = ttk.Frame(main_frame)
            actions.pack(fill="x", expand=True)

            def _open_defaults():
                try:
                    top.destroy()
                except Exception:
                    pass
                self._open_defaults_dialog(parent_window=parent_window or owner)

            def _open_analyzer():
                try:
                    top.destroy()
                except Exception:
                    pass
                self._open_template_values_analyzer_dialog(parent_window=parent_window or owner)

            def _open_generator():
                try:
                    top.destroy()
                except Exception:
                    pass
                try:
                    if parent_window and hasattr(parent_window, '_hide_for_parent_switch'):
                        parent_window._hide_for_parent_switch(fullscreen_parent=True)
                    try:
                        self.winfo_toplevel().deiconify()
                    except Exception:
                        pass
                    top_level = self.winfo_toplevel()
                    generate_fn = getattr(top_level, 'generate_documents', None)
                    if callable(generate_fn):
                        generate_fn()
                except Exception:
                    logger.exception("Erreur lors de l'ouverture du générateur de documents")

            def _open_word_pdf_batch():
                try:
                    top.destroy()
                except Exception:
                    pass
                self._open_word_pdf_batch_dialog(parent_window=parent_window or owner)

            WidgetFactory.create_button(
                actions,
                text="🧾 Générateur de Documents",
                command=_open_generator,
                style="Manage.TButton",
            ).pack(fill="x", pady=(0, 8))

            WidgetFactory.create_button(
                actions,
                text="📄 Convertisseur Word -> PDF (lot)",
                command=_open_word_pdf_batch,
                style="Upload.TButton",
            ).pack(fill="x", pady=(0, 8))

            WidgetFactory.create_button(
                actions,
                text="⚙ Valeurs par défaut",
                command=_open_defaults,
                style="Success.TButton",
            ).pack(fill="x", pady=(0, 8))

            WidgetFactory.create_button(
                actions,
                text="📊 Analyse des valeurs templates",
                command=_open_analyzer,
                style="Refresh.TButton",
            ).pack(fill="x", pady=(0, 8))

            buttons = ttk.Frame(main_frame)
            buttons.pack(fill="x")
            WidgetFactory.create_button(
                buttons,
                text="❌ Fermer",
                command=top.destroy,
                style="Close.TButton",
            ).pack(side="right")

            WindowManager.center_window(top)
            try:
                top.lift()
                top.focus_force()
            except Exception:
                pass
        except Exception as e:
            logger.exception("Erreur lors de l'ouverture des outils")
            messagebox.showerror("Erreur", f"Impossible d'ouvrir les outils: {e}")

    def _open_word_pdf_batch_dialog(self, parent_window=None):
        """Open tool dialog to convert many DOCX files to PDF with report."""
        try:
            from ..utils.word_pdf_batch import convert_docx_batch
        except Exception as e:
            messagebox.showerror(
                "Outil indisponible",
                f"Impossible de charger le convertisseur Word -> PDF:\n{e}",
            )
            return

        owner = parent_window or self.winfo_toplevel()
        top = tk.Toplevel(owner)
        top.transient(owner)
        top.title("Outils - Conversion Word vers PDF")
        top.geometry("840x520")
        top.resizable(True, True)

        try:
            top.grab_set()
        except Exception:
            pass

        container = ttk.Frame(top, padding=14)
        container.pack(fill="both", expand=True)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(6, weight=1)

        ttk.Label(
            container,
            text="📄 Conversion Word -> PDF (lot)",
            style="SectionHeader.TLabel",
        ).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))

        source_var = tk.StringVar()
        recursive_var = tk.BooleanVar(value=True)
        status_var = tk.StringVar(value="Sélectionnez un dossier source puis lancez la conversion.")
        progress_var = tk.StringVar(value="0 / 0")
        progress_value = tk.DoubleVar(value=0.0)
        progress_percent_var = tk.StringVar(value="0%")
        selection_var = tk.StringVar(value="Sélection : 0 / 0")
        report_html = {"path": None}
        pending_files = {"paths": []}
        file_items = {"items": {}, "by_path": {}, "labels": {}, "total": 0}

        ttk.Label(container, text="Dossier source (.docx):").grid(row=1, column=0, sticky="w", padx=(0, 8))
        source_entry = ttk.Entry(container, textvariable=source_var)
        source_entry.grid(row=1, column=1, sticky="ew")

        def _browse_source():
            selected = filedialog.askdirectory(
                title="Sélectionner le dossier contenant les fichiers Word (.docx)"
            )
            if selected:
                source_var.set(selected)
                _scan_source_folder()

        browse_btn = WidgetFactory.create_button(
            container,
            text="📁 Parcourir",
            command=_browse_source,
            style="Secondary.TButton",
        )
        browse_btn.grid(row=1, column=2, sticky="e", padx=(8, 0))

        recursive_chk = ttk.Checkbutton(
            container,
            text="Inclure sous-dossiers (V1, activé)",
            variable=recursive_var,
        )
        recursive_chk.state(["selected", "disabled"])
        recursive_chk.grid(row=2, column=0, columnspan=3, sticky="w", pady=(10, 8))

        status_row = ttk.Frame(container)
        status_row.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 8))
        status_row.columnconfigure(0, weight=1)
        ttk.Label(status_row, textvariable=status_var).grid(row=0, column=0, sticky="w")
        ttk.Label(status_row, textvariable=progress_var).grid(row=0, column=1, sticky="e")

        progress_style = "WordPdf.Horizontal.TProgressbar"
        progress_label_style = "WordPdf.ProgressText.TLabel"
        try:
            colors = getattr(self.theme_manager, "colors", {})
            trough_color = colors.get("section_bg", "#2b2b2b")
            bar_color = colors.get("success", "#2ecc71")
            self.style.configure(
                progress_style,
                troughcolor=trough_color,
                background=bar_color,
                bordercolor=trough_color,
                lightcolor=bar_color,
                darkcolor=bar_color,
            )
            self.style.configure(
                progress_label_style,
                background=trough_color,
                foreground="#ffffff",
                font=("Segoe UI", 9, "bold"),
            )
        except Exception:
            pass

        progress_frame = ttk.Frame(container)
        progress_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(0, 8))
        progress_frame.columnconfigure(0, weight=1)
        progress = ttk.Progressbar(
            progress_frame,
            mode="determinate",
            variable=progress_value,
            maximum=1,
            style=progress_style,
        )
        progress.grid(row=0, column=0, sticky="ew")
        progress_label = ttk.Label(
            progress_frame,
            textvariable=progress_percent_var,
            style=progress_label_style,
        )
        progress_label.place(relx=0.5, rely=0.5, anchor="center")

        selection_row = ttk.Frame(container)
        selection_row.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 6))
        selection_row.columnconfigure(0, weight=1)

        ttk.Label(selection_row, textvariable=selection_var).grid(row=0, column=0, sticky="w")

        select_all_btn = WidgetFactory.create_button(
            selection_row,
            text="✅ Tout sélectionner",
            command=lambda: _select_all_files(),
            style="Secondary.TButton",
        )
        select_all_btn.grid(row=0, column=1, sticky="e", padx=(8, 4))

        clear_sel_btn = WidgetFactory.create_button(
            selection_row,
            text="❌ Désélectionner",
            command=lambda: _clear_file_selection(),
            style="Secondary.TButton",
        )
        clear_sel_btn.grid(row=0, column=2, sticky="e")

        files_frame = ttk.Frame(container)
        files_frame.grid(row=6, column=0, columnspan=3, sticky="nsew")
        files_frame.columnconfigure(0, weight=1)
        files_frame.rowconfigure(0, weight=1)

        files_tree = ttk.Treeview(
            files_frame,
            columns=("file", "progress", "status"),
            show="headings",
            selectmode="extended",
            height=12,
        )
        files_tree.heading("file", text="Fichier")
        files_tree.heading("progress", text="Progression")
        files_tree.heading("status", text="Statut")
        files_tree.column("file", width=520, anchor="w")
        files_tree.column("progress", width=140, anchor="center")
        files_tree.column("status", width=140, anchor="center")
        files_tree.grid(row=0, column=0, sticky="nsew")

        files_scroll = ttk.Scrollbar(files_frame, orient="vertical", command=files_tree.yview)
        files_scroll.grid(row=0, column=1, sticky="ns")
        files_tree.configure(yscrollcommand=files_scroll.set)

        tag_colors = getattr(self.theme_manager, "colors", {})
        files_tree.tag_configure("pending", foreground="#f2994a")
        files_tree.tag_configure("running", foreground="#f2994a")
        files_tree.tag_configure(
            "ok",
            foreground=tag_colors.get("success_row_fg", "#e7f7ed"),
            background=tag_colors.get("success_row_bg", "#1f5f3a"),
        )
        files_tree.tag_configure(
            "error",
            foreground=tag_colors.get("error_row_fg", "#f4d7d6"),
            background=tag_colors.get("error_row_bg", "#5a2a2a"),
        )

        buttons = ttk.Frame(container)
        buttons.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(10, 0))

        def _status_display(status: str) -> tuple[str, str]:
            status_key = (status or "pending").lower()
            if status_key == "ok":
                return "✅ OK", "ok"
            if status_key == "skipped":
                return "✅ OK", "ok"
            if status_key == "running":
                return "🟠 En cours", "running"
            if status_key == "error":
                return "❌ Non converti", "error"
            return "🟠 En attente", "pending"

        def _progress_bar(percent: int) -> str:
            blocks = 14
            percent = max(0, min(100, int(percent)))
            filled = int(round((percent / 100) * blocks))
            bar = "▰" * filled + "▱" * (blocks - filled)
            return f"{bar} {percent}%"

        def _progress_display(status: str) -> str:
            status_key = (status or "pending").lower()
            if status_key in {"ok", "skipped", "error"}:
                return _progress_bar(100)
            if status_key == "running":
                return _progress_bar(50)
            return _progress_bar(0)

        def _set_global_progress(processed: int, total: int):
            total_safe = max(total, 1)
            progress_var.set(f"{processed} / {total}")
            progress_value.set(processed)
            percent = int(round((processed / total_safe) * 100))
            progress_percent_var.set(f"{percent}%")

        def _update_selection_summary():
            total = file_items.get("total", 0)
            selected = len(files_tree.selection())
            selection_var.set(f"Sélection : {selected} / {total}")

        def _select_all_files():
            items = files_tree.get_children()
            files_tree.selection_set(items)
            _update_selection_summary()

        def _clear_file_selection():
            items = files_tree.get_children()
            files_tree.selection_remove(items)
            _update_selection_summary()

        def _toggle_tree_selection(event):
            try:
                region = files_tree.identify("region", event.x, event.y)
                if region in {"heading", "separator"}:
                    return None
                row_id = files_tree.identify_row(event.y)
                if not row_id:
                    return None
                if row_id in files_tree.selection():
                    files_tree.selection_remove(row_id)
                else:
                    files_tree.selection_add(row_id)
                    files_tree.focus(row_id)
                _update_selection_summary()
                return "break"
            except Exception:
                return None

        files_tree.bind("<<TreeviewSelect>>", lambda _e: _update_selection_summary())
        files_tree.bind("<Button-1>", _toggle_tree_selection, add=True)

        def _scan_source_folder():
            source_raw = source_var.get().strip()
            report_html["path"] = None
            open_report_btn.configure(state="disabled")
            pending_files["paths"] = []
            file_items["items"].clear()
            file_items["by_path"].clear()
            file_items["labels"].clear()
            file_items["total"] = 0

            if not source_raw:
                for item in files_tree.get_children():
                    files_tree.delete(item)
                status_var.set("Sélectionnez un dossier source puis lancez la conversion.")
                _set_global_progress(0, 0)
                progress_value.set(0)
                progress.configure(maximum=1)
                selection_var.set("Sélection : 0 / 0")
                launch_btn.configure(state="disabled")
                return

            source_dir = Path(source_raw).expanduser()
            if not source_dir.exists() or not source_dir.is_dir():
                for item in files_tree.get_children():
                    files_tree.delete(item)
                status_var.set("Dossier source invalide.")
                _set_global_progress(0, 0)
                progress_value.set(0)
                progress.configure(maximum=1)
                selection_var.set("Sélection : 0 / 0")
                launch_btn.configure(state="disabled")
                return

            paths = sorted(p for p in source_dir.rglob("*.docx") if p.is_file())
            pending_files["paths"] = paths
            total = len(paths)
            _set_global_progress(0, total)
            progress_value.set(0)
            progress.configure(maximum=max(total, 1))
            file_items["total"] = total
            for item in files_tree.get_children():
                files_tree.delete(item)

            if total == 0:
                status_var.set("Aucun document .docx trouvé dans ce dossier.")
                selection_var.set("Sélection : 0 / 0")
                launch_btn.configure(state="disabled")
                return

            status_var.set(f"{total} document(s) .docx détecté(s), prêt(s) à convertir.")
            for docx_path in paths:
                try:
                    rel = docx_path.relative_to(source_dir)
                    label = str(rel)
                except Exception:
                    label = str(docx_path)
                status_text, status_tag = _status_display("pending")
                progress_text = _progress_display("pending")
                item_id = files_tree.insert(
                    "",
                    "end",
                    values=(label, progress_text, status_text),
                    tags=(status_tag,),
                )
                resolved = str(docx_path.expanduser().resolve())
                file_items["items"][item_id] = docx_path
                file_items["by_path"][resolved] = item_id
                file_items["labels"][item_id] = label
            _select_all_files()
            launch_btn.configure(state="normal")

        def _open_report():
            raw_path = report_html.get("path")
            if not raw_path:
                return
            try:
                self._open_path_in_system(Path(raw_path))
            except Exception as e:
                messagebox.showerror("Ouverture rapport", f"Impossible d'ouvrir le rapport:\n{e}")

        open_report_btn = WidgetFactory.create_button(
            buttons,
            text="📊 Ouvrir le rapport",
            command=_open_report,
            style="View.TButton",
        )
        open_report_btn.pack(side="left")
        open_report_btn.configure(state="disabled")

        def _set_running(is_running: bool):
            run_state = "disabled" if is_running else "normal"
            try:
                source_entry.configure(state=run_state)
            except Exception:
                pass
            browse_btn.configure(state=run_state)
            launch_btn.configure(state=run_state)
            close_btn.configure(state=run_state)
            select_all_btn.configure(state=run_state)
            clear_sel_btn.configure(state=run_state)
            if is_running:
                open_report_btn.configure(state="disabled")
            else:
                report_path = report_html.get("path")
                if report_path and Path(report_path).exists():
                    open_report_btn.configure(state="normal")
                else:
                    open_report_btn.configure(state="disabled")

        def _safe_ui_call(fn):
            try:
                if top.winfo_exists():
                    fn()
            except Exception:
                pass

        def _safe_after(fn):
            try:
                if top.winfo_exists():
                    top.after(1, fn)
            except Exception:
                pass

        def _start_batch_conversion():
            source_raw = source_var.get().strip()
            if not source_raw:
                messagebox.showwarning(
                    "Dossier source",
                    "Veuillez sélectionner un dossier source.",
                    parent=top,
                )
                return

            source_dir = Path(source_raw).expanduser()
            if not source_dir.exists() or not source_dir.is_dir():
                messagebox.showerror(
                    "Dossier source",
                    f"Dossier invalide:\n{source_dir}",
                    parent=top,
                )
                return

            if not pending_files["paths"]:
                _scan_source_folder()
            if not pending_files["paths"]:
                messagebox.showwarning(
                    "Conversion Word -> PDF",
                    "Aucun fichier .docx à convertir dans le dossier sélectionné.",
                    parent=top,
                )
                return
            selected_items = files_tree.selection()
            if not selected_items:
                messagebox.showwarning(
                    "Conversion Word -> PDF",
                    "Veuillez sélectionner au moins un fichier à convertir.",
                    parent=top,
                )
                return

            report_html["path"] = None
            _set_global_progress(0, 0)
            status_var.set("Conversion en cours...")
            for item_id in selected_items:
                label = file_items["labels"].get(item_id, "")
                status_text, status_tag = _status_display("pending")
                progress_text = _progress_display("pending")
                files_tree.item(item_id, values=(label, progress_text, status_text), tags=(status_tag,))
            progress_value.set(0)
            _set_running(True)

            def _progress_callback(processed: int, total: int, filename: str, entry: dict):
                def _update_ui():
                    if not top.winfo_exists():
                        return
                    status_key = str(entry.get("status", "pending")).lower()
                    if status_key != "running":
                        _set_global_progress(processed, total)
                    source_key = entry.get("source_docx") or filename
                    try:
                        source_key = str(Path(source_key).expanduser().resolve())
                    except Exception:
                        source_key = str(source_key)
                    item_id = file_items["by_path"].get(source_key)
                    if item_id:
                        label = file_items["labels"].get(item_id, filename)
                        status_text, status_tag = _status_display(entry.get("status"))
                        progress_text = _progress_display(entry.get("status"))
                        files_tree.item(
                            item_id,
                            values=(label, progress_text, status_text),
                            tags=(status_tag,),
                        )
                        if status_key in {"ok", "skipped", "error"}:
                            try:
                                files_tree.selection_remove(item_id)
                            except Exception:
                                pass

                _safe_after(_update_ui)

            def _worker():
                try:
                    selected_paths = []
                    for item_id in selected_items:
                        path = file_items["items"].get(item_id)
                        if path is not None:
                            selected_paths.append(path)
                    total_selected = len(selected_paths)
                    progress_value.set(0)
                    progress.configure(maximum=max(total_selected, 1))
                    progress_percent_var.set("0%")
                    result = convert_docx_batch(
                        source_dir=source_dir,
                        recursive=True,
                        files=selected_paths,
                        progress_callback=_progress_callback,
                    )

                    def _done():
                        if not top.winfo_exists():
                            return
                        report_html["path"] = result.get("report_html")
                        _set_running(False)
                        status_var.set("Conversion terminée.")
                        total_files = result.get("total_files", 0)
                        _set_global_progress(total_files, total_files)
                        _update_selection_summary()
                        try:
                            files_tree.selection_remove(files_tree.selection())
                        except Exception:
                            pass
                        location_path = source_dir
                        try:
                            output_dirs = {
                                str(Path(item.get("out_pdf", "")).expanduser().resolve().parent)
                                for item in result.get("files", [])
                                if item.get("status") in {"ok", "skipped"} and item.get("out_pdf")
                            }
                            if len(output_dirs) == 1:
                                location_path = Path(next(iter(output_dirs)))
                        except Exception:
                            location_path = source_dir
                        summary = (
                            f"Conversion terminée.\n\n"
                            f"Total: {result.get('total_files', 0)}\n"
                            f"Succès: {result.get('success_count', 0)}\n"
                            f"Ignorés: {result.get('skipped_count', 0)}\n"
                            f"Erreurs: {result.get('error_count', 0)}\n\n"
                            f"Emplacement: {location_path}"
                        )
                        if result.get("global_error"):
                            summary += f"\n\nErreur globale:\n{result.get('global_error')}"
                        try:
                            top.lift()
                            top.focus_force()
                        except Exception:
                            pass
                        try:
                            messagebox.showinfo("Conversion Word -> PDF", summary, parent=top)
                        except Exception:
                            pass
                        try:
                            self._open_path_in_system(Path(location_path))
                        except Exception:
                            pass

                    _safe_after(_done)
                except Exception as e:
                    def _failed():
                        if not top.winfo_exists():
                            return
                        _set_running(False)
                        status_var.set("Échec de conversion.")
                        try:
                            top.lift()
                            top.focus_force()
                        except Exception:
                            pass
                        try:
                            messagebox.showerror(
                                "Conversion Word -> PDF",
                                f"Erreur:\n{e}",
                                parent=top,
                            )
                        except Exception:
                            pass

                    _safe_after(_failed)

            threading.Thread(target=_worker, daemon=True).start()

        launch_btn = WidgetFactory.create_button(
            buttons,
            text="▶ Lancer la conversion",
            command=_start_batch_conversion,
            style="Success.TButton",
        )
        launch_btn.pack(side="left", padx=(8, 0))
        launch_btn.configure(state="disabled")

        source_entry.bind("<Return>", lambda _e: _scan_source_folder())
        source_entry.bind("<FocusOut>", lambda _e: _scan_source_folder())

        close_btn = WidgetFactory.create_button(
            buttons,
            text="❌ Fermer",
            command=top.destroy,
            style="Close.TButton",
        )
        close_btn.pack(side="right")

        WindowManager.center_window(top)

    def _open_defaults_dialog(self, parent_window=None):
        """Open tools dialog to manage default values for the entire application."""
        try:
            from ..utils.defaults_manager import get_defaults_manager
            from ..utils import constants
            
            owner = parent_window or self.winfo_toplevel()
            top = tk.Toplevel(owner)
            top.transient(owner)
            top.title('Outils - Valeurs par défaut')
            top.geometry('960x560')
            top.resizable(True, True)
            
            # Apply theme to the dialog window
            theme_manager = self.theme_manager
            bg_color = theme_manager.colors.get('bg', '#2b2b2b')
            fg_color = theme_manager.colors.get('fg', '#ffffff')
            top.configure(bg=bg_color)
            
            # Make modal
            try:
                top.grab_set()
            except Exception:
                pass

            # Get defaults manager
            defaults_mgr = get_defaults_manager()
            current_defaults = defaults_mgr.get_all_defaults()

            # Main container with notebook (tabs)
            main_frame = ttk.Frame(top)
            main_frame.pack(fill='both', expand=True, padx=10, pady=10)

            # Title
            title_label = ttk.Label(main_frame, text='⚙ Gestion des valeurs par défaut', 
                                   font=('Segoe UI', 12, 'bold'))
            title_label.pack(anchor='w', pady=(0, 10))

            # Create notebook for tabs
            notebook = ttk.Notebook(main_frame)
            notebook.pack(fill='both', expand=True, pady=(0, 10))

            # Dictionary to hold all entry widgets for saving
            entry_vars = {}

            # Create a tab for each section (societe, associe, contrat)
            sections_config = {
                'societe': {
                    'label': '🏢 Entreprise',
                    'fields': [
                        ('DenSte', 'Dénomination sociale', constants.DenSte),
                        ('FormJur', 'Forme juridique', constants.Formjur),
                        ('ModeSignatureGerance', 'Mode signature gérance (SARL)', ['separee', 'conjointe']),
                        ('Ice', 'ICE', []),
                        ('DateIce', 'Date certificat négatif', []),
                        ('DateExpCertNeg', 'Date expiration certificat négatif', []),
                        ('Capital', 'Capital', constants.Capital),
                        ('PartsSocial', 'Parts sociales', constants.PartsSocial),
                        ('ValeurNominale', 'Valeur nominale', ['100']),
                        ('SteAdresse', 'Adresse', constants.SteAdresse),
                        ('Tribunal', 'Tribunal', constants.Tribunnaux),
                    ]
                },
                'associe': {
                    'label': '👤 Associé',
                    'fields': [
                        ('Civility', 'Civilité', constants.Civility),
                        ('Nom', 'Nom', []),
                        ('Prenom', 'Prénom', []),
                        ('Nationality', 'Nationalité', constants.Nationalite),
                        ('NumPiece', 'N° CIN', []),
                        ('Telephone', 'Téléphone', []),
                        ('Email', 'Email', []),
                        ('Adresse', 'Adresse', []),
                        ('Quality', 'Qualité', constants.QualityAssocie),
                    ]
                },
                'contrat': {
                    'label': '📋 Contrat',
                    'fields': [
                        ('NbMois', 'Période (mois)', constants.Nbmois),
                        ('TypeContratDomiciliation', 'Type contrat domiciliation', constants.TypeContratDomiciliation),
                        ('CollaborateurNom', 'Nom collaborateur', []),
                        ('TypeRenouvellement', 'Type renouvellement', constants.TypeRenouvellement),
                        ('Tva', 'TVA initiale (%)', ['20']),
                        ('DhHt', 'Loyer HT initial (DH)', ['83.3333']),
                        ('TvaRenouvellement', 'TVA renouvellement (%)', ['20']),
                        ('DhHtRenouvellement', 'Loyer HT renouvellement (DH)', ['166.667']),
                    ]
                },
                'collaborateur': {
                    'label': '🤝 Collaborateur',
                    'fields': [
                        ('Type', 'Type collaborateur', constants.Collaborateurs),
                        ('Code', 'Code collaborateur', []),
                        ('Nom', 'Nom / Raison sociale', []),
                        ('Ice', 'ICE', []),
                        ('Tp', 'TP', []),
                        ('Rc', 'RC', []),
                        ('If', 'IF', []),
                        ('TelFixe', 'Téléphone fixe', []),
                        ('TelMobile', 'Téléphone mobile', []),
                        ('Adresse', 'Adresse', []),
                        ('Email', 'Email', []),
                    ]
                }
            }

            def _is_full_width_field(field_key: str, field_label: str) -> bool:
                if str(field_key or '').strip() in {'SteAdresse', 'Adresse'}:
                    return True
                return 'adresse' in str(field_label or '').strip().lower()

            for section_key, section_info in sections_config.items():
                # Create frame for this tab
                tab_frame = ttk.Frame(notebook)
                notebook.add(tab_frame, text=section_info['label'])

                # Create scrollable area
                canvas = tk.Canvas(tab_frame, highlightthickness=0, bg=bg_color)
                scrollbar = ttk.Scrollbar(tab_frame, orient='vertical', command=canvas.yview)
                scrollable_frame = ttk.Frame(canvas)

                scrollable_frame.bind(
                    "<Configure>",
                    lambda _event, c=canvas: c.configure(scrollregion=c.bbox("all"))
                )

                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)

                canvas.pack(side='left', fill='both', expand=True)
                scrollbar.pack(side='right', fill='y')

                def _bind_canvas_mousewheel(target_canvas):
                    def _on_wheel(event):
                        try:
                            delta = int(getattr(event, "delta", 0))
                        except Exception:
                            delta = 0
                        if delta == 0:
                            return "break"
                        steps = max(1, abs(delta) // 120)
                        direction = -1 if delta > 0 else 1
                        target_canvas.yview_scroll(direction * steps, "units")
                        return "break"

                    # Linux compatibility
                    def _on_wheel_up(_event):
                        target_canvas.yview_scroll(-1, "units")
                        return "break"

                    def _on_wheel_down(_event):
                        target_canvas.yview_scroll(1, "units")
                        return "break"

                    target_canvas.bind("<MouseWheel>", _on_wheel)
                    target_canvas.bind("<Button-4>", _on_wheel_up)
                    target_canvas.bind("<Button-5>", _on_wheel_down)

                _bind_canvas_mousewheel(canvas)

                # Add fields for this section
                entry_vars[section_key] = {}
                fields_grid = ttk.Frame(scrollable_frame, padding=(8, 8, 8, 8))
                fields_grid.pack(fill='both', expand=True)
                fields_grid.columnconfigure(0, weight=1, uniform=f'{section_key}_defaults_cols')
                fields_grid.columnconfigure(1, weight=1, uniform=f'{section_key}_defaults_cols')
                fields_grid.columnconfigure(2, weight=1, uniform=f'{section_key}_defaults_cols')

                grid_row = 0
                grid_col = 0
                for field_key, field_label, field_options in section_info['fields']:
                    full_width = _is_full_width_field(field_key, field_label)
                    col_span = 3 if full_width else 1

                    if full_width and grid_col != 0:
                        grid_row += 1
                        grid_col = 0

                    if full_width:
                        cell_padx = (0, 0)
                    elif grid_col == 0:
                        cell_padx = (0, 6)
                    elif grid_col == 1:
                        cell_padx = (3, 3)
                    else:
                        cell_padx = (6, 0)

                    field_frame = ttk.Frame(fields_grid)
                    field_frame.grid(
                        row=grid_row,
                        column=0 if full_width else grid_col,
                        columnspan=col_span,
                        sticky='ew',
                        padx=cell_padx,
                        pady=5,
                    )
                    field_frame.columnconfigure(0, weight=1)

                    ttk.Label(field_frame, text=f'{field_label}:', anchor='w').grid(
                        row=0, column=0, sticky='w', pady=(0, 3)
                    )

                    current_value = current_defaults.get(section_key, {}).get(field_key, '')
                    var = tk.StringVar(value=str(current_value))
                    entry_vars[section_key][field_key] = var

                    if field_options and len(field_options) > 0:
                        combo = ttk.Combobox(field_frame, textvariable=var, values=field_options, state='readonly')
                        combo.grid(row=1, column=0, sticky='ew')
                    else:
                        entry = ttk.Entry(field_frame, textvariable=var)
                        entry.grid(row=1, column=0, sticky='ew')

                    if full_width:
                        grid_row += 1
                        grid_col = 0
                    elif grid_col == 0:
                        grid_col = 1
                    elif grid_col == 1:
                        grid_col = 2
                    else:
                        grid_col = 0
                        grid_row += 1

            # Button frame at bottom
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill='x', pady=(10, 0))

            def _save():
                try:
                    # Collect all values from tabs
                    new_defaults = {}
                    for section_key, fields in entry_vars.items():
                        new_defaults[section_key] = {}
                        for field_key, var in fields.items():
                            new_defaults[section_key][field_key] = var.get()
                    
                    # Save to defaults manager
                    defaults_mgr.set_all_defaults(new_defaults)
                    
                    try:
                        self._on_defaults_changed()
                    except Exception:
                        pass
                    if hasattr(self.parent, '_on_defaults_changed'):
                        self.parent._on_defaults_changed()
                    
                        messagebox.showinfo('Outils', 'Valeurs par défaut enregistrées avec succès.')
                    try:
                        top.destroy()
                    except Exception:
                        pass
                except Exception as e:
                    messagebox.showerror('Erreur', f"Impossible d'enregistrer: {e}")

            def _reset():
                """Reset to initial defaults."""
                if messagebox.askyesno('Confirmation', 'Réinitialiser tous les défauts?\n\nCette action ne peut pas être annulée.'):
                    try:
                        defaults_mgr.reset_to_initial()
                        # Reload values
                        current_defaults = defaults_mgr.get_all_defaults()
                        for section_key, fields in entry_vars.items():
                            for field_key, var in fields.items():
                                new_val = current_defaults.get(section_key, {}).get(field_key, '')
                                var.set(str(new_val))
                        
                        try:
                            self._on_defaults_changed()
                        except Exception:
                            pass
                        if hasattr(self.parent, '_on_defaults_changed'):
                            self.parent._on_defaults_changed()
                        
                        messagebox.showinfo('Outils', 'Défauts réinitialisés aux valeurs initiales.')
                    except Exception as e:
                        messagebox.showerror('Erreur', f"Impossible de réinitialiser: {e}")

            # Buttons
            save_btn = WidgetFactory.create_button(button_frame, text='💾 Enregistrer', 
                                                 command=_save, style='Success.TButton')
            save_btn.pack(side='right', padx=4)

            reset_btn = WidgetFactory.create_button(button_frame, text='🔄 Réinitialiser', 
                                                   command=_reset, style='Secondary.TButton')
            reset_btn.pack(side='right', padx=4)

            close_btn = WidgetFactory.create_button(button_frame, text='❌ Fermer', 
                                                  command=lambda: top.destroy(), style='Close.TButton')
            close_btn.pack(side='right', padx=4)

            # Center the dialog
            try:
                from ..utils.utils import WindowManager
                WindowManager.center_window(top)
            except Exception:
                pass
        except Exception as e:
            logger.exception("Erreur lors de l'ouverture des outils")
            messagebox.showerror("Erreur", f"Impossible d'ouvrir les outils: {e}")

    def _open_template_values_analyzer_dialog(self, parent_window=None):
        """Open template variable analyzer with global and detailed views."""
        try:
            from ..utils.doc_generator import get_expected_context_keys
            from ..utils.template_value_analyzer import (
                analyze_templates,
                export_analysis_rows,
                filter_analysis_rows,
            )

            owner = parent_window or self.winfo_toplevel()
            top = tk.Toplevel(owner)
            top.transient(owner)
            top.title("Outils - Analyse des valeurs templates")
            top.geometry("1200x760")
            top.resizable(True, True)

            try:
                top.grab_set()
            except Exception:
                pass

            main_frame = ttk.Frame(top, padding=10)
            main_frame.pack(fill="both", expand=True)
            main_frame.columnconfigure(0, weight=1)
            main_frame.rowconfigure(3, weight=1)

            ttk.Label(
                main_frame,
                text="📊 Analyse des valeurs templates",
                font=("Segoe UI", 12, "bold"),
            ).grid(row=0, column=0, sticky="w", pady=(0, 8))

            filter_frame = ttk.Frame(main_frame)
            filter_frame.grid(row=1, column=0, sticky="ew", pady=(0, 8))
            for idx in range(8):
                filter_frame.columnconfigure(idx, weight=1 if idx % 2 else 0)
            filter_frame.columnconfigure(8, weight=0)

            search_var = tk.StringVar()
            template_var = tk.StringVar(value="Tous")
            section_var = tk.StringVar(value="Tous")
            coverage_var = tk.StringVar(value="Tous")
            status_var = tk.StringVar(value="")
            errors_var = tk.StringVar(value="")

            ttk.Label(filter_frame, text="Recherche:").grid(row=0, column=0, sticky="w", padx=(0, 6))
            search_entry = ttk.Entry(filter_frame, textvariable=search_var)
            search_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))

            ttk.Label(filter_frame, text="Template:").grid(row=0, column=2, sticky="w", padx=(0, 6))
            template_combo = ttk.Combobox(filter_frame, textvariable=template_var, state="readonly", values=["Tous"])
            template_combo.grid(row=0, column=3, sticky="ew", padx=(0, 10))

            ttk.Label(filter_frame, text="Section:").grid(row=0, column=4, sticky="w", padx=(0, 6))
            section_combo = ttk.Combobox(
                filter_frame,
                textvariable=section_var,
                state="readonly",
                values=["Tous", "societe", "associe", "contrat", "autre"],
            )
            section_combo.grid(row=0, column=5, sticky="ew", padx=(0, 10))

            ttk.Label(filter_frame, text="Couverture:").grid(row=0, column=6, sticky="w", padx=(0, 6))
            coverage_combo = ttk.Combobox(
                filter_frame,
                textvariable=coverage_var,
                state="readonly",
                values=["Tous", "couvert", "non couvert"],
            )
            coverage_combo.grid(row=0, column=7, sticky="ew", padx=(0, 10))

            actions = ttk.Frame(filter_frame)
            actions.grid(row=0, column=8, sticky="e")

            notebook = ttk.Notebook(main_frame)
            notebook.grid(row=3, column=0, sticky="nsew", pady=(0, 8))
            global_tab = ttk.Frame(notebook)
            detail_tab = ttk.Frame(notebook)
            notebook.add(global_tab, text="Vue globale")
            notebook.add(detail_tab, text="Vue détaillée")

            for frame in (global_tab, detail_tab):
                frame.columnconfigure(0, weight=1)
                frame.rowconfigure(0, weight=1)

            global_cols = ("variable", "occurrences", "templates_count", "section", "coverage")
            detail_cols = ("template", "variable", "occurrences", "section", "coverage")

            def _create_tree(parent, columns, headings):
                container = ttk.Frame(parent)
                container.grid(row=0, column=0, sticky="nsew")
                container.columnconfigure(0, weight=1)
                container.rowconfigure(0, weight=1)

                tree = ttk.Treeview(container, columns=columns, show="headings")
                y_scroll = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
                x_scroll = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
                tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

                tree.grid(row=0, column=0, sticky="nsew")
                y_scroll.grid(row=0, column=1, sticky="ns")
                x_scroll.grid(row=1, column=0, sticky="ew")

                sort_state = {col: False for col in columns}

                def _sort_by(col_name):
                    reverse = sort_state.get(col_name, False)
                    sort_state[col_name] = not reverse
                    rows = [(tree.set(item_id, col_name), item_id) for item_id in tree.get_children("")]

                    def _coerce(value):
                        txt = str(value).replace(" ", "").replace(",", ".")
                        try:
                            return float(txt)
                        except Exception:
                            return str(value).lower()

                    rows.sort(key=lambda item: _coerce(item[0]), reverse=reverse)
                    for idx, (_, item_id) in enumerate(rows):
                        tree.move(item_id, "", idx)

                for col_name, heading in zip(columns, headings):
                    tree.heading(col_name, text=heading, command=lambda c=col_name: _sort_by(c))
                    tree.column(
                        col_name,
                        width=220 if col_name in ("variable", "template") else 130,
                        anchor="center",
                    )
                return tree

            global_tree = _create_tree(
                global_tab,
                global_cols,
                ("Variable", "Occurrences", "Nb templates", "Section", "Couverture"),
            )
            detail_tree = _create_tree(
                detail_tab,
                detail_cols,
                ("Template", "Variable", "Occurrences", "Section", "Couverture"),
            )

            kpi_frame = ttk.Frame(main_frame)
            kpi_frame.grid(row=2, column=0, sticky="ew", pady=(0, 8))
            for idx in range(5):
                kpi_frame.columnconfigure(idx, weight=1)

            total_templates_var = tk.StringVar(value="Templates: 0")
            total_occurrences_var = tk.StringVar(value="Occurrences: 0")
            total_distinct_var = tk.StringVar(value="Variables distinctes: 0")
            total_covered_var = tk.StringVar(value="Couvertes: 0")
            total_uncovered_var = tk.StringVar(value="Non couvertes: 0")

            ttk.Label(kpi_frame, textvariable=total_templates_var).grid(row=0, column=0, sticky="w")
            ttk.Label(kpi_frame, textvariable=total_occurrences_var).grid(row=0, column=1, sticky="w")
            ttk.Label(kpi_frame, textvariable=total_distinct_var).grid(row=0, column=2, sticky="w")
            ttk.Label(kpi_frame, textvariable=total_covered_var).grid(row=0, column=3, sticky="w")
            ttk.Label(kpi_frame, textvariable=total_uncovered_var).grid(row=0, column=4, sticky="w")

            footer = ttk.Frame(main_frame)
            footer.grid(row=4, column=0, sticky="ew")
            footer.columnconfigure(0, weight=1)
            ttk.Label(footer, textvariable=status_var).grid(row=0, column=0, sticky="w")
            ttk.Label(footer, textvariable=errors_var, foreground="#c5865d").grid(row=1, column=0, sticky="w")

            analysis_data = {"variables": [], "details": [], "summary": {}, "templates": [], "errors": []}
            current_global_rows = []
            current_detail_rows = []
            detail_item_map = {}

            def _populate_tree(tree, rows, columns, item_map=None):
                for item_id in tree.get_children(""):
                    tree.delete(item_id)
                if item_map is not None:
                    item_map.clear()
                for row in rows:
                    item_id = tree.insert("", "end", values=[row.get(col, "") for col in columns])
                    if item_map is not None:
                        item_map[item_id] = row

            def _apply_filters(_event=None):
                nonlocal current_global_rows, current_detail_rows
                global_rows = filter_analysis_rows(
                    analysis_data.get("variables", []),
                    search_text=search_var.get(),
                    template_name=template_var.get(),
                    section=section_var.get(),
                    coverage=coverage_var.get(),
                )
                detail_rows = filter_analysis_rows(
                    analysis_data.get("details", []),
                    search_text=search_var.get(),
                    template_name=template_var.get(),
                    section=section_var.get(),
                    coverage=coverage_var.get(),
                )
                current_global_rows = global_rows
                current_detail_rows = detail_rows
                _populate_tree(global_tree, current_global_rows, global_cols)
                _populate_tree(detail_tree, current_detail_rows, detail_cols, item_map=detail_item_map)
                status_var.set(
                    f"Global: {len(global_rows)} / {len(analysis_data.get('variables', []))} | "
                    f"Détail: {len(detail_rows)} / {len(analysis_data.get('details', []))}"
                )

            def _export_current_view(default_ext: str):
                if notebook.index(notebook.select()) == 0:
                    rows = current_global_rows
                    title = "vue_globale"
                    export_columns = (
                        "variable",
                        "occurrences",
                        "templates_count",
                        "section",
                        "coverage",
                        "templates",
                    )
                else:
                    rows = current_detail_rows
                    title = "vue_detaillee"
                    export_columns = (
                        "template",
                        "template_path",
                        "variable",
                        "occurrences",
                        "section",
                        "coverage",
                    )

                if not rows:
                    messagebox.showwarning("Export", "Aucune ligne à exporter pour la vue courante.")
                    return

                filename = filedialog.asksaveasfilename(
                    title="Exporter l'analyse",
                    defaultextension=default_ext,
                    filetypes=[
                        ("Fichier Excel", "*.xlsx"),
                        ("Fichier CSV", "*.csv"),
                    ],
                    initialfile=f"analyse_templates_{title}{default_ext}",
                )
                if not filename:
                    return
                try:
                    export_analysis_rows(rows, filename, columns=export_columns)
                    messagebox.showinfo("Export", f"Export terminé:\n{filename}")
                except Exception as exc:
                    logger.exception("Erreur lors de l'export de l'analyse templates")
                    messagebox.showerror("Erreur", f"Impossible d'exporter: {exc}")

            def _open_selected_template():
                selected = detail_tree.selection()
                if not selected:
                    messagebox.showwarning("Ouvrir template", "Sélectionnez d'abord une ligne dans la vue détaillée.")
                    return
                row = detail_item_map.get(selected[0]) or {}
                template_path_value = row.get("template_path")
                template_name = row.get("template")

                candidate = Path(template_path_value) if template_path_value else None
                if not candidate or not candidate.exists():
                    if template_name:
                        matches = sorted(PathManager.MODELS_DIR.rglob(str(template_name)))
                        if not matches:
                            messagebox.showerror(
                                "Ouvrir template",
                                f"Template introuvable pour '{template_name}'.",
                            )
                            return
                        candidate = matches[0]

                try:
                    self._open_path_in_system(candidate)
                except Exception as exc:
                    logger.exception("Erreur lors de l'ouverture du template")
                    messagebox.showerror("Erreur", f"Impossible d'ouvrir le template: {exc}")

            def _refresh_analysis():
                try:
                    data = analyze_templates(
                        PathManager.MODELS_DIR,
                        context_keys=get_expected_context_keys(),
                    )
                    analysis_data.update(data)

                    summary = data.get("summary", {})
                    total_templates_var.set(f"Templates: {summary.get('total_templates', 0)}")
                    total_occurrences_var.set(f"Occurrences: {summary.get('total_variable_occurrences', 0)}")
                    total_distinct_var.set(f"Variables distinctes: {summary.get('total_distinct_variables', 0)}")
                    total_covered_var.set(f"Couvertes: {summary.get('covered_variables', 0)}")
                    total_uncovered_var.set(f"Non couvertes: {summary.get('uncovered_variables', 0)}")

                    template_values = ["Tous"] + sorted(set(data.get("templates", [])))
                    template_combo.configure(values=template_values)
                    if template_var.get() not in template_values:
                        template_var.set("Tous")

                    if section_var.get() not in ("Tous", "societe", "associe", "contrat", "autre"):
                        section_var.set("Tous")
                    if coverage_var.get() not in ("Tous", "couvert", "non couvert"):
                        coverage_var.set("Tous")

                    error_count = len(data.get("errors", []))
                    if error_count:
                        errors_var.set(f"⚠ {error_count} template(s) non analysé(s). Voir logs.")
                    else:
                        errors_var.set("")

                    _apply_filters()
                except Exception as exc:
                    logger.exception("Erreur lors de l'actualisation de l'analyse templates")
                    messagebox.showerror("Erreur", f"Impossible d'analyser les templates: {exc}")

            WidgetFactory.create_button(
                actions,
                text="🔄 Actualiser",
                command=_refresh_analysis,
                style="Refresh.TButton",
            ).pack(side="left", padx=(0, 6))
            WidgetFactory.create_button(
                actions,
                text="📤 Export CSV",
                command=lambda: _export_current_view(".csv"),
                style="Secondary.TButton",
            ).pack(side="left", padx=(0, 6))
            WidgetFactory.create_button(
                actions,
                text="📗 Export Excel",
                command=lambda: _export_current_view(".xlsx"),
                style="Secondary.TButton",
            ).pack(side="left", padx=(0, 6))
            WidgetFactory.create_button(
                actions,
                text="📂 Ouvrir template",
                command=_open_selected_template,
                style="View.TButton",
            ).pack(side="left", padx=(0, 6))
            WidgetFactory.create_button(
                actions,
                text="❌ Fermer",
                command=top.destroy,
                style="Close.TButton",
            ).pack(side="left")

            search_entry.bind("<KeyRelease>", _apply_filters)
            template_var.trace_add("write", lambda *_args: _apply_filters())
            section_var.trace_add("write", lambda *_args: _apply_filters())
            coverage_var.trace_add("write", lambda *_args: _apply_filters())

            _refresh_analysis()
            WindowManager.center_window(top)
        except Exception as e:
            logger.exception("Erreur lors de l'ouverture de l'analyse des valeurs templates")
            messagebox.showerror("Erreur", f"Impossible d'ouvrir l'analyse: {e}")

    @staticmethod
    def _open_path_in_system(path: Path):
        """Open a file/folder with the system default application."""
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"Chemin introuvable: {p}")

        if os.name == "nt":
            os.startfile(str(p))  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.run(["open", str(p)], check=False)
            return
        subprocess.run(["xdg-open", str(p)], check=False)

    def _on_defaults_changed(self):
        """Recharger les formulaires quand les défauts changent."""
        try:
            # Recharger les valeurs par défaut dans les formulaires existants
            if hasattr(self, 'societe_form') and self.societe_form:
                self.societe_form.initialize_variables()
            
            if hasattr(self, 'associe_form') and self.associe_form:
                # Pour AssocieForm, on réinitialise les vars pour les nouveaux associés
                # Les associés existants gardent leurs valeurs
                pass
            
            if hasattr(self, 'contrat_form') and self.contrat_form:
                self.contrat_form.initialize_variables()
            
            logger.info("Formulaires rechargés avec les nouvelles valeurs par défaut")
        except Exception as e:
            logger.exception('Erreur lors du rechargement des formulaires après changement de défauts')

    def get_values(self):
        """Get values from all forms"""
        # Ensure latest current page is saved before returning
        try:
            # call get_values on each form present
            values = {}
            for key, _, form in self.pages:
                if hasattr(form, 'get_values'):
                    values[key] = form.get_values()
            societe_vals = values.get('societe', {}) if isinstance(values, dict) else {}
            contrat_vals = values.get('contrat', {}) if isinstance(values, dict) else {}
            legal_form = ''
            if isinstance(societe_vals, dict):
                legal_form = str(societe_vals.get('forme_juridique', '') or '').strip()
            if legal_form and isinstance(contrat_vals, dict):
                contrat_vals['contrat_forme_juridique'] = legal_form
            self.values = values
        except Exception:
            # fallback to existing structure if any form is missing
            self.values = {
                'societe': getattr(self, 'societe_form', None) and self.societe_form.get_values(),
                'associes': getattr(self, 'associe_form', None) and self.associe_form.get_values(),
                'contrat': getattr(self, 'contrat_form', None) and self.contrat_form.get_values(),
            }
        return self.values

    def set_values(self, values):
        """Set values for all forms"""
        self.values = values
        if values:
            for key, _, form in self.pages:
                if key in values and hasattr(form, 'set_values'):
                    form.set_values(values[key])
            self._sync_legal_form_dependents()

    def reset(self):
        """Reset all forms to their default state"""
        self.values = {}
        # Réinitialiser chaque formulaire individuellement
        for key, _, form in self.pages:
            if hasattr(form, 'set_values'):
                # default empty structure
                default = [] if key == 'associes' else {}
                form.set_values(default)
        # Garantir au moins un associé visible après "Nouvelle".
        try:
            if hasattr(self, 'associe_form') and self.associe_form is not None:
                if len(getattr(self.associe_form, 'associe_vars', [])) == 0:
                    self.associe_form.add_associe()
        except Exception:
            pass
        try:
            self._sync_legal_form_dependents()
        except Exception:
            pass

    # --- Navigation helpers ---
    def show_page(self, index: int):
        """Show the page at `index` and hide others."""
        if index < 0 or index >= len(self.pages):
            return
        self.current_page = index
        for i, (_key, frame, _form) in enumerate(self.pages):
            if i == index:
                frame.tkraise()
                frame.grid()
            else:
                frame.grid_remove()
        # After showing the requested page, allow the page/form to perform
        # any 'on-show' updates. This is useful for forms that need to
        # recompute derived fields (for example, the Contrat form that
        # recalculates 'Date Fin' from 'Date Début' + 'Période'). We look for
        # a public hook `recalculate_date_fin` first, then fall back to the
        # private `_update_date_fin` to preserve backward compatibility.
        try:
            key, frame, form = self.pages[index]
            hook = getattr(form, 'recalculate_date_fin', None) or getattr(form, '_update_date_fin', None)
            if callable(hook):
                try:
                    hook()
                except Exception:
                    # Non-fatal: ignore form-level errors to avoid breaking navigation
                    pass
        except Exception:
            # Conservative: ignore any problems with calling the hook
            pass

        self.update_nav_buttons()

    def next_page(self):
        if self.current_page < len(self.pages) - 1:
            self.show_page(self.current_page + 1)

    def prev_page(self):
        if self.current_page > 0:
            self.show_page(self.current_page - 1)

    def save_current(self):
        """Save current page values into `self.values`."""
        key, _frame, form = self.pages[self.current_page]
        if hasattr(form, 'get_values'):
            try:
                vals = form.get_values()

                if key == 'associes' and hasattr(form, 'validate_for_submit'):
                    valid, _errors = form.validate_for_submit(show_dialog=True)
                    if not valid:
                        return

                # If we're saving the societe page, check for existing company name and forbid duplicates
                if key == 'societe':
                    try:
                        from ..utils.utils import societe_exists
                        name = vals.get('denomination') or vals.get('den_ste') or vals.get('DEN_STE')
                        editing = getattr(self, '_dashboard_edit_context', None)
                        if name and societe_exists(name) and not editing:
                            messagebox.showerror('Société existante', f"La société '{name}' existe déjà dans la base. Enregistrement interdit pour éviter les doublons.")
                            return
                    except Exception:
                        # If the check fails, log but allow save to proceed (conservative)
                        logger = __import__('logging').getLogger(__name__)
                        logger.exception('Failed to run societe_exists check')

                self.values[key] = vals
                try:
                    if getattr(self, '_dashboard_edit_context', None):
                        self._apply_dashboard_edit_for_page(key, vals)
                except Exception:
                    logger = __import__('logging').getLogger(__name__)
                    logger.exception('Failed to apply dashboard edit update')
                messagebox.showinfo("Sauvegarde", f"Section '{key}' sauvegardée.")
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de sauvegarder la section: {e}")

    def _trigger_generate_documents(self):
        """Run the exact same generation flow as the main 'Generer' button."""
        try:
            top = self.winfo_toplevel()
            generate_fn = getattr(top, 'generate_documents', None)
            if callable(generate_fn):
                return generate_fn()
        except Exception:
            logger.exception("Echec du declenchement de la generation depuis Terminer")
        # Fallback to legacy finish behavior if top-level hook is unavailable.
        return self.finish()

    def finish(self):
        """Save all pages and emit a finished event."""
        # Gather all values from forms (do this silently to avoid per-section popups)
        try:
            # get_values calls each form.get_values() and updates self.values
            all_values = self.get_values()
        except Exception:
            # Fallback: attempt to collect each form's values without showing dialogs
            try:
                values = {}
                for key, _, form in self.pages:
                    try:
                        values[key] = form.get_values() if hasattr(form, 'get_values') else {}
                    except Exception:
                        values[key] = {}
                self.values = values
                all_values = values
            except Exception:
                all_values = self.values
        # Ensure the Excel database exists and has expected sheets
        # Build the DB path explicitly to avoid static-analysis warnings about
        # PathManager.get_database_path. Use the centralized DB filename from
        # constants so the path is consistent.
        try:
            db_path = Path(PathManager.DATABASE_DIR) / _const.DB_FILENAME
        except Exception:
            db_path = Path.cwd() / _const.DB_FILENAME

        try:
            # Create the Excel DB using centralized headers from constants; ensure idempotent creation
            ensure_excel_db(db_path, _const.excel_sheets)
            # Initialize reference sheets with default data from constants
            from ..utils.utils import initialize_reference_sheets
            initialize_reference_sheets(db_path)
        except Exception as e:
            # non-fatal: log and show an error to the user
            try:
                from ..utils.utils import ErrorHandler
                ErrorHandler.handle_error(e, 'Erreur lors de la création de la base de données', show_dialog=True)
            except Exception:
                messagebox.showerror('Erreur', f"Impossible de créer la base de données: {e}")
            return

        # Now attempt to persist the collected values to the workbook by calling
        # the top-level application's save_to_db. This will collect values from
        # the forms and write them to the Excel file.
        try:
            top = self.winfo_toplevel()
            save_fn = getattr(top, 'save_to_db', None)
            saved_db = None
            if callable(save_fn):
                # save_to_db now returns the DB path on success (or None on failure)
                try:
                    saved_db = save_fn()
                except Exception:
                    # If the top-level save raised, surface a user-friendly message
                    saved_db = None
        except PermissionError as pe:
            # Common on Windows when the file is open in Excel
            messagebox.showerror('Erreur lors de la sauvegarde des données', 'Le fichier Excel est ouvert dans une autre application. Fermez Excel et réessayez.')
            return
        except Exception as e:
            try:
                from ..utils.utils import ErrorHandler
                ErrorHandler.handle_error(e, 'Erreur lors de la sauvegarde des données', show_dialog=True)
            except Exception:
                messagebox.showerror('Erreur', f"Erreur lors de la sauvegarde: {e}")
            return

        # If save was aborted or failed (saved_db is None), do not show success message
        if saved_db is None:
            # save_to_db already reports errors to the user where appropriate
            return

        # Present a single, clear success message (include the DB path when available)
        try:
            messagebox.showinfo("Sauvegarde réussie", f"Toutes les sections ont été sauvegardées dans le fichier Excel :\n{saved_db}")
        except Exception:
            # If messagebox fails for any reason, log and continue
            try:
                logger = __import__('logging').getLogger(__name__)
                logger.info('Sauvegarde terminée (message box failed to show)')
            except Exception:
                pass

        # Emit a virtual event so outer code can handle finalization if needed
        try:
            self.event_generate('<<FormsFinished>>')
        except Exception:
            pass

    def update_nav_buttons(self):
        # Disable Prev on first page.
        # Use explicit None checks so static analyzers (Pylance) know the
        # attribute is not None before calling widget methods.
        _prev = getattr(self, 'prev_btn', None)
        if _prev is not None:
            _prev.configure(state=('disabled' if self.current_page == 0 else 'normal'))

        _next = getattr(self, 'next_btn', None)
        _finish = getattr(self, 'finish_btn', None)

        is_last_page = self.current_page == len(self.pages) - 1

        # Keep the dedicated finish button hidden; we reuse "Suivant" as the
        # final action to preserve order and avoid layout shifts.
        if _finish is not None:
            try:
                _finish.pack_forget()
            except Exception:
                pass

        # Replace "Suivant" by "Terminer" on the last step, and restore when
        # navigating back.
        if _next is not None:
            if is_last_page:
                _next.configure(
                    text='Terminer',
                    style=getattr(self, 'next_finish_style', 'Success.TButton'),
                    command=self._trigger_generate_documents,
                    width=getattr(self, 'next_finish_width', 20),
                    state='normal',
                )
            else:
                _next.configure(
                    text='Suivant ▶',
                    style=getattr(self, 'next_default_style', 'Secondary.TButton'),
                    command=self.next_page,
                    width=getattr(self, 'next_default_width', 12),
                    state='normal',
                )

    @staticmethod
    def _infer_dashboard_page_key(payload: dict | None, page_key: str | None) -> str:
        """Infer dashboard page when older callers do not pass it explicitly."""
        if page_key in {'societe', 'associe', 'contrat', 'collaborateur'}:
            return page_key
        if not isinstance(payload, dict):
            return 'societe'
        if 'id_associe' in payload:
            return 'associe'
        if 'id_contrat' in payload:
            return 'contrat'
        if 'id_collaborateur' in payload:
            return 'collaborateur'
        return 'societe'

    @staticmethod
    def _dashboard_page_index(page_key: str) -> int:
        return {'societe': 0, 'associe': 1, 'contrat': 2}.get(page_key, 0)

    @staticmethod
    def _filter_dashboard_rows(df, sid: str, den: str):
        """Filter a worksheet DataFrame to rows linked to the selected company."""
        try:
            if sid and not df.empty and 'id_societe' in df.columns:
                return df[df['id_societe'].astype(str).str.strip() == sid]
            if den and not df.empty and 'den_ste' in df.columns:
                return df[df['den_ste'].astype(str).str.strip().str.lower() == den.lower()]
        except Exception:
            pass
        try:
            return df.iloc[0:0]
        except Exception:
            return df

    def _load_dashboard_workbook_frames(self):
        """Load dashboard workbook sheets once for edit/delete operations."""
        import pandas as _pd

        db_path = Path(PathManager.DATABASE_DIR) / _const.DB_FILENAME
        if not db_path.exists():
            raise FileNotFoundError('Fichier de base de données introuvable.')

        def _apply_aliases(df: _pd.DataFrame, alias_map: dict) -> _pd.DataFrame:
            if df is None or df.empty or not alias_map:
                return df
            out = df.copy()
            for old_col, new_col in alias_map.items():
                if old_col not in out.columns:
                    continue
                if new_col not in out.columns:
                    out[new_col] = out[old_col]
                else:
                    try:
                        old_vals = out[old_col].fillna('').astype(str).str.strip()
                        new_vals = out[new_col].fillna('').astype(str).str.strip()
                        mask = (new_vals == '') & (old_vals != '')
                        out.loc[mask, new_col] = out.loc[mask, old_col]
                    except Exception:
                        pass
                try:
                    out.drop(columns=[old_col], inplace=True)
                except Exception:
                    pass
            return out

        try:
            soc_df = _pd.read_excel(db_path, sheet_name='Societes', dtype=str).fillna('')
        except Exception:
            soc_df = _pd.DataFrame(columns=_const.societe_headers)
        soc_df = _apply_aliases(soc_df, getattr(_const, 'societe_header_aliases', {}) or {})

        try:
            assoc_df = _pd.read_excel(db_path, sheet_name='Associes', dtype=str).fillna('')
        except Exception:
            assoc_df = _pd.DataFrame(columns=_const.associe_headers)
        assoc_df = _apply_aliases(assoc_df, getattr(_const, 'associe_header_aliases', {}) or {})

        try:
            contrat_df = _pd.read_excel(db_path, sheet_name='Contrats', dtype=str).fillna('')
        except Exception:
            contrat_df = _pd.DataFrame(columns=_const.contrat_headers)
        contrat_df = _apply_aliases(contrat_df, getattr(_const, 'contrat_header_aliases', {}) or {})

        try:
            collab_df = _pd.read_excel(db_path, sheet_name='Collaborateurs', dtype=str).fillna('')
        except Exception:
            collab_df = _pd.DataFrame(columns=getattr(_const, 'collaborateur_headers', []))
        collab_df = _apply_aliases(collab_df, getattr(_const, 'collaborateur_header_aliases', {}) or {})

        return db_path, soc_df, assoc_df, contrat_df, collab_df

    def _resolve_company_payload_from_dashboard(self, page_key: str, payload: dict | None, soc_df):
        """Resolve the selected dashboard row to its company row."""
        if not isinstance(payload, dict):
            return None

        sid = str(payload.get('id_societe') or '').strip()
        den = str(payload.get('den_ste') or '').strip()

        if page_key == 'societe':
            if sid and not soc_df.empty and 'id_societe' in soc_df.columns:
                matches = soc_df[soc_df['id_societe'].astype(str).str.strip() == sid]
                if not matches.empty:
                    return matches.iloc[0].to_dict()
            if den and not soc_df.empty and 'den_ste' in soc_df.columns:
                matches = soc_df[soc_df['den_ste'].astype(str).str.strip().str.lower() == den.lower()]
                if not matches.empty:
                    return matches.iloc[0].to_dict()
            return payload

        matches = self._filter_dashboard_rows(soc_df, sid, den)
        if matches is None or matches.empty:
            return None
        return matches.iloc[0].to_dict()

    def _build_dashboard_edit_values(self, company_payload: dict, assoc_df, contrat_df, collab_df):
        """Build complete form values from the selected company and linked sheets."""
        soc_map = {
            'den_ste': 'denomination',
            'forme_jur': 'forme_juridique',
            'ice': 'ice',
            'date_ice': 'date_ice',
            'date_exp_cert_neg': 'date_expiration_certificat_negatif',
            'capital': 'capital',
            'part_social': 'parts_social',
            'valeur_nominale': 'valeur_nominale',
            'ste_adress': 'adresse',
            'tribunal': 'tribunal',
            'type_generation': 'type_generation',
            'procedure_creation': 'procedure_creation',
            'mode_depot_creation': 'mode_depot_creation',
        }
        inverse_assoc_map = {
            'civil': 'civilite',
            'prenom': 'prenom',
            'nom': 'nom',
            'parts': 'num_parts',
            'date_naiss': 'date_naiss',
            'lieu_naiss': 'lieu_naiss',
            'nationality': 'nationalite',
            'cin_num': 'num_piece',
            'cin_validaty': 'validite_piece',
            'adresse': 'adresse',
            'phone': 'telephone',
            'email': 'email',
            'is_gerant': 'est_gerant',
            'quality': 'qualite',
            'capital_detenu': 'capital_detenu',
            'part_percent': 'percentage',
        }
        inverse_contrat_map = {
            'date_contrat': 'date_contrat',
            'duree_contrat_mois': 'period',
            'type_contrat_domiciliation': 'type_contrat_domiciliation',
            'type_contrat_domiciliation_autre': 'type_contrat_domiciliation_autre',
            'loyer_mensuel_ttc': 'prix_mensuel',
            'frais_intermediaire_contrat': 'prix_inter',
            'date_debut_contrat': 'date_debut',
            'date_fin_contrat': 'date_fin',
            'taux_tva_pourcent': 'tva',
            'loyer_mensuel_ht': 'dh_ht',
            'montant_total_ht_contrat': 'montant_ht',
            'montant_pack_demarrage_ttc': 'pack_demarrage_montant',
            'loyer_mensuel_pack_demarrage_ttc': 'pack_demarrage_loyer',
            'type_renouvellement': 'type_renouvellement',
            'taux_tva_renouvellement_pourcent': 'tva_renouvellement',
            'loyer_mensuel_ht_renouvellement': 'dh_ht_renouvellement',
            'montant_total_ht_renouvellement': 'montant_ht_renouvellement',
            'loyer_mensuel_renouvellement_ttc': 'loyer_renouvellement_mensuel',
            'loyer_annuel_renouvellement_ttc': 'loyer_renouvellement_annuel',
            'period_domcil': 'period',
            'prix_contrat': 'prix_mensuel',
            'prix_intermediare_contrat': 'prix_inter',
            'dom_datedeb': 'date_debut',
            'dom_datefin': 'date_fin',
            'pack_demarrage_montant_ttc': 'pack_demarrage_montant',
            'pack_demarrage_loyer_mensuel_ttc': 'pack_demarrage_loyer',
            'loyer_renouvellement_mensuel_ttc': 'loyer_renouvellement_mensuel',
            'loyer_renouvellement_annuel_ttc': 'loyer_renouvellement_annuel',
        }
        inverse_collab_map = {
            'collaborateur_type': 'type',
            'collaborateur_code': 'code',
            'collaborateur_nom': 'nom',
            'collaborateur_ice': 'ice',
            'collaborateur_tp': 'tp',
            'collaborateur_rc': 'rc',
            'collaborateur_if': 'if',
            'collaborateur_tel_fixe': 'tel_fixe',
            'collaborateur_tel_mobile': 'tel_mobile',
            'collaborateur_adresse': 'adresse',
            'collaborateur_email': 'email',
        }

        soc_vals = {}
        for key, value in company_payload.items():
            if key in soc_map:
                soc_vals[soc_map[key]] = value

        sid = str(company_payload.get('id_societe') or '').strip()
        den = str(company_payload.get('den_ste') or '').strip()
        associes_list = []
        contrat_vals = {}
        collaborateur_vals = {}

        assoc_matches = self._filter_dashboard_rows(assoc_df, sid, den)
        if assoc_matches is not None and not assoc_matches.empty:
            for _, assoc_row in assoc_matches.iterrows():
                assoc_payload = {}
                for column in assoc_row.index:
                    if column in inverse_assoc_map:
                        assoc_payload[inverse_assoc_map[column]] = assoc_row.get(column)
                associes_list.append(assoc_payload)

        contrat_matches = self._filter_dashboard_rows(contrat_df, sid, den)
        if contrat_matches is not None and not contrat_matches.empty:
            contrat_row = contrat_matches.iloc[0]
            for column in contrat_row.index:
                if column in inverse_contrat_map:
                    contrat_vals[inverse_contrat_map[column]] = contrat_row.get(column)

        collab_matches = self._filter_dashboard_rows(collab_df, sid, den)
        if collab_matches is not None and not collab_matches.empty:
            collab_row = collab_matches.iloc[0]
            for column in collab_row.index:
                if column in inverse_collab_map:
                    collaborateur_vals[inverse_collab_map[column]] = collab_row.get(column)

        return {
            'societe': soc_vals or {},
            'associes': associes_list,
            'contrat': contrat_vals,
            'collaborateur': collaborateur_vals,
        }

    def _apply_dashboard_edit_for_page(self, page_key: str, page_values):
        """Persist edits coming from the dashboard into the Excel DB."""
        try:
            import pandas as _pd
            from ..utils.utils import normalize_canonical_dataframe_for_storage, normalize_excel_storage
        except Exception:
            return

        context = getattr(self, '_dashboard_edit_context', None) or {}
        sid = str(context.get('id_societe') or '').strip()
        den_old = str(context.get('den_ste') or '').strip()
        if not sid and not den_old:
            return

        db_path, soc_df, assoc_df, contrat_df, collab_df = self._load_dashboard_workbook_frames()

        def _update_societe(df, vals):
            if df.empty:
                return df
            mask = None
            if sid and 'id_societe' in df.columns:
                mask = df['id_societe'].astype(str).str.strip() == sid
            elif den_old and 'den_ste' in df.columns:
                mask = df['den_ste'].astype(str).str.strip().str.lower() == den_old.lower()
            if mask is None:
                return df
            mapping = {
                'denomination': 'den_ste',
                'den_ste': 'den_ste',
                'forme_juridique': 'forme_jur',
                'forme_jur': 'forme_jur',
                'ice': 'ice',
                'date_ice': 'date_ice',
                'date_certificat_negatif': 'date_ice',
                'date_expiration_certificat_negatif': 'date_exp_cert_neg',
                'capital': 'capital',
                'parts_social': 'part_social',
                'valeur_nominale': 'valeur_nominale',
                'adresse': 'ste_adress',
                'ste_adress': 'ste_adress',
                'tribunal': 'tribunal',
                'type_generation': 'type_generation',
                'generation_type': 'type_generation',
                'procedure_creation': 'procedure_creation',
                'creation_procedure': 'procedure_creation',
                'mode_depot_creation': 'mode_depot_creation',
                'creation_depot_mode': 'mode_depot_creation',
            }
            for k, col in mapping.items():
                if k in vals and col in df.columns:
                    df.loc[mask, col] = vals.get(k)
            if sid and 'id_societe' in df.columns:
                df.loc[mask, 'id_societe'] = sid
            return df

        def _build_associes_rows(values_list):
            rows = []
            if not values_list:
                return rows
            existing_ids = _pd.to_numeric(assoc_df.get('id_associe', []), errors='coerce').dropna()
            next_id = int(existing_ids.max()) + 1 if not existing_ids.empty else 1
            for item in values_list:
                if not isinstance(item, dict):
                    continue
                row = {h: '' for h in _const.associe_headers}
                row['id_associe'] = next_id
                next_id += 1
                row['id_societe'] = sid or row.get('id_societe', '')
                row['den_ste'] = item.get('den_ste') or context.get('den_ste') or den_old or ''
                mapping = {
                    'civilite': 'civil',
                    'prenom': 'prenom',
                    'nom': 'nom',
                    'nationalite': 'nationality',
                    'num_piece': 'cin_num',
                    'validite_piece': 'cin_validaty',
                    'date_naiss': 'date_naiss',
                    'lieu_naiss': 'lieu_naiss',
                    'adresse': 'adresse',
                    'telephone': 'phone',
                    'email': 'email',
                    'percentage': 'part_percent',
                    'part_percentage': 'part_percent',
                    'parts': 'parts',
                    'num_parts': 'parts',
                    'capital_detenu': 'capital_detenu',
                    'est_gerant': 'is_gerant',
                    'qualite': 'quality',
                }
                for k, col in mapping.items():
                    if k in item and col in row:
                        row[col] = item.get(k)
                rows.append(row)
            return rows

        def _build_contrat_row(values_dict):
            row = {h: '' for h in _const.contrat_headers}
            existing_ids = _pd.to_numeric(contrat_df.get('id_contrat', []), errors='coerce').dropna()
            row['id_contrat'] = int(existing_ids.max()) + 1 if not existing_ids.empty else 1
            row['id_societe'] = sid or row.get('id_societe', '')
            row['den_ste'] = values_dict.get('den_ste') or context.get('den_ste') or den_old or ''
            mapping = {
                'date_contrat': 'date_contrat',
                'period': 'duree_contrat_mois',
                'type_contrat_domiciliation': 'type_contrat_domiciliation',
                'type_contrat_domiciliation_autre': 'type_contrat_domiciliation_autre',
                'prix_mensuel': 'loyer_mensuel_ttc',
                'prix_inter': 'frais_intermediaire_contrat',
                'date_debut': 'date_debut_contrat',
                'date_fin': 'date_fin_contrat',
                'tva': 'taux_tva_pourcent',
                'dh_ht': 'loyer_mensuel_ht',
                'montant_ht': 'montant_total_ht_contrat',
                'pack_demarrage_montant': 'montant_pack_demarrage_ttc',
                'pack_demarrage_loyer': 'loyer_mensuel_pack_demarrage_ttc',
                'type_renouvellement': 'type_renouvellement',
                'tva_renouvellement': 'taux_tva_renouvellement_pourcent',
                'dh_ht_renouvellement': 'loyer_mensuel_ht_renouvellement',
                'montant_ht_renouvellement': 'montant_total_ht_renouvellement',
                'loyer_renouvellement_mensuel': 'loyer_mensuel_renouvellement_ttc',
                'loyer_renouvellement_annuel': 'loyer_annuel_renouvellement_ttc',
            }
            for k, col in mapping.items():
                if k in values_dict and col in row:
                    row[col] = values_dict.get(k)
            return row

        if page_key == 'societe':
            soc_df = _update_societe(soc_df, page_values or {})
            try:
                new_den = str((page_values or {}).get('denomination') or (page_values or {}).get('den_ste') or '').strip()
                if new_den:
                    if not assoc_df.empty and 'den_ste' in assoc_df.columns and 'id_societe' in assoc_df.columns:
                        assoc_df.loc[assoc_df['id_societe'].astype(str).str.strip() == sid, 'den_ste'] = new_den
                    if not contrat_df.empty and 'den_ste' in contrat_df.columns and 'id_societe' in contrat_df.columns:
                        contrat_df.loc[contrat_df['id_societe'].astype(str).str.strip() == sid, 'den_ste'] = new_den
                    if not collab_df.empty and 'den_ste' in collab_df.columns and 'id_societe' in collab_df.columns:
                        collab_df.loc[collab_df['id_societe'].astype(str).str.strip() == sid, 'den_ste'] = new_den
                    context['den_ste'] = new_den
                    self._dashboard_edit_context = context
            except Exception:
                pass
        elif page_key == 'associes':
            if 'id_societe' in assoc_df.columns:
                assoc_df = assoc_df[assoc_df['id_societe'].astype(str).str.strip() != sid]
            rows = _build_associes_rows(page_values or [])
            if rows:
                assoc_df = _pd.concat([assoc_df, _pd.DataFrame(rows)], ignore_index=True)
        elif page_key == 'contrat':
            if 'id_societe' in contrat_df.columns:
                contrat_df = contrat_df[contrat_df['id_societe'].astype(str).str.strip() != sid]
            if page_values:
                row = _build_contrat_row(page_values)
                contrat_df = _pd.concat([contrat_df, _pd.DataFrame([row])], ignore_index=True)
        else:
            return

        soc_df = normalize_canonical_dataframe_for_storage(soc_df.reindex(columns=_const.societe_headers, fill_value=''))
        assoc_df = normalize_canonical_dataframe_for_storage(assoc_df.reindex(columns=_const.associe_headers, fill_value=''))
        contrat_df = normalize_canonical_dataframe_for_storage(contrat_df.reindex(columns=_const.contrat_headers, fill_value=''))
        collab_df = normalize_canonical_dataframe_for_storage(collab_df.reindex(columns=_const.collaborateur_headers, fill_value=''))

        try:
            with _pd.ExcelWriter(db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                soc_df.to_excel(writer, sheet_name='Societes', index=False)
                assoc_df.to_excel(writer, sheet_name='Associes', index=False)
                contrat_df.to_excel(writer, sheet_name='Contrats', index=False)
                collab_df.to_excel(writer, sheet_name='Collaborateurs', index=False)
        except TypeError:
            from openpyxl import load_workbook
            wb = load_workbook(db_path)
            for sname in ('Societes', 'Associes', 'Contrats', 'Collaborateurs'):
                if sname in wb.sheetnames:
                    try:
                        wb.remove(wb[sname])
                    except Exception:
                        pass
            wb.save(db_path)
            with _pd.ExcelWriter(db_path, engine='openpyxl', mode='a') as writer:
                soc_df.to_excel(writer, sheet_name='Societes', index=False)
                assoc_df.to_excel(writer, sheet_name='Associes', index=False)
                contrat_df.to_excel(writer, sheet_name='Contrats', index=False)
                collab_df.to_excel(writer, sheet_name='Collaborateurs', index=False)

        try:
            normalize_excel_storage(db_path)
        except Exception:
            pass

        try:
            dashboard = getattr(self, '_dashboard_window', None)
            if dashboard is not None and dashboard.winfo_exists():
                dashboard._load_data()
                dashboard._show_page(getattr(dashboard, '_current_page', 'societe'))
        except Exception:
            pass

    def handle_dashboard_action(self, action: str, payload: dict | None, page_key: str | None = None):
        """Handle actions coming from the DashboardView.

        Supported actions:
        - 'add'    : clear forms and show the societe page for a new entry
        - 'edit'   : prefill forms with the selected company and show the relevant page
        - 'delete' : remove the company (and related rows) from the Excel DB after confirmation
        - other    : ignored
        """
        try:
            page_key = self._infer_dashboard_page_key(payload, page_key)
            # Bring main window to front
            top = self.winfo_toplevel()
            try:
                top.deiconify()
                top.lift()
                top.focus_force()
            except Exception:
                pass

            if action == 'add':
                if page_key == 'collaborateur':
                    self._open_collaborateur_dialog()
                    return {'status': 'opened', 'page': 'collaborateur'}

                # Reset forms to default/new state and show first page
                self.reset()
                self._dashboard_edit_context = None
                self.show_page(0)
                return {'status': 'opened', 'page': 'societe'}

            if action == 'edit':
                if not payload:
                    messagebox.showwarning('Modifier', 'Aucune donnée fournie pour modification.')
                    return {'status': 'invalid'}

                try:
                    _db_path, soc_df, assoc_df, contrat_df, collab_df = self._load_dashboard_workbook_frames()
                except FileNotFoundError:
                    messagebox.showerror('Erreur', 'Fichier de base de données introuvable.')
                    return {'status': 'error'}
                except Exception as e:
                    try:
                        from ..utils.utils import ErrorHandler
                        ErrorHandler.handle_error(e, 'Erreur lors du chargement des données')
                    except Exception:
                        messagebox.showerror('Erreur', f'Impossible de charger les données: {e}')
                    return {'status': 'error'}

                company_payload = self._resolve_company_payload_from_dashboard(page_key, payload, soc_df)
                if not company_payload:
                    messagebox.showwarning(
                        'Modifier',
                        "Impossible de retrouver la société liée à l'enregistrement sélectionné."
                    )
                    return {'status': 'not_found'}

                try:
                    self._dashboard_edit_context = {
                        'page_key': page_key,
                        'id_societe': company_payload.get('id_societe'),
                        'den_ste': company_payload.get('den_ste'),
                    }
                except Exception:
                    self._dashboard_edit_context = None

                values = self._build_dashboard_edit_values(company_payload, assoc_df, contrat_df, collab_df)
                if page_key == 'collaborateur':
                    try:
                        self._dashboard_edit_context = {
                            'page_key': page_key,
                            'id_societe': company_payload.get('id_societe'),
                            'den_ste': company_payload.get('den_ste'),
                            'id_collaborateur': payload.get('id_collaborateur') if isinstance(payload, dict) else None,
                        }
                    except Exception:
                        self._dashboard_edit_context = None
                    self._open_collaborateur_dialog(values.get('collaborateur', {}), edit_context=self._dashboard_edit_context)
                    return {'status': 'opened', 'page': 'collaborateur'}

                self.set_values(values)
                self.show_page(self._dashboard_page_index(page_key))
                return {'status': 'opened', 'page': page_key}

            if action == 'delete':
                if not payload:
                    messagebox.showwarning('Supprimer', 'Aucune société sélectionnée pour suppression.')
                    return {'status': 'invalid'}

                # Remove rows from the Excel workbook
                try:
                    import pandas as _pd
                    db_path, soc_df, assoc_df, contrat_df, collab_df = self._load_dashboard_workbook_frames()
                    company_payload = self._resolve_company_payload_from_dashboard(page_key, payload, soc_df)
                    if not company_payload:
                        messagebox.showwarning(
                            'Supprimer',
                            "Impossible de retrouver la société liée à l'enregistrement sélectionné."
                        )
                        return {'status': 'not_found'}

                    den = str(company_payload.get('den_ste') or '').strip()
                    sid = str(company_payload.get('id_societe') or '').strip()
                    if not messagebox.askyesno(
                        'Confirmation',
                        f"Voulez-vous vraiment supprimer le dossier société '{den}' ?"
                    ):
                        return {'status': 'cancelled'}

                    if sid:
                        if not soc_df.empty and 'id_societe' in soc_df.columns:
                            soc_df = soc_df[~(soc_df['id_societe'].astype(str).str.strip() == sid)]
                        if not assoc_df.empty and 'id_societe' in assoc_df.columns:
                            assoc_df = assoc_df[~(assoc_df['id_societe'].astype(str).str.strip() == sid)]
                        if not contrat_df.empty and 'id_societe' in contrat_df.columns:
                            contrat_df = contrat_df[~(contrat_df['id_societe'].astype(str).str.strip() == sid)]
                        if not collab_df.empty and 'id_societe' in collab_df.columns:
                            collab_df = collab_df[~(collab_df['id_societe'].astype(str).str.strip() == sid)]
                    else:
                        # fallback to DEN_STE match
                        if not soc_df.empty and 'den_ste' in soc_df.columns:
                            soc_df = soc_df[~(soc_df['den_ste'].astype(str).str.strip().str.lower() == den.strip().lower())]
                        if not assoc_df.empty and 'den_ste' in assoc_df.columns:
                            assoc_df = assoc_df[~(assoc_df['den_ste'].astype(str).str.strip().str.lower() == den.strip().lower())]
                        if not contrat_df.empty and 'den_ste' in contrat_df.columns:
                            contrat_df = contrat_df[~(contrat_df['den_ste'].astype(str).str.strip().str.lower() == den.strip().lower())]
                        if not collab_df.empty and 'den_ste' in collab_df.columns:
                            collab_df = collab_df[~(collab_df['den_ste'].astype(str).str.strip().str.lower() == den.strip().lower())]

                    # Write back sheets replacing them
                    try:
                        from ..utils.utils import normalize_canonical_dataframe_for_storage, normalize_excel_storage

                        soc_df = normalize_canonical_dataframe_for_storage(soc_df.reindex(columns=_const.societe_headers, fill_value=''))
                        assoc_df = normalize_canonical_dataframe_for_storage(assoc_df.reindex(columns=_const.associe_headers, fill_value=''))
                        contrat_df = normalize_canonical_dataframe_for_storage(contrat_df.reindex(columns=_const.contrat_headers, fill_value=''))
                        collab_df = normalize_canonical_dataframe_for_storage(collab_df.reindex(columns=_const.collaborateur_headers, fill_value=''))

                        with _pd.ExcelWriter(db_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            soc_df.to_excel(writer, sheet_name='Societes', index=False)
                            assoc_df.to_excel(writer, sheet_name='Associes', index=False)
                            contrat_df.to_excel(writer, sheet_name='Contrats', index=False)
                            collab_df.to_excel(writer, sheet_name='Collaborateurs', index=False)
                    except TypeError:
                        # pandas older version fallback
                        from openpyxl import load_workbook
                        wb = load_workbook(db_path)
                        for sname, df in (('Societes', soc_df), ('Associes', assoc_df), ('Contrats', contrat_df), ('Collaborateurs', collab_df)):
                            if sname in wb.sheetnames:
                                try:
                                    std = wb[sname]
                                    wb.remove(std)
                                except Exception:
                                    pass
                        wb.save(db_path)
                        with _pd.ExcelWriter(db_path, engine='openpyxl', mode='a') as writer:
                            soc_df.to_excel(writer, sheet_name='Societes', index=False)
                            assoc_df.to_excel(writer, sheet_name='Associes', index=False)
                            contrat_df.to_excel(writer, sheet_name='Contrats', index=False)
                            collab_df.to_excel(writer, sheet_name='Collaborateurs', index=False)

                    try:
                        normalize_excel_storage(db_path)
                    except Exception:
                        pass

                    messagebox.showinfo('Succès', f"Dossier société '{den}' supprimé avec succès.")
                    return {'status': 'deleted', 'company_name': den}
                except PermissionError:
                    messagebox.showerror('Erreur', 'Le fichier Excel est ouvert dans une autre application. Fermez Excel et réessayez.')
                    return {'status': 'error'}
                except Exception as e:
                    try:
                        from ..utils.utils import ErrorHandler
                        ErrorHandler.handle_error(e, 'Erreur lors de la suppression')
                    except Exception:
                        messagebox.showerror('Erreur', f'Impossible de supprimer: {e}')
                    return {'status': 'error'}

        except Exception as e:
            try:
                logger = __import__('logging').getLogger(__name__)
                logger.exception('handle_dashboard_action failed: %s', e)
            except Exception:
                pass
            return {'status': 'error'}
