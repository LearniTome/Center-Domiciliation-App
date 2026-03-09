import tkinter as tk
from tkinter import ttk, messagebox
from .styles import ModernTheme
import os
import pandas as pd
from pathlib import Path
import json
import logging
import traceback
from typing import Optional, Callable, Any
import datetime
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from fractions import Fraction
from urllib import request as _urlrequest
from urllib import error as _urlerror
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pathlib import Path as _Path
import pandas as _pd
import shutil

# Configuration du logging
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Process-local cache for reference sheets loaded from the Excel database.
# Keyed by resolved DB path and invalidated when file mtime changes.
_REFERENCE_DATA_CACHE = {}

_DB_DATE_COLUMNS = {
    'DATE_ICE',
    'DATE_EXP_CERT_NEG',
    'CIN_VALIDATY',
    'DATE_NAISS',
    'DATE_CONTRAT',
    'DATE_DEBUT_CONTRAT',
    'DATE_FIN_CONTRAT',
}

_DB_BOOLEAN_COLUMNS = {
    'IS_GERANT',
}

_DB_INTEGER_COLUMNS = {
    'CAPITAL',
    'PART_SOCIAL',
    'VALEUR_NOMINALE',
    'PART_PERCENT',
    'PARTS',
    'CAPITAL_DETENU',
    'DUREE_CONTRAT_MOIS',
    'TAUX_TVA_POURCENT',
    'TAUX_TVA_RENOUVELLEMENT_POURCENT',
}

_DB_AMOUNT_COLUMNS = {
    'LOYER_MENSUEL_TTC',
    'FRAIS_INTERMEDIAIRE_CONTRAT',
    'LOYER_MENSUEL_HT',
    'MONTANT_TOTAL_HT_CONTRAT',
    'MONTANT_PACK_DEMARRAGE_TTC',
    'LOYER_MENSUEL_PACK_DEMARRAGE_TTC',
    'LOYER_MENSUEL_HT_RENOUVELLEMENT',
    'MONTANT_TOTAL_HT_RENOUVELLEMENT',
    'LOYER_MENSUEL_RENOUVELLEMENT_TTC',
    'LOYER_ANNUEL_RENOUVELLEMENT_TTC',
}


def _storage_empty_string(value) -> str:
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    text = str(value).strip()
    return '' if text.lower() == 'nan' else text


def _format_storage_date(value) -> str:
    text = _storage_empty_string(value)
    if not text:
        return ''
    if len(text) == 10 and text[2] == '/' and text[5] == '/':
        return text

    parsed = None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            break
        except Exception:
            continue

    if parsed is None:
        try:
            parsed = _pd.to_datetime(text, dayfirst='/' in text, errors='coerce')
        except Exception:
            parsed = None
    if parsed is None or _pd.isna(parsed):
        return text
    return parsed.strftime('%d/%m/%Y')


def _format_storage_bool(value) -> str:
    text = _storage_empty_string(value)
    if not text:
        return ''
    lowered = text.lower()
    if lowered in {'1', 'true', 'vrai', 'yes', 'oui'}:
        return 'Oui'
    if lowered in {'0', 'false', 'faux', 'no', 'non'}:
        return 'Non'
    return text


def _format_storage_number(value, min_decimals: int = 0, max_decimals: int = 4) -> str:
    text = _storage_empty_string(value)
    if not text:
        return ''

    normalized = text.replace('\xa0', ' ').replace(' ', '')
    if ',' in normalized and '.' in normalized:
        normalized = normalized.replace(',', '')
    elif ',' in normalized:
        normalized = normalized.replace(',', '.')

    try:
        number = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return text

    if min_decimals == 0 and number == number.to_integral():
        return f"{int(number):,}".replace(',', ' ')

    original_decimals = ''
    if '.' in normalized:
        original_decimals = normalized.split('.', 1)[1]
    trimmed_len = len(original_decimals.rstrip('0'))
    decimals = max(min_decimals, min(max_decimals, trimmed_len if trimmed_len else min_decimals))
    if decimals <= 0:
        return f"{int(number):,}".replace(',', ' ')

    formatted = f"{number:,.{decimals}f}"
    integer_part, fractional_part = formatted.split('.')
    integer_part = integer_part.replace(',', ' ')
    if decimals > min_decimals:
        fractional_part = fractional_part.rstrip('0')
        if len(fractional_part) < min_decimals:
            fractional_part = fractional_part.ljust(min_decimals, '0')
    return f"{integer_part},{fractional_part}"


def normalize_canonical_dataframe_for_storage(df: _pd.DataFrame) -> _pd.DataFrame:
    """Normalize canonical workbook values to user-facing strings in the Excel file."""
    if df is None:
        return _pd.DataFrame()

    out = df.copy()
    if out.empty:
        return out.fillna('')

    for col in out.columns:
        if col in _DB_DATE_COLUMNS:
            out[col] = out[col].apply(_format_storage_date)
        elif col in _DB_BOOLEAN_COLUMNS:
            out[col] = out[col].apply(_format_storage_bool)
        elif col in _DB_AMOUNT_COLUMNS:
            out[col] = out[col].apply(lambda value: _format_storage_number(value, min_decimals=2, max_decimals=4))
        elif col in _DB_INTEGER_COLUMNS:
            out[col] = out[col].apply(_format_storage_number)
        else:
            out[col] = out[col].apply(_storage_empty_string)

    return out.fillna('')

class ErrorHandler:
    """Gestionnaire d'erreurs centralisé pour l'application"""
    @staticmethod
    def handle_error(error: Exception,
                    message: str = "Une erreur s'est produite",
                    show_dialog: bool = True,
                    callback: Optional[Callable] = None) -> None:
        """
        Gère une erreur de manière centralisée
        Args:
            error: L'exception à gérer
            message: Message à afficher à l'utilisateur
            show_dialog: Si True, affiche une boîte de dialogue
            callback: Fonction à appeler après le traitement de l'erreur
        """
        # Log l'erreur
        logger.error(f"{message}: {str(error)}\n{traceback.format_exc()}")

        # Affiche la boîte de dialogue si demandé
        if show_dialog:
            messagebox.showerror("Erreur", f"{message}\n\nDétails: {str(error)}")

        # Exécute le callback si fourni
        if callback:
            try:
                callback()
            except Exception as e:
                logger.error(f"Erreur dans le callback: {str(e)}")


class ThemedMessageBox:
    """Dark-mode replacements for tkinter.messagebox."""

    _installed = False
    _originals = {}

    @classmethod
    def install(cls):
        if cls._installed:
            return
        cls._originals = {
            'showinfo': messagebox.showinfo,
            'showwarning': messagebox.showwarning,
            'showerror': messagebox.showerror,
            'askyesno': messagebox.askyesno,
        }
        messagebox.showinfo = cls.showinfo
        messagebox.showwarning = cls.showwarning
        messagebox.showerror = cls.showerror
        messagebox.askyesno = cls.askyesno
        cls._installed = True

    @classmethod
    def _coerce_message(cls, message=None, detail=None) -> str:
        main = "" if message is None else str(message)
        extra = "" if detail is None else str(detail)
        if main and extra:
            return f"{main}\n\n{extra}"
        return main or extra

    @classmethod
    def _resolve_parent(cls, parent=None):
        if parent is not None:
            return parent
        try:
            return tk._default_root
        except Exception:
            return None

    @classmethod
    def _fallback(cls, kind: str, title: str, message: str):
        original = cls._originals.get('askyesno' if kind == 'question' else f"show{kind}")
        if original is None:
            return False if kind == 'question' else 'ok'
        return original(title, message)

    @classmethod
    def _show_dialog(cls, kind: str, title: str, message: str, parent=None):
        host = cls._resolve_parent(parent)
        if host is None:
            return cls._fallback(kind, title, message)

        dialog = tk.Toplevel(host)
        dialog.withdraw()
        dialog.title(title or "")
        dialog.resizable(False, False)

        try:
            dialog.transient(host)
            dialog.grab_set()
        except Exception:
            pass

        try:
            theme_manager = ThemeManager(dialog)
            colors = theme_manager.colors
        except Exception:
            colors = {
                'bg': '#1f1f1f',
                'fg': '#f3f3f3',
                'label_fg': '#e6e6e6',
                'info': '#17a2b8',
                'warning': '#ffc107',
                'error': '#ff6b6b',
                'accent': '#6c7783',
            }
            try:
                dialog.configure(bg=colors['bg'])
            except Exception:
                pass

        icon_map = {
            'info': ('i', colors.get('info', '#17a2b8')),
            'warning': ('!', colors.get('warning', '#ffc107')),
            'error': ('x', colors.get('error', '#ff6b6b')),
            'question': ('?', colors.get('accent', '#6c7783')),
        }
        icon_text, icon_color = icon_map.get(kind, ('i', colors.get('info', '#17a2b8')))

        outer = ttk.Frame(dialog, padding=16)
        outer.pack(fill='both', expand=True)

        body = ttk.Frame(outer)
        body.pack(fill='both', expand=True)

        icon_label = tk.Label(
            body,
            text=icon_text,
            font=('Segoe UI', 22, 'bold'),
            bg=colors.get('bg', '#1f1f1f'),
            fg=icon_color,
            width=3,
        )
        icon_label.pack(side='left', anchor='n', padx=(0, 12), pady=(4, 0))

        ttk.Label(
            body,
            text=message,
            justify='left',
            style='FieldLabel.TLabel',
            wraplength=500,
        ).pack(side='left', fill='both', expand=True)

        button_row = ttk.Frame(outer)
        button_row.pack(fill='x', pady=(16, 0))

        result = {'value': False if kind == 'question' else 'ok'}

        def _finish(value):
            result['value'] = value
            try:
                dialog.grab_release()
            except Exception:
                pass
            dialog.destroy()

        if kind == 'question':
            ttk.Button(
                button_row,
                text="Non",
                command=lambda: _finish(False),
                style='Secondary.TButton',
            ).pack(side='right')
            ttk.Button(
                button_row,
                text="Oui",
                command=lambda: _finish(True),
                style='Success.TButton',
            ).pack(side='right', padx=(0, 8))
            dialog.bind('<Return>', lambda _e: _finish(True))
            dialog.bind('<Escape>', lambda _e: _finish(False))
        else:
            ttk.Button(
                button_row,
                text="OK",
                command=lambda: _finish('ok'),
                style='Success.TButton',
            ).pack(side='right')
            dialog.bind('<Return>', lambda _e: _finish('ok'))
            dialog.bind('<Escape>', lambda _e: _finish('ok'))

        try:
            dialog.update_idletasks()
            req_w = max(420, min(640, dialog.winfo_reqwidth() + 20))
            req_h = max(180, min(360, dialog.winfo_reqheight() + 20))
            dialog.geometry(f"{req_w}x{req_h}")
            dialog.update_idletasks()
            WindowManager.center_window(dialog)
        except Exception:
            pass

        dialog.deiconify()
        try:
            dialog.after_idle(lambda: WindowManager.center_window(dialog))
        except Exception:
            pass
        dialog.focus_force()
        dialog.wait_window()
        return result['value']

    @classmethod
    def showinfo(cls, title=None, message=None, **kwargs):
        return cls._show_dialog(
            'info',
            title or "Information",
            cls._coerce_message(message, kwargs.get('detail')),
            parent=kwargs.get('parent'),
        )

    @classmethod
    def showwarning(cls, title=None, message=None, **kwargs):
        return cls._show_dialog(
            'warning',
            title or "Attention",
            cls._coerce_message(message, kwargs.get('detail')),
            parent=kwargs.get('parent'),
        )

    @classmethod
    def showerror(cls, title=None, message=None, **kwargs):
        return cls._show_dialog(
            'error',
            title or "Erreur",
            cls._coerce_message(message, kwargs.get('detail')),
            parent=kwargs.get('parent'),
        )

    @classmethod
    def askyesno(cls, title=None, message=None, **kwargs):
        return bool(
            cls._show_dialog(
                'question',
                title or "Confirmation",
                cls._coerce_message(message, kwargs.get('detail')),
                parent=kwargs.get('parent'),
            )
        )

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20

        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")

        label = ttk.Label(self.tooltip, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1)
        label.pack(padx=1)

    def leave(self, event=None):
        if self.tooltip:
            try:
                self.tooltip.destroy()
            except Exception:
                pass
            self.tooltip = None

class ThemeManager:
    """ThemeManager simplifié : délègue tout à ModernTheme dans `src.utils.styles`.

    Ceci évite la duplication des styles dans plusieurs fichiers et garde
    un seul endroit (`ModernTheme`) responsable des apparences.
    """
    def __init__(self, root):
        self.root = root
        self.pref_path = Path(__file__).resolve().parent.parent.parent / 'config' / 'preferences.json'
        mode = 'dark'
        try:
            if self.pref_path.exists():
                with self.pref_path.open('r', encoding='utf-8') as f:
                    prefs = json.load(f)
                    mode = prefs.get('theme', 'dark')
        except Exception:
            mode = 'dark'

        # Use the centralized ModernTheme from src.utils.styles
        self.theme = ModernTheme(root, mode=mode)
        self.style = self.theme.style
        self.colors = self.theme.colors
        self.is_dark_mode = self.theme.mode == 'dark'
        # Apply background to root and existing canvases to keep tk widgets in sync
        try:
            self._apply_root_background()
            self._update_canvas_backgrounds()
        except Exception:
            logger.debug('Failed to update canvas/background on init', exc_info=True)
        # Start background monitor to ensure dynamically created non-ttk widgets
        # (like Combobox popdown Listbox) get themed shortly after creation.
        try:
            # _start_non_ttk_monitor is defined on this class below
            self._start_non_ttk_monitor()
        except Exception:
            logger.debug('Failed to start non-ttk monitor', exc_info=True)
        try:
            ThemedMessageBox.install()
        except Exception:
            logger.debug('Failed to install themed message boxes', exc_info=True)

    def set_theme(self, mode: str):
        if mode not in ('light', 'dark'):
            return
        # Recreate centralized theme
        self.theme = ModernTheme(self.root, mode=mode)
        self.style = self.theme.style
        self.colors = self.theme.colors
        self.is_dark_mode = self.theme.mode == 'dark'
        # Persist
        try:
            prefs = {'theme': mode}
            self.pref_path.parent.mkdir(parents=True, exist_ok=True)
            with self.pref_path.open('w', encoding='utf-8') as f:
                json.dump(prefs, f, ensure_ascii=False, indent=2)
        except Exception:
            logger.exception('Failed to persist theme preference')
        # Update root and existing canvases so widgets created before theme change update
        try:
            self._apply_root_background()
            self._update_canvas_backgrounds()
            # Update non-ttk widget styles for existing widgets
            self._apply_non_ttk_styles()
        except Exception:
            logger.debug('Failed to update canvas/background after set_theme', exc_info=True)

    def toggle_theme(self):
        new_mode = 'dark' if self.theme.mode == 'light' else 'light'
        self.set_theme(new_mode)

    def _apply_root_background(self):
        """Configure the top-level root background to the current theme bg color."""
        try:
            if hasattr(self.root, 'configure'):
                # Some platforms expect 'background' or 'bg'
                try:
                    self.root.configure(background=self.colors['bg'])
                except Exception:
                    try:
                        self.root.configure(bg=self.colors['bg'])
                    except Exception:
                        pass
        except Exception:
            logger.debug('Failed to apply root background', exc_info=True)

    def _update_canvas_backgrounds(self):
        """Recursively find tk.Canvas widgets under root and update their background.

        This ensures Canvas widgets (which are not ttk and do not follow ttk styles)
        reflect the current theme background when the theme is changed at runtime.
        """
        try:
            # recursive walk
            def _walk(widget):
                for child in widget.winfo_children():
                    # tk.Canvas class is available as tk.Canvas
                    if isinstance(child, tk.Canvas):
                        try:
                            child.configure(background=self.colors['bg'])
                        except Exception:
                            try:
                                child.configure(bg=self.colors['bg'])
                            except Exception:
                                pass
                    # recurse
                    try:
                        _walk(child)
                    except Exception:
                        pass

            _walk(self.root)
        except Exception:
            logger.debug('Failed to update canvas backgrounds', exc_info=True)

    def _apply_non_ttk_styles(self):
        """Apply colors to non-ttk widgets (Listbox, Menu, Text) and update existing instances.

        This uses root.option_add to set defaults for new widgets and walks the widget tree
        to update already-created widgets so their selection colors match the theme.
        """
        try:
            # Use section_bg for Listbox in dark mode for better contrast
            listbox_bg = self.colors.get('section_bg', self.colors['bg']) if self.theme.mode == 'dark' else self.colors['bg']
            input_bg = self.colors.get('input_bg', self.colors['bg'])

            # Set global defaults for new widgets
            try:
                self.root.option_add('*Listbox.background', listbox_bg)
                self.root.option_add('*Listbox.foreground', self.colors['fg'])
                self.root.option_add('*Listbox.selectBackground', self.colors['accent'])
                self.root.option_add('*Listbox.selectForeground', 'white')
                self.root.option_add('*Listbox.relief', 'solid')
                self.root.option_add('*Listbox.borderwidth', '1')

                self.root.option_add('*Text.background', input_bg)
                self.root.option_add('*Text.foreground', self.colors['fg'])
                self.root.option_add('*Text.insertBackground', self.colors['fg'])

                self.root.option_add('*Menu.background', listbox_bg)
                self.root.option_add('*Menu.foreground', self.colors['fg'])
                self.root.option_add('*Menu.activeBackground', self.colors['accent'])
                self.root.option_add('*Menu.activeForeground', 'white')
            except Exception:
                pass

            # Walk existing widgets and update instances
            def _walk_update(widget):
                for child in widget.winfo_children():
                    try:
                        if isinstance(child, tk.Listbox):
                            child.configure(background=listbox_bg, foreground=self.colors['fg'], selectbackground=self.colors['accent'], selectforeground='white', relief='solid', borderwidth=1)
                        elif isinstance(child, tk.Text):
                            child.configure(background=input_bg, foreground=self.colors['fg'], insertbackground=self.colors['fg'])
                        elif isinstance(child, tk.Menu):
                            try:
                                child.configure(background=listbox_bg, foreground=self.colors['fg'], activebackground=self.colors['accent'], activeforeground='white')
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # Recurse
                    try:
                        _walk_update(child)
                    except Exception:
                        pass

            _walk_update(self.root)
        except Exception:
            logger.debug('Failed to apply non-ttk styles', exc_info=True)

    def _start_non_ttk_monitor(self, interval_ms: int = 400):
        """Start a periodic after() loop that reapplies non-ttk styles.

        This ensures popdown widgets created later (combobox lists) are themed.
        """
        try:
            # cancel previous if any
            if hasattr(self, '_monitor_id') and self._monitor_id:
                try:
                    self.root.after_cancel(self._monitor_id)
                except Exception:
                    pass

            def _monitor():
                try:
                    self._apply_non_ttk_styles()
                except Exception:
                    pass
                try:
                    self._monitor_id = self.root.after(interval_ms, _monitor)
                except Exception:
                    pass

            # schedule first run
            self._monitor_id = self.root.after(interval_ms, _monitor)
        except Exception:
            logger.debug('Failed to start monitor', exc_info=True)

    def apply_widget_styles(self, widget):
        """Applique un style cohérent en utilisant les noms définis dans ModernTheme.

        Cette méthode remplace les styles locaux et force l'utilisation du
        fichier unique `src/utils/styles.py` comme source de vérité.
        """
        try:
            if isinstance(widget, ttk.Entry):
                widget.configure(style='TEntry')
            elif isinstance(widget, ttk.Combobox):
                widget.configure(style='TCombobox')
            elif isinstance(widget, ttk.Label):
                widget.configure(style='FieldLabel.TLabel')
            elif isinstance(widget, ttk.LabelFrame):
                # Use Section frame style for labeled sections
                widget.configure(style='Section.TFrame')
            elif isinstance(widget, ttk.Frame):
                widget.configure(style='TFrame')
            elif isinstance(widget, ttk.Checkbutton):
                # fall back to default checkbutton style
                widget.configure(style='TCheckbutton')
        except Exception:
            # Defensive: ignore failures to avoid breaking UI creation
            logger.debug('apply_widget_styles: failed to configure widget style', exc_info=True)
        return widget

class WidgetFactory:
    _ICON_REGISTRY = {
        'settings': 'settings.png',
        'dashboard': 'dashboard.png',
        'close': 'close.png',
        'next': 'next.png',
        'prev': 'prev.png',
        'finish': 'finish.png',
        'save': 'save.png',
        'new': 'new.png',
        'add': 'add.png',
        'edit': 'edit.png',
        'delete': 'delete.png',
        'refresh': 'refresh.png',
        'folder': 'folder.png',
        'upload': 'upload.png',
        'confirm': 'confirm.png',
        'document': 'document.png',
        'company': 'company.png',
        'contract': 'contract.png',
        'person': 'person.png',
        'view': 'view.png',
    }
    _EMOJI_ICON_KEY = {
        '⚙': 'settings',
        '📊': 'dashboard',
        '❌': 'close',
        '▶': 'next',
        '◀': 'prev',
        '🏁': 'finish',
        '💾': 'save',
        '🆕': 'new',
        '➕': 'add',
        '✏️': 'edit',
        '✏': 'edit',
        '🗑️': 'delete',
        '🗑': 'delete',
        '🔄': 'refresh',
        '📁': 'folder',
        '⬆️': 'upload',
        '⬆': 'upload',
        '✅': 'confirm',
        '📄': 'document',
        '🏢': 'company',
        '👤': 'person',
        '👁': 'view',
    }
    _ICON_CACHE = {}
    _ASSETS_ENSURED = False
    _MIN_ICON_SIZE_BYTES = 900
    _MATERIAL_ICON_NAME = {
        'settings': 'settings',
        'dashboard': 'dashboard',
        'close': 'close',
        'next': 'arrow_forward',
        'prev': 'arrow_back',
        'finish': 'flag',
        'save': 'save',
        'new': 'add_box',
        'add': 'add',
        'edit': 'edit',
        'delete': 'delete',
        'refresh': 'refresh',
        'folder': 'folder',
        'upload': 'upload',
        'confirm': 'check_circle',
        'document': 'description',
        'company': 'business',
        'contract': 'description',
        'person': 'person',
        'view': 'visibility',
    }
    _TEXT_ICON_HINTS = [
        ('tableau de bord', 'dashboard'),
        ('generer les documents', 'document'),
        ('generation', 'document'),
        ('configuration', 'settings'),
        ('parametre', 'settings'),
        ('parametres', 'settings'),
        ('quitter', 'close'),
        ('annuler', 'close'),
        ('fermer', 'close'),
        ('suivant', 'next'),
        ('precedent', 'prev'),
        ('terminer', 'finish'),
        ('sauvegarder', 'save'),
        ('enregistrer', 'save'),
        ('nouvelle', 'new'),
        ('ajouter', 'add'),
        ('modifier', 'edit'),
        ('supprimer', 'delete'),
        ('actualiser', 'refresh'),
        ('reinitialiser', 'refresh'),
        ('consulter', 'view'),
        ('uploader', 'upload'),
        ('upload', 'upload'),
        ('proceder', 'confirm'),
        ('confirmer', 'confirm'),
        ('societe', 'company'),
        ('associe', 'person'),
        ('contrat', 'contract'),
    ]
    _TEXT_ICON_WORDS = {
        'oui': 'confirm',
        'ok': 'confirm',
        'non': 'close',
    }

    @classmethod
    def get_icon_registry(cls):
        return dict(cls._ICON_REGISTRY)

    @staticmethod
    def _normalize_text(text: str) -> str:
        normalized = unicodedata.normalize('NFKD', str(text or ''))
        normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = normalized.lower()
        normalized = re.sub(r'[^a-z0-9]+', ' ', normalized).strip()
        return normalized

    @classmethod
    def _infer_icon_key_from_text(cls, text: str) -> Optional[str]:
        norm = cls._normalize_text(text)
        if not norm:
            return None
        words = set(norm.split())
        for word, key in cls._TEXT_ICON_WORDS.items():
            if word in words:
                return key
        for hint, key in cls._TEXT_ICON_HINTS:
            if hint in norm:
                return key
        return None

    @classmethod
    def _download_icon_from_material(cls, icon_key: str, target_path: Path) -> bool:
        """Attempt to download a modern icon PNG from a material CDN."""
        icon_name = cls._MATERIAL_ICON_NAME.get(icon_key)
        if not icon_name:
            return False

        urls = [
            f"https://cdn.jsdelivr.net/npm/@material-symbols/png-400/outlined/{icon_name}.png",
            f"https://cdn.jsdelivr.net/npm/@material-symbols/png-400/rounded/{icon_name}.png",
        ]
        target_path.parent.mkdir(parents=True, exist_ok=True)
        for url in urls:
            try:
                with _urlrequest.urlopen(url, timeout=8) as response:
                    data = response.read()
                    if data and len(data) > 64:
                        target_path.write_bytes(data)
                        logger.info("Downloaded icon '%s' from %s", icon_key, url)
                        return True
            except _urlerror.URLError:
                continue
            except Exception:
                continue
        return False

    @classmethod
    def _is_low_quality_icon(cls, icon_path: Path) -> bool:
        """Heuristic: tiny png files are usually placeholder glyphs."""
        try:
            return icon_path.exists() and icon_path.stat().st_size < cls._MIN_ICON_SIZE_BYTES
        except Exception:
            return False

    @classmethod
    def _create_fallback_icon_copy(cls, target_path: Path) -> bool:
        """Fallback when download is unavailable: copy a local existing icon."""
        fallback_candidates = [
            PathManager.ICONS_DIR / 'document.png',
            PathManager.ICONS_DIR / 'view.png',
            PathManager.ICONS_DIR / 'close.png',
        ]
        for fallback in fallback_candidates:
            if fallback.exists():
                target_path.parent.mkdir(parents=True, exist_ok=True)
                try:
                    shutil.copy2(fallback, target_path)
                    logger.warning("Fallback icon copy created: %s", target_path)
                    return True
                except Exception:
                    continue
        return False

    @classmethod
    def ensure_icon_assets(cls, download_missing: bool = True):
        """Ensure icon files exist. Download missing modern icons when possible."""
        if cls._ASSETS_ENSURED:
            return
        cls._ASSETS_ENSURED = True
        try:
            PathManager.ensure_directories()
            for icon_key, filename in cls._ICON_REGISTRY.items():
                # Prefer modern icons in assets/icons/material.
                target_path = PathManager.MATERIAL_ICONS_DIR / filename
                if target_path.exists() and not cls._is_low_quality_icon(target_path):
                    continue
                downloaded = False
                if download_missing:
                    downloaded = cls._download_icon_from_material(icon_key, target_path)
                if not downloaded and not target_path.exists():
                    cls._create_fallback_icon_copy(target_path)
            # Invalidate cache after any ensure/refresh pass.
            cls._ICON_CACHE = {}
        except Exception:
            logger.debug("Failed to ensure icon assets", exc_info=True)

    @classmethod
    def register_icon(cls, icon_key: str, filename: str):
        """Register or override an icon mapping at runtime."""
        if icon_key and filename:
            cls._ICON_REGISTRY[str(icon_key).strip()] = str(filename).strip()

    @classmethod
    def _resolve_icon_path(cls, icon_key: str) -> Optional[Path]:
        """Resolve icon path from registry, direct filename, or material namespace.

        Supported formats:
        - registry key: `settings`
        - direct file in assets/icons: `my_icon.png` or `my_icon`
        - material namespace: `mi:home` / `material:home` / `material/home`
        """
        key = str(icon_key or "").strip()
        if not key:
            return None

        # Absolute path support
        try:
            direct_path = Path(key)
            if direct_path.is_absolute() and direct_path.exists():
                return direct_path
        except Exception:
            pass

        candidates: list[Path] = []
        # Prefer modern material icons over legacy icons.
        roots = [PathManager.MATERIAL_ICONS_DIR, PathManager.ICONS_DIR]

        # 1) Logical registry key
        reg_filename = cls._ICON_REGISTRY.get(key)
        if reg_filename:
            for root in roots:
                candidates.append(root / reg_filename)

        # 2) Namespace for material icon folder
        material_prefixes = ("mi:", "material:", "material/")
        if key.startswith(material_prefixes):
            if ":" in key:
                material_name = key.split(":", 1)[1].strip()
            else:
                material_name = key.split("/", 1)[1].strip()
            if material_name:
                material_rel = Path(material_name)
                if material_rel.suffix:
                    candidates.append(PathManager.MATERIAL_ICONS_DIR / material_rel)
                else:
                    candidates.append(PathManager.MATERIAL_ICONS_DIR / f"{material_name}.png")
        else:
            # 3) Direct filename in icons folders
            rel = Path(key)
            if rel.suffix:
                for root in roots:
                    candidates.append(root / rel)
            else:
                for root in roots:
                    candidates.append(root / f"{key}.png")

        for candidate in candidates:
            try:
                if candidate.exists():
                    return candidate
            except Exception:
                continue
        return None

    @classmethod
    def _extract_icon_key_and_clean_text(cls, text: str, icon_key: Optional[str]):
        raw_text = str(text or '')
        chosen_key = icon_key
        if chosen_key is None:
            for emoji, key in cls._EMOJI_ICON_KEY.items():
                if emoji in raw_text:
                    chosen_key = key
                    raw_text = raw_text.replace(emoji, ' ')
        clean_text = ' '.join(raw_text.split())
        if chosen_key is None:
            chosen_key = cls._infer_icon_key_from_text(clean_text)
        return chosen_key, clean_text

    @classmethod
    def _load_icon(cls, widget, icon_key: str):
        try:
            icon_path = cls._resolve_icon_path(icon_key)
            if not icon_path:
                return None
            if cls._is_low_quality_icon(icon_path):
                # Skip placeholder icons; caller will fallback to text/emoji.
                return None

            cache_key = (str(icon_path.resolve()), str(widget.winfo_toplevel()))
            if cache_key in cls._ICON_CACHE:
                return cls._ICON_CACHE[cache_key]

            image = tk.PhotoImage(file=str(icon_path))
            image = cls._normalize_icon_image(image, target_px=16)
            cls._ICON_CACHE[cache_key] = image
            return image
        except Exception:
            logger.debug("Failed to load icon '%s'", icon_key, exc_info=True)
            return None

    @classmethod
    def _normalize_icon_image(cls, image, target_px: int = 16):
        """Normalize icon size for more consistent rendering across buttons."""
        try:
            w = max(1, int(image.width()))
            h = max(1, int(image.height()))
            max_side = max(w, h)
            if max_side == target_px:
                return image

            scale = float(target_px) / float(max_side)
            scale = max(0.25, min(2.0, scale))
            frac = Fraction(scale).limit_denominator(8)
            if frac.numerator == frac.denominator:
                return image
            return image.zoom(frac.numerator, frac.numerator).subsample(frac.denominator, frac.denominator)
        except Exception:
            return image

    @staticmethod
    def create_button(
        parent,
        text,
        command,
        style='Secondary.TButton',
        tooltip=None,
        icon_key: Optional[str] = None,
        compound: str = 'left',
        icon_gap: int = 1,
    ):
        # Keep startup fast: don't perform network downloads while creating UI buttons.
        WidgetFactory.ensure_icon_assets(download_missing=False)
        original_text = str(text or '')
        detected_icon_key, clean_text = WidgetFactory._extract_icon_key_and_clean_text(text, icon_key)
        btn = ttk.Button(parent, text=clean_text, command=command, style=style, takefocus=False, compound=compound)

        if detected_icon_key:
            icon_image = WidgetFactory._load_icon(parent, detected_icon_key)
            if icon_image is not None:
                if clean_text:
                    clean_text = (" " * max(1, int(icon_gap))) + clean_text
                btn.configure(image=icon_image)
                btn.configure(text=clean_text)
                # Keep an instance reference to avoid Tk image GC.
                btn._icon_image = icon_image
            else:
                # Fallback when modern icon file is unavailable/unusable.
                # Keep original text so emoji/icon prefix remains visible.
                btn.configure(text=original_text)

        if tooltip:
            WidgetFactory.create_tooltip(btn, tooltip)
        return btn

    @staticmethod
    def create_tooltip(widget, text):
        def show_tooltip(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")

            label = ttk.Label(tooltip, text=text, justify='left',
                            relief='solid', borderwidth=1)
            label.pack()

            def hide_tooltip():
                tooltip.destroy()

            widget.tooltip = tooltip
            widget.bind('<Leave>', lambda e: hide_tooltip())
            tooltip.bind('<Leave>', lambda e: hide_tooltip())

        widget.bind('<Enter>', show_tooltip)

class WindowManager:
    @staticmethod
    def setup_window(root, title=""):
        # Obtenir les dimensions de l'écran
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        # Définir la taille minimale
        root.minsize(1024, 768)

        # Centrer la fenêtre
        x = (screen_width - 1200) // 2
        y = (screen_height - 800) // 2
        root.geometry(f"1200x800+{x}+{y}")

        # Configurer le titre
        if title:
            root.title(title)

        # Configurer le plein écran selon le système
        import platform
        if platform.system() == "Linux":
            root.attributes('-fullscreen', True)
        else:  # Windows
            root.state('zoomed')

        # Configurer la touche Échap pour quitter le plein écran
        root.bind('<Escape>', lambda e: root.state('normal'))

    @staticmethod
    def center_window(win, center_on_parent: bool = True):
        """Center a Toplevel or root window on the parent or on the screen.

        Args:
            win: Window instance to center.
            center_on_parent: When True, center on mapped parent if available.
                When False, always center on the current screen.
        """
        try:
            win.update_idletasks()
            # If requested and window has a parent (transient), center on parent.
            parent = getattr(win, 'master', None)
            if center_on_parent and parent and parent.winfo_ismapped():
                pw = parent.winfo_width()
                ph = parent.winfo_height()
                px = parent.winfo_rootx()
                py = parent.winfo_rooty()
                ww = win.winfo_width()
                wh = win.winfo_height()
                x = px + max(0, (pw - ww) // 2)
                y = py + max(0, (ph - wh) // 2)
            else:
                screen_w = win.winfo_screenwidth()
                screen_h = win.winfo_screenheight()
                ww = win.winfo_width()
                wh = win.winfo_height()
                x = (screen_w - ww) // 2
                y = (screen_h - wh) // 2
            win.geometry(f"+{x}+{y}")
        except Exception:
            pass

# Fonction utilitaire pour appliquer les styles
def apply_style(widget, theme_manager):
    """
    Applique le style approprié à un widget en utilisant le ThemeManager fourni.

    Args:
        widget: Le widget auquel appliquer le style
        theme_manager: L'instance de ThemeManager à utiliser
    """
    return theme_manager.apply_widget_styles(widget)

class PathManager:
    """Gestionnaire centralisé des chemins de fichiers de l'application"""
    BASE_DIR = Path(__file__).resolve().parent.parent.parent
    ASSETS_DIR = BASE_DIR / "assets"
    ICONS_DIR = ASSETS_DIR / "icons"
    MATERIAL_ICONS_DIR = ICONS_DIR / "material"
    MODELS_DIR = BASE_DIR / "Models"
    DATABASE_DIR = BASE_DIR / "databases"
    CONFIG_DIR = BASE_DIR / "config"
    ALLOWED_EXTENSIONS = {
        'models': ['.docx', '.doc'],
        'database': ['.xlsx', '.xls'],
        'config': ['.json']
    }

    @classmethod
    def ensure_directories(cls) -> None:
        """Crée les répertoires nécessaires s'ils n'existent pas"""
        try:
            for dir_path in [
                cls.MODELS_DIR,
                cls.DATABASE_DIR,
                cls.CONFIG_DIR,
                cls.ASSETS_DIR,
                cls.ICONS_DIR,
                cls.MATERIAL_ICONS_DIR,
            ]:
                dir_path.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            ErrorHandler.handle_error(e, "Erreur lors de la création des répertoires")

    @classmethod
    def validate_file(cls, filepath: Path, file_type: str) -> bool:
        """
        Valide un fichier selon son type
        Args:
            filepath: Chemin du fichier
            file_type: Type de fichier ('models', 'database', 'config')
        Returns:
            bool: True si le fichier est valide
        """
        if not filepath.exists():
            logger.warning(f"Fichier non trouvé: {filepath}")
            return False

        if not filepath.suffix.lower() in cls.ALLOWED_EXTENSIONS.get(file_type, []):
            logger.warning(f"Extension non autorisée: {filepath}")
            return False

        return True

    @classmethod
    def get_model_path(cls, filename: str) -> Path:
        """Retourne le chemin d'un fichier modèle"""
        path = cls.MODELS_DIR / filename
        if not cls.validate_file(path, 'models'):
            raise FileNotFoundError(f"Modèle invalide ou non trouvé: {filename}")
        return path

    @classmethod
    def get_database_path(cls, filename: str) -> Path:
        """Retourne le chemin d'un fichier de base de données"""
        path = cls.DATABASE_DIR / filename
        if not cls.validate_file(path, 'database'):
            raise FileNotFoundError(f"Base de données invalide ou non trouvée: {filename}")
        return path

def ensure_excel_db(path, sheets: dict):
    """Create an Excel workbook at `path` with given sheets dict (name -> columns).

    Idempotent: if the file exists, ensure missing sheets are added with headers.
    Also attempts to set basic date column formatting where column names contain 'date'.
    """
    try:
        import openpyxl
    except Exception:
        raise RuntimeError('openpyxl is required for ensure_excel_db')

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if not path.exists():
        # create new workbook
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for name, cols in sheets.items():
                pd.DataFrame(columns=cols).to_excel(writer, sheet_name=name, index=False)
        return

    # If exists, open and add missing sheets
    wb = openpyxl.load_workbook(path)
    modified = False
    for name, cols in sheets.items():
        if name not in wb.sheetnames:
            # create sheet with header row
            ws = wb.create_sheet(title=name)
            for c, col in enumerate(cols, start=1):
                ws.cell(row=1, column=c, value=col)
            modified = True
    if modified:
        wb.save(path)

    # Align canonical columns for existing sheets (add newly introduced columns).
    try:
        for name, cols in sheets.items():
            try:
                existing = _pd.read_excel(path, sheet_name=name, dtype=str)
            except Exception:
                existing = _pd.DataFrame(columns=cols)

            aligned = existing.reindex(columns=cols, fill_value='')
            if list(existing.columns) == list(cols):
                continue

            try:
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    aligned.to_excel(writer, sheet_name=name, index=False)
            except TypeError:
                wb = openpyxl.load_workbook(path)
                if name in wb.sheetnames:
                    try:
                        std = wb[name]
                        wb.remove(std)
                        wb.save(path)
                    except Exception:
                        pass
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                    aligned.to_excel(writer, sheet_name=name, index=False)
    except Exception:
        logger.exception('Failed to align workbook columns in ensure_excel_db')
    return


def _reference_fallback_map() -> dict:
    from . import constants as _const
    return {
        'SteAdresses': list(_const.SteAdresse),
        'Tribunaux': list(_const.Tribunnaux),
        'Activites': list(_const.Activities),
        'Nationalites': list(_const.Nationalite),
        'LieuxNaissance': ["Casablanca", "Rabat", "Fes", "Marrakech", "Agadir"],
    }


def _normalize_reference_values(values) -> list:
    normalized = []
    for val in values:
        text = str(val or '').strip()
        if text:
            normalized.append(text)
    return normalized


def _load_reference_sheets_cached(db_path: _Path) -> dict:
    fallback = _reference_fallback_map()
    try:
        path_obj = _Path(db_path).resolve()
    except Exception:
        path_obj = _Path(db_path)

    cache_key = str(path_obj)
    try:
        mtime_ns = path_obj.stat().st_mtime_ns if path_obj.exists() else None
    except Exception:
        mtime_ns = None

    cached = _REFERENCE_DATA_CACHE.get(cache_key)
    if cached and cached.get('mtime_ns') == mtime_ns:
        return dict(cached.get('sheets', {}))

    sheets_data = {}
    if path_obj.exists():
        try:
            excel_file = _pd.ExcelFile(path_obj)
            for sheet_name in fallback.keys():
                try:
                    df = excel_file.parse(sheet_name=sheet_name, dtype=str)
                    if df.empty or len(df.columns) == 0:
                        sheets_data[sheet_name] = list(fallback[sheet_name])
                    else:
                        col_name = df.columns[0]
                        values = _normalize_reference_values(df[col_name].fillna('').tolist())
                        sheets_data[sheet_name] = values if values else list(fallback[sheet_name])
                except Exception:
                    sheets_data[sheet_name] = list(fallback[sheet_name])
        except Exception:
            logger.exception('Failed to load reference workbook cache from %s', path_obj)

    for sheet_name, default_values in fallback.items():
        sheets_data.setdefault(sheet_name, list(default_values))

    _REFERENCE_DATA_CACHE[cache_key] = {
        'mtime_ns': mtime_ns,
        'sheets': dict(sheets_data),
    }
    return sheets_data


def get_reference_data(sheet_name: str, path: Optional[_Path] = None) -> list:
    """Load reference values from cache (invalidated by workbook mtime)."""
    try:
        if path is None:
            from . import constants as _const
            db_path = Path(__file__).resolve().parent.parent.parent / 'databases' / _const.DB_FILENAME
        else:
            db_path = _Path(path)

        sheet_data = _load_reference_sheets_cached(db_path)
        if sheet_name in sheet_data:
            return list(sheet_data[sheet_name])

        return list(_reference_fallback_map().get(sheet_name, []))
    except Exception as e:
        logger.exception('Failed to get reference data for %s: %s', sheet_name, e)
        try:
            return list(_reference_fallback_map().get(sheet_name, []))
        except Exception:
            return []


def initialize_reference_sheets(path):
    """Initialize the reference sheets (SteAdresses, Tribunaux, Activites, Nationalites, LieuxNaissance)
    with default data from constants if they are empty.

    This is called after ensure_excel_db to populate lookup tables.
    """
    try:
        from . import constants as _const
        import pandas as _pd

        path = Path(path)
        if not path.exists():
            return

        # Mapping of sheet names to data lists from constants
        ref_data = {
            'SteAdresses': _const.SteAdresse,
            'Tribunaux': _const.Tribunnaux,
            'Activites': _const.Activities,
            'Nationalites': _const.Nationalite,
            'LieuxNaissance': []  # Will be populated from default list below
        }

        # Default lieu de naissance for initial setup
        default_lieux = ["Casablanca", "Rabat", "Fes", "Marrakech", "Agadir"]
        ref_data['LieuxNaissance'] = default_lieux

        # For each reference sheet, check if empty and populate
        for sheet_name, data_list in ref_data.items():
            try:
                existing = _pd.read_excel(path, sheet_name=sheet_name, dtype=str)
                # If sheet has only header row (no data rows), populate it
                if existing.empty:
                    # Get the header name for this sheet
                    if sheet_name == 'SteAdresses':
                        col_name = 'STE_ADRESSE'
                    elif sheet_name == 'Tribunaux':
                        col_name = 'TRIBUNAL'
                    elif sheet_name == 'Activites':
                        col_name = 'ACTIVITE'
                    elif sheet_name == 'Nationalites':
                        col_name = 'NATIONALITE'
                    else:  # LieuxNaissance
                        col_name = 'LIEU_NAISSANCE'

                    # Create DataFrame from data list
                    df = _pd.DataFrame({col_name: data_list})

                    # Write to sheet (replace mode)
                    try:
                        with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except TypeError:
                        # Fallback for older pandas versions
                        from openpyxl import load_workbook
                        wb = load_workbook(path)
                        if sheet_name in wb.sheetnames:
                            std = wb[sheet_name]
                            wb.remove(std)
                            wb.save(path)
                        with _pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception:
                # If sheet doesn't exist or has errors, skip (it was just created by ensure_excel_db)
                pass
    except Exception as e:
        logger.exception('Failed to initialize reference sheets: %s', e)


def write_records_to_db(path, societe_vals: dict, associes_list: list, contrat_vals: dict):
    """Write the provided records into the Excel workbook at `path`.

    This function is idempotent and will compute incremental integer IDs
    for Societes/Associes/Contrats based on existing rows in the workbook.
    Date-like fields are converted to datetime so Excel stores them as dates.
    """
    path = _Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # import constants lazily to avoid circular imports
    from . import constants as _const

    def _normalize_contrat_columns(df):
        """Rename legacy contract columns to the canonical names and merge values."""
        if df is None or df.empty:
            return df if df is not None else _pd.DataFrame(columns=_const.contrat_headers)
        aliases = getattr(_const, 'contrat_header_aliases', {}) or {}
        out = df.copy()
        for old_col, new_col in aliases.items():
            if old_col not in out.columns:
                continue
            if new_col not in out.columns:
                out[new_col] = out[old_col]
            else:
                try:
                    old_vals = out[old_col].fillna('').astype(str).str.strip()
                    new_vals = out[new_col].fillna('').astype(str).str.strip()
                    mask = (new_vals == '') & (old_vals != '')
                    out.loc[mask, new_col] = out.loc[mask, old_col]
                except Exception:
                    pass
            try:
                out.drop(columns=[old_col], inplace=True)
            except Exception:
                pass
        return out

    # Helper to load existing sheet into DataFrame safely
    def _load_sheet_df(sheet_name):
        try:
            return _pd.read_excel(path, sheet_name=sheet_name, dtype=str)
        except Exception:
            return _pd.DataFrame(columns=_const.excel_sheets.get(sheet_name, []))

    # Helper next ID
    def _next_id(sheet_name, id_col):
        df = _load_sheet_df(sheet_name)
        if id_col in df.columns and not df.empty:
            try:
                nums = _pd.to_numeric(df[id_col], errors='coerce').dropna()
                if not nums.empty:
                    return int(nums.max()) + 1
            except Exception:
                pass
            return len(df) + 1
        return 1

    def _to_datetime(val):
        # Return pandas.Timestamp or None
        if val is None or (isinstance(val, str) and val.strip() == ''):
            return None
        try:
            return _pd.to_datetime(val, dayfirst=True, errors='coerce')
        except Exception:
            return None

    # Build rows aligned with headers
    soc_df = _pd.DataFrame(columns=_const.societe_headers)
    assoc_df = _pd.DataFrame(columns=_const.associe_headers)
    contrat_df = _pd.DataFrame(columns=_const.contrat_headers)

    # Societe
    if societe_vals:
        sid = _next_id('Societes', 'ID_SOCIETE')
        # initialize with None so columns can hold datetimes or numbers
        row: dict = {h: None for h in _const.societe_headers}
        row['ID_SOCIETE'] = sid
        # mapping from form keys to headers (best-effort)
        mapping = {
            'denomination': 'DEN_STE',
            'forme_juridique': 'FORME_JUR',
            'ice': 'ICE',
            'date_ice': 'DATE_ICE',
            'date_certificat_negatif': 'DATE_ICE',
            'date_expiration_certificat_negatif': 'DATE_EXP_CERT_NEG',
            'capital': 'CAPITAL',
            'parts_social': 'PART_SOCIAL',
            'valeur_nominale': 'VALEUR_NOMINALE',
            'adresse': 'STE_ADRESS',
            'tribunal': 'TRIBUNAL',
            'type_generation': 'TYPE_GENERATION',
            'generation_type': 'TYPE_GENERATION',
            'procedure_creation': 'PROCEDURE_CREATION',
            'creation_procedure': 'PROCEDURE_CREATION',
            'mode_depot_creation': 'MODE_DEPOT_CREATION',
            'creation_depot_mode': 'MODE_DEPOT_CREATION',
        }
        for k, h in mapping.items():
            if k in societe_vals:
                v = societe_vals.get(k)
                # try to parse dates into datetime
                if h.upper().find('DATE') >= 0:
                    dt = _to_datetime(v)
                    row[h] = dt.to_pydatetime() if dt is not None else None
                else:
                    if v is None:
                        row[h] = None
                    elif isinstance(v, bool):
                        row[h] = int(v)
                    elif isinstance(v, (int, float)):
                        row[h] = v
                    else:
                        row[h] = str(v)
        soc_df = _pd.DataFrame([row])

    # Associes
    if associes_list:
        aid = _next_id('Associes', 'ID_ASSOCIE')
        assoc_rows = []
        # Determine linked societe id
        linked_sid = soc_df['ID_SOCIETE'].iloc[0] if not soc_df.empty else ''
        for a in associes_list:
            if not isinstance(a, dict):
                continue
            r: dict = {h: None for h in _const.associe_headers}
            r['ID_ASSOCIE'] = aid
            aid += 1
            r['ID_SOCIETE'] = linked_sid
            map_a = {
                'civilite': 'CIVIL', 'prenom': 'PRENOM', 'nom': 'NOM',
                'nationalite': 'NATIONALITY', 'num_piece': 'CIN_NUM',
                'validite_piece': 'CIN_VALIDATY', 'date_naiss': 'DATE_NAISS',
                'lieu_naiss': 'LIEU_NAISS', 'adresse': 'ADRESSE',
                'telephone': 'PHONE', 'email': 'EMAIL',
                'percentage': 'PART_PERCENT', 'part_percentage': 'PART_PERCENT',
                # forms historically used either 'parts' or 'num_parts'
                'parts': 'PARTS', 'num_parts': 'PARTS',
                # form uses 'capital_detenu' variable, store it in CAPITAL_DETENU
                'capital_detenu': 'CAPITAL_DETENU',
                'est_gerant': 'IS_GERANT', 'qualite': 'QUALITY'
            }
            for k, h in map_a.items():
                if k in a:
                    v = a.get(k)
                    if h.upper().find('DATE') >= 0:
                        dt = _to_datetime(v)
                        r[h] = dt.to_pydatetime() if dt is not None else None
                    else:
                        if v is None:
                            r[h] = None
                        elif isinstance(v, bool):
                            r[h] = int(v)
                        elif isinstance(v, (int, float)):
                            r[h] = v
                        else:
                            s = str(v).strip()
                            # Try numeric conversion for parts / capital
                            if h in ('PART_PERCENT', 'PARTS', 'CAPITAL_DETENU'):
                                try:
                                    # remove spaces and parse comma/point
                                    ns = s.replace(' ', '').replace(',', '.')
                                    if '.' in ns:
                                        r[h] = float(ns)
                                    else:
                                        r[h] = int(ns)
                                except Exception:
                                    r[h] = s
                            else:
                                r[h] = s
            assoc_rows.append(r)
        if assoc_rows:
            assoc_df = _pd.DataFrame(assoc_rows)

    # Contrat
    if contrat_vals:
        cid = _next_id('Contrats', 'ID_CONTRAT')
        r: dict = {h: None for h in _const.contrat_headers}
        r['ID_CONTRAT'] = cid
        r['ID_SOCIETE'] = soc_df['ID_SOCIETE'].iloc[0] if not soc_df.empty else None
        # Map keys used by ContratForm -> canonical headers
        map_c = {
            'date_contrat': 'DATE_CONTRAT',
            'period': 'DUREE_CONTRAT_MOIS',
            'type_contrat_domiciliation': 'TYPE_CONTRAT_DOMICILIATION',
            'type_contrat_domiciliation_autre': 'TYPE_CONTRAT_DOMICILIATION_AUTRE',
            'prix_mensuel': 'LOYER_MENSUEL_TTC',
            'prix_inter': 'FRAIS_INTERMEDIAIRE_CONTRAT',
            'date_debut': 'DATE_DEBUT_CONTRAT',
            'date_fin': 'DATE_FIN_CONTRAT',
            'tva': 'TAUX_TVA_POURCENT',
            'dh_ht': 'LOYER_MENSUEL_HT',
            'montant_ht': 'MONTANT_TOTAL_HT_CONTRAT',
            'pack_demarrage_montant': 'MONTANT_PACK_DEMARRAGE_TTC',
            'pack_demarrage_loyer': 'LOYER_MENSUEL_PACK_DEMARRAGE_TTC',
            'type_renouvellement': 'TYPE_RENOUVELLEMENT',
            'tva_renouvellement': 'TAUX_TVA_RENOUVELLEMENT_POURCENT',
            'dh_ht_renouvellement': 'LOYER_MENSUEL_HT_RENOUVELLEMENT',
            'montant_ht_renouvellement': 'MONTANT_TOTAL_HT_RENOUVELLEMENT',
            'loyer_renouvellement_mensuel': 'LOYER_MENSUEL_RENOUVELLEMENT_TTC',
            'loyer_renouvellement_annuel': 'LOYER_ANNUEL_RENOUVELLEMENT_TTC',
        }
        numeric_contract_cols = {
            'LOYER_MENSUEL_TTC', 'FRAIS_INTERMEDIAIRE_CONTRAT',
            'TAUX_TVA_POURCENT', 'LOYER_MENSUEL_HT', 'MONTANT_TOTAL_HT_CONTRAT',
            'MONTANT_PACK_DEMARRAGE_TTC', 'LOYER_MENSUEL_PACK_DEMARRAGE_TTC',
            'TAUX_TVA_RENOUVELLEMENT_POURCENT', 'LOYER_MENSUEL_HT_RENOUVELLEMENT',
            'MONTANT_TOTAL_HT_RENOUVELLEMENT', 'LOYER_MENSUEL_RENOUVELLEMENT_TTC',
            'LOYER_ANNUEL_RENOUVELLEMENT_TTC',
        }
        for k, h in map_c.items():
            if k in contrat_vals:
                v = contrat_vals.get(k)
                if h.upper().find('DATE') >= 0:
                    dt = _to_datetime(v)
                    r[h] = dt.to_pydatetime() if dt is not None else None
                else:
                    if v is None:
                        r[h] = None
                    elif isinstance(v, (int, float)):
                        r[h] = v
                    else:
                        s = str(v).strip()
                        # Try to parse prices/numeric fields into numbers
                        if h in numeric_contract_cols:
                            try:
                                ns = s.replace(' ', '').replace(',', '.')
                                if '.' in ns:
                                    r[h] = float(ns)
                                else:
                                    r[h] = int(ns)
                            except Exception:
                                r[h] = s
                        else:
                            r[h] = s
        contrat_df = _pd.DataFrame([r])

    soc_df = normalize_canonical_dataframe_for_storage(soc_df)
    assoc_df = normalize_canonical_dataframe_for_storage(assoc_df)
    contrat_df = normalize_canonical_dataframe_for_storage(contrat_df)

    # Write into workbook
    # If file exists, append; otherwise create fresh workbook
    if not path.exists():
        with _pd.ExcelWriter(path, engine='openpyxl') as writer:
            if not soc_df.empty:
                soc_df.to_excel(writer, sheet_name='Societes', index=False)
            else:
                # ensure header exists
                _pd.DataFrame(columns=_const.societe_headers).to_excel(writer, sheet_name='Societes', index=False)
            if not assoc_df.empty:
                assoc_df.to_excel(writer, sheet_name='Associes', index=False)
            else:
                _pd.DataFrame(columns=_const.associe_headers).to_excel(writer, sheet_name='Associes', index=False)
            if not contrat_df.empty:
                contrat_df.to_excel(writer, sheet_name='Contrats', index=False)
            else:
                _pd.DataFrame(columns=_const.contrat_headers).to_excel(writer, sheet_name='Contrats', index=False)
        return

    # Append to existing workbook — safer approach:
    # For each canonical sheet, read existing data, align columns to canonical headers,
    # concat the new rows, then write back replacing the sheet. This avoids column
    # shifts when the existing workbook has a different header layout.
    try:
        from . import constants as _const
        sheets_to_write = [
            ("Societes", soc_df, _const.societe_headers),
            ("Associes", assoc_df, _const.associe_headers),
            ("Contrats", contrat_df, _const.contrat_headers),
        ]

        for sheet_name, new_df, headers in sheets_to_write:
            if new_df.empty:
                # still ensure the sheet exists with correct headers
                try:
                    existing = _pd.read_excel(path, sheet_name=sheet_name, dtype=str)
                except Exception:
                    existing = _pd.DataFrame(columns=headers)
                if sheet_name == 'Contrats':
                    existing = _normalize_contrat_columns(existing)
                existing = normalize_canonical_dataframe_for_storage(existing.reindex(columns=headers, fill_value=''))
                if set(existing.columns) != set(headers):
                    # rewrite sheet with canonical headers but keep existing rows aligned if possible
                    existing_aligned = existing
                    try:
                        with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            existing_aligned.to_excel(writer, sheet_name=sheet_name, index=False)
                    except TypeError:
                        # pandas older versions may not support if_sheet_exists; fallback
                        wb = load_workbook(path)
                        if sheet_name in wb.sheetnames:
                            std = wb[sheet_name]
                            wb.remove(std)
                            wb.save(path)
                        with _pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                            existing_aligned.to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            # Read existing sheet if present
            try:
                existing = _pd.read_excel(path, sheet_name=sheet_name, dtype=str)
            except Exception:
                existing = _pd.DataFrame(columns=headers)
            if sheet_name == 'Contrats':
                existing = _normalize_contrat_columns(existing)
                new_df = _normalize_contrat_columns(new_df)

            # Reindex both to canonical headers to avoid column shifts
            existing_aligned = normalize_canonical_dataframe_for_storage(
                existing.reindex(columns=headers, fill_value='')
            )
            new_aligned = normalize_canonical_dataframe_for_storage(
                new_df.reindex(columns=headers, fill_value='')
            )

            # Avoid concatenating empty or all-NA frames to prevent pandas FutureWarning
            parts = []
            try:
                if not existing_aligned.dropna(how='all').empty:
                    parts.append(existing_aligned)
            except Exception:
                # if dropna fails for any reason, fall back to using the raw frame
                if not existing_aligned.empty:
                    parts.append(existing_aligned)
            try:
                if not new_aligned.dropna(how='all').empty:
                    parts.append(new_aligned)
            except Exception:
                if not new_aligned.empty:
                    parts.append(new_aligned)

            if parts:
                combined = _pd.concat(parts, ignore_index=True)
            else:
                # both frames empty/all-NA -> produce an empty canonical DataFrame
                combined = _pd.DataFrame(columns=headers)
            combined = normalize_canonical_dataframe_for_storage(combined.reindex(columns=headers, fill_value=''))

            # Write back replacing the sheet — use if_sheet_exists='replace' when available
            try:
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    combined.to_excel(writer, sheet_name=sheet_name, index=False)
            except TypeError:
                # Fallback: remove sheet via openpyxl then append
                wb = load_workbook(path)
                if sheet_name in wb.sheetnames:
                    try:
                        std = wb[sheet_name]
                        wb.remove(std)
                        wb.save(path)
                    except Exception:
                        pass
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                    combined.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception:
        # In case of any failure fall back to the previous overlay append method
        try:
            with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if not soc_df.empty:
                    if 'Societes' in writer.book.sheetnames:
                        start = writer.book['Societes'].max_row
                        soc_df.to_excel(writer, sheet_name='Societes', index=False, header=False, startrow=start)
                    else:
                        soc_df.to_excel(writer, sheet_name='Societes', index=False)
                if not assoc_df.empty:
                    if 'Associes' in writer.book.sheetnames:
                        start = writer.book['Associes'].max_row
                        assoc_df.to_excel(writer, sheet_name='Associes', index=False, header=False, startrow=start)
                    else:
                        assoc_df.to_excel(writer, sheet_name='Associes', index=False)
                if not contrat_df.empty:
                    if 'Contrats' in writer.book.sheetnames:
                        start = writer.book['Contrats'].max_row
                        contrat_df.to_excel(writer, sheet_name='Contrats', index=False, header=False, startrow=start)
                    else:
                        contrat_df.to_excel(writer, sheet_name='Contrats', index=False)
        except Exception:
            logger.exception('Failed to append records to workbook')

    # After writing/appending, ensure date columns have a proper Excel number format
    try:
        from . import constants as _const
        wb = load_workbook(path)
        # iterate canonical sheets and apply format to columns whose header contains 'DATE'
        for sheet_name, headers in [('Societes', _const.societe_headers), ('Associes', _const.associe_headers), ('Contrats', _const.contrat_headers)]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for idx, hdr in enumerate(headers, start=1):
                if not hdr:
                    continue
                if 'DATE' in hdr.upper():
                    col_letter = get_column_letter(idx)
                    for r_i in range(2, ws.max_row + 1):
                        try:
                            cell = ws[f"{col_letter}{r_i}"]
                            cell.number_format = 'DD/MM/YYYY'
                        except Exception:
                            pass
        wb.save(path)
    except Exception:
        logger.exception('Failed to apply date number formats after writing records')

    # Try to autofit column widths for all sheets to improve readability
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            try:
                # compute max length per column (include header row)
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    col_letter = get_column_letter(idx)
                    for cell in col_cells:
                        try:
                            val = cell.value
                            if val is None:
                                l = 0
                            else:
                                l = len(str(val))
                            if l > max_len:
                                max_len = l
                        except Exception:
                            continue
                    # set width with some padding; guard minimum width
                    width = max(8, float(max_len) + 2)
                    try:
                        ws.column_dimensions[col_letter].width = width
                    except Exception:
                        pass
            except Exception:
                continue
        # Apply header style and column-specific formatting
        try:
            from openpyxl.styles import Font, Alignment, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(fill_type='solid', fgColor='DDDDDD')
            header_align = Alignment(horizontal='center', vertical='center')
            wrap_align = Alignment(wrap_text=True, vertical='top')

            for ws in wb.worksheets:
                try:
                    # header row styling
                    for cell in list(ws[1]):
                        try:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_align
                        except Exception:
                            pass
                    # per-column formatting based on header name
                    for idx in range(1, ws.max_column + 1):
                        try:
                            hdr = ws.cell(row=1, column=idx).value
                            if not hdr:
                                continue
                            h = str(hdr).upper()
                            col_letter = get_column_letter(idx)
                            # numeric columns (integers)
                            if h in ('CAPITAL', 'CAPITAL_DETENU', 'PARTS'):
                                for row_idx in range(2, ws.max_row + 1):
                                    try:
                                        c = ws[f"{col_letter}{row_idx}"]
                                        c.number_format = '#,##0'
                                        c.alignment = Alignment(horizontal='right', vertical='top')
                                    except Exception:
                                        pass
                            # pricing / currency columns
                            if h in (
                                'PRIX_CONTRAT', 'PRIX_INTERMEDIARE_CONTRAT',
                                'LOYER_MENSUEL_TTC', 'FRAIS_INTERMEDIAIRE_CONTRAT',
                                'LOYER_MENSUEL_HT', 'MONTANT_TOTAL_HT_CONTRAT',
                                'MONTANT_PACK_DEMARRAGE_TTC', 'LOYER_MENSUEL_PACK_DEMARRAGE_TTC',
                                'LOYER_MENSUEL_HT_RENOUVELLEMENT', 'MONTANT_TOTAL_HT_RENOUVELLEMENT',
                                'LOYER_MENSUEL_RENOUVELLEMENT_TTC', 'LOYER_ANNUEL_RENOUVELLEMENT_TTC'
                            ):
                                for row_idx in range(2, ws.max_row + 1):
                                    try:
                                        c = ws[f"{col_letter}{row_idx}"]
                                        c.number_format = '#,##0.00'
                                        c.alignment = Alignment(horizontal='right', vertical='top')
                                    except Exception:
                                        pass
                            # phone as text
                            if h in ('PHONE',):
                                for row_idx in range(2, ws.max_row + 1):
                                    try:
                                        c = ws[f"{col_letter}{row_idx}"]
                                        c.number_format = '@'
                                        c.alignment = Alignment(horizontal='left', vertical='top')
                                    except Exception:
                                        pass
                            # long text fields -> wrap
                            if h in ('ADRESSE', 'STE_ADRESS', 'LIEU_NAISS'):
                                for row_idx in range(2, ws.max_row + 1):
                                    try:
                                        c = ws[f"{col_letter}{row_idx}"]
                                        c.alignment = wrap_align
                                    except Exception:
                                        pass
                        except Exception:
                            # if anything fails for this header/column, continue with next
                            continue
                    # freeze header row for easier navigation
                    try:
                        ws.freeze_panes = ws['A2']
                    except Exception:
                        pass
                except Exception:
                    continue
        except Exception:
            logger.exception('Failed to apply header/column formatting')

        wb.save(path)
    except Exception:
        logger.exception('Failed to autofit column widths after writing records')


def normalize_excel_storage(path):
    """Rewrite canonical sheets so workbook values are stored in normalized user-facing format."""
    path = _Path(path)
    if not path.exists():
        return

    from . import constants as _const

    aliases = getattr(_const, 'contrat_header_aliases', {}) or {}
    sheets = (
        ('Societes', _const.societe_headers),
        ('Associes', _const.associe_headers),
        ('Contrats', _const.contrat_headers),
    )

    normalized_frames = {}
    for sheet_name, headers in sheets:
        try:
            df = _pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna('')
        except Exception:
            df = _pd.DataFrame(columns=headers)
        if sheet_name == 'Contrats':
            for old_col, new_col in aliases.items():
                if old_col not in df.columns:
                    continue
                if new_col not in df.columns:
                    df[new_col] = df[old_col]
                else:
                    try:
                        old_vals = df[old_col].fillna('').astype(str).str.strip()
                        new_vals = df[new_col].fillna('').astype(str).str.strip()
                        mask = (new_vals == '') & (old_vals != '')
                        df.loc[mask, new_col] = df.loc[mask, old_col]
                    except Exception:
                        pass
                try:
                    df.drop(columns=[old_col], inplace=True)
                except Exception:
                    pass
        normalized_frames[sheet_name] = normalize_canonical_dataframe_for_storage(
            df.reindex(columns=headers, fill_value='')
        )

    try:
        with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for sheet_name, headers in sheets:
                normalized_frames[sheet_name].reindex(columns=headers, fill_value='').to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
    except TypeError:
        wb = load_workbook(path)
        for sheet_name, _headers in sheets:
            if sheet_name in wb.sheetnames:
                try:
                    wb.remove(wb[sheet_name])
                except Exception:
                    pass
        wb.save(path)
        with _pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
            for sheet_name, headers in sheets:
                normalized_frames[sheet_name].reindex(columns=headers, fill_value='').to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )


def cleanup_old_backups(db_path, max_backups=5):
    """Keep only the most recent N backups, delete older ones.

    Args:
        db_path: Path to the main database file
        max_backups: Maximum number of backups to keep (default: 5)
    """
    try:
        db_path = _Path(db_path)
        if not db_path.exists():
            return

        # Find all backup files for this database
        backup_pattern = f"{db_path.stem}_backup_*.xlsx"
        backup_dir = db_path.parent
        backups = sorted(backup_dir.glob(backup_pattern), reverse=True)

        # Delete backups beyond the limit
        if len(backups) > max_backups:
            for backup in backups[max_backups:]:
                try:
                    backup.unlink()
                    logger.info(f"Deleted old backup: {backup.name}")
                except Exception as e:
                    logger.warning(f"Failed to delete old backup {backup.name}: {e}")
    except Exception as e:
        logger.warning(f"Error during backup cleanup: {e}")


def migrate_excel_workbook(path):
    """Detects sheets that look like canonical sheets but have different names
    and merges their rows into the canonical sheet, then removes the old sheet.
    """
    path = _Path(path)
    if not path.exists():
        return
    try:
        # Create a timestamped backup before modifying the workbook
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{path.stem}_backup_{timestamp}{path.suffix}"
        backup_path = path.parent / backup_name
        shutil.copy2(path, backup_path)
        logger.info(f"Created backup of workbook before migration: {backup_path}")
        # Clean up old backups (keep only the 5 most recent)
        cleanup_old_backups(path, max_backups=5)
        # Also add the backup path to the generation report (if present)
        try:
            tmp_out = Path(__file__).resolve().parent.parent.parent / 'tmp_out'
            tmp_out.mkdir(parents=True, exist_ok=True)

            # Try to find an HTML generation report and update its embedded JSON
            updated = False
            try:
                for html in tmp_out.glob('*_Raport_Docs_generer.html'):
                    try:
                        text = html.read_text(encoding='utf-8')
                        start_tag = '<pre id="genjson">'
                        end_tag = '</pre>'
                        sidx = text.find(start_tag)
                        if sidx != -1:
                            sidx += len(start_tag)
                            eidx = text.find(end_tag, sidx)
                            if eidx != -1:
                                raw = text[sidx:eidx]
                                try:
                                    rep = json.loads(raw)
                                except Exception:
                                    rep = {}
                                rep['migration_backup'] = str(backup_path)
                                # replace the JSON block
                                new_raw = json.dumps(rep, ensure_ascii=False, indent=2)
                                new_text = text[:sidx] + new_raw + text[eidx:]
                                html.write_text(new_text, encoding='utf-8')
                                updated = True
                                break
                    except Exception:
                        continue
            except Exception:
                updated = False

            if not updated:
                # Fallback: write a named JSON report matching the HTML convention
                # so external tooling can find it more reliably. Use a safe
                # default company string if none is available.
                try:
                    import datetime as _dt
                    gen_date = _dt.date.today().isoformat()
                    gen_time = _dt.datetime.now().strftime('%H-%M-%S')
                except Exception:
                    gen_date = 'unknown_date'
                    import time as _time
                    gen_time = _time.strftime('%H-%M-%S')
                company_clean = 'UnknownCompany'
                gen_name = f"{gen_date}_{company_clean}_Raport_Docs_generer_{gen_time}.json"
                gen_report = tmp_out / gen_name
                if gen_report.exists():
                    try:
                        with gen_report.open('r', encoding='utf-8') as gf:
                            rep = json.load(gf)
                    except Exception:
                        rep = {}
                else:
                    rep = {}
                rep['migration_backup'] = str(backup_path)
                with gen_report.open('w', encoding='utf-8') as gf:
                    json.dump(rep, gf, ensure_ascii=False, indent=2)
        except Exception:
            logger.exception('Failed to update generation report with migration backup')
    except Exception:
        logger.exception('Failed to create backup before migration; continuing without backup')
    from . import constants as _const
    wb = load_workbook(path)
    changed = False
    # Build set of canonical header sets for quick matching
    canonical = {name: set([h.upper() for h in cols]) for name, cols in _const.excel_sheets.items()}
    to_remove = []
    for sheet in list(wb.sheetnames):
        if sheet in canonical:
            continue
        try:
            df = _pd.read_excel(path, sheet_name=sheet, dtype=str)
        except Exception:
            continue
        hdrs = set([c.upper() for c in df.columns])
        # find best matching canonical sheet
        for cname, cheaders in canonical.items():
            # if overlap is large relative to the legacy sheet size, consider it a match
            # this lets small legacy extracts (few columns) be merged
            overlap = len(hdrs & cheaders)
            if overlap >= max(1, int(len(hdrs) * 0.5)):
                # append df to canonical sheet
                try:
                    existing = _pd.read_excel(path, sheet_name=cname, dtype=str)
                except Exception:
                    existing = _pd.DataFrame(columns=_const.excel_sheets.get(cname, []))
                if cname == 'Contrats':
                    aliases = getattr(_const, 'contrat_header_aliases', {}) or {}
                    for old_col, new_col in aliases.items():
                        if old_col not in df.columns:
                            continue
                        if new_col not in df.columns:
                            df[new_col] = df[old_col]
                        else:
                            try:
                                old_vals = df[old_col].fillna('').astype(str).str.strip()
                                new_vals = df[new_col].fillna('').astype(str).str.strip()
                                mask = (new_vals == '') & (old_vals != '')
                                df.loc[mask, new_col] = df.loc[mask, old_col]
                            except Exception:
                                pass
                        try:
                            df.drop(columns=[old_col], inplace=True)
                        except Exception:
                            pass
                    for old_col, new_col in aliases.items():
                        if old_col not in existing.columns:
                            continue
                        if new_col not in existing.columns:
                            existing[new_col] = existing[old_col]
                        else:
                            try:
                                old_vals = existing[old_col].fillna('').astype(str).str.strip()
                                new_vals = existing[new_col].fillna('').astype(str).str.strip()
                                mask = (new_vals == '') & (old_vals != '')
                                existing.loc[mask, new_col] = existing.loc[mask, old_col]
                            except Exception:
                                pass
                        try:
                            existing.drop(columns=[old_col], inplace=True)
                        except Exception:
                            pass
                # Align legacy df to canonical headers to ensure correct column placement
                canonical_cols = _const.excel_sheets.get(cname, [])
                df_aligned = df.reindex(columns=canonical_cols, fill_value='')
                # write back by appending
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    startrow = writer.book[cname].max_row if cname in writer.book.sheetnames else 0
                    df_aligned.to_excel(writer, sheet_name=cname, index=False, header=False, startrow=startrow)
                to_remove.append(sheet)
                changed = True
                break
    if to_remove:
        # reload workbook to ensure any appended data is present before removing old sheets
        wb = load_workbook(path)
        for s in to_remove:
            try:
                std = wb[s]
                wb.remove(std)
            except Exception:
                pass
        wb.save(path)

    # Ensure canonical sheets always use the latest canonical columns
    # (adds newly introduced columns to existing workbooks without data loss).
    try:
        for cname, headers in _const.excel_sheets.items():
            try:
                existing = _pd.read_excel(path, sheet_name=cname, dtype=str)
            except Exception:
                existing = _pd.DataFrame(columns=headers)

            if cname == 'Contrats':
                aliases = getattr(_const, 'contrat_header_aliases', {}) or {}
                for old_col, new_col in aliases.items():
                    if old_col not in existing.columns:
                        continue
                    if new_col not in existing.columns:
                        existing[new_col] = existing[old_col]
                    else:
                        try:
                            old_vals = existing[old_col].fillna('').astype(str).str.strip()
                            new_vals = existing[new_col].fillna('').astype(str).str.strip()
                            mask = (new_vals == '') & (old_vals != '')
                            existing.loc[mask, new_col] = existing.loc[mask, old_col]
                        except Exception:
                            pass
                    try:
                        existing.drop(columns=[old_col], inplace=True)
                    except Exception:
                        pass

            aligned = existing.reindex(columns=headers, fill_value='')
            if list(existing.columns) == list(headers):
                continue

            try:
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    aligned.to_excel(writer, sheet_name=cname, index=False)
            except TypeError:
                wb = load_workbook(path)
                if cname in wb.sheetnames:
                    try:
                        std = wb[cname]
                        wb.remove(std)
                        wb.save(path)
                    except Exception:
                        pass
                with _pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                    aligned.to_excel(writer, sheet_name=cname, index=False)
    except Exception:
        logger.exception('Failed to align canonical workbook columns during migration')

    try:
        normalize_excel_storage(path)
    except Exception:
        logger.exception('Failed to normalize canonical workbook values during migration')

    # Ensure Excel date columns use a readable date number format
    try:
        # Reload constants to get canonical headers
        from . import constants as _const
        wb = load_workbook(path)
        for cname, cols in _const.excel_sheets.items():
            if cname not in wb.sheetnames:
                continue
            ws = wb[cname]
            for idx, col_name in enumerate(cols, start=1):
                if not col_name:
                    continue
                if 'DATE' in col_name.upper():
                    col_letter = get_column_letter(idx)
                    # Apply number_format to all data cells in this column
                    for row in range(2, ws.max_row + 1):
                        try:
                            cell = ws[f"{col_letter}{row}"]
                            # set format regardless; Excel/openpyxl will ignore non-dates
                            cell.number_format = 'DD/MM/YYYY'
                        except Exception:
                            pass
        wb.save(path)
    except Exception:
        logger.exception('Failed to apply date number formats after migration')
    # Autofit columns for all sheets after migration adjustments
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            try:
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    col_letter = get_column_letter(idx)
                    for cell in col_cells:
                        try:
                            val = cell.value
                            if val is None:
                                l = 0
                            else:
                                l = len(str(val))
                            if l > max_len:
                                max_len = l
                        except Exception:
                            continue
                    width = max(8, float(max_len) + 2)
                    try:
                        ws.column_dimensions[col_letter].width = width
                    except Exception:
                        pass
            except Exception:
                continue
        wb.save(path)
    except Exception:
        logger.exception('Failed to autofit column widths after migration')


def societe_exists(name: str, path: Optional[_Path] = None) -> bool:
    """Check whether a société with the given name exists in the Excel database.

    Args:
        name: Company name to search for (case-insensitive, trimmed)
        path: Optional path to the Excel workbook. If not provided, uses
              the default databases path and filename from package constants.

    Returns:
        True if a matching company name is present in the 'Societes' sheet,
        False otherwise.
    """
    try:
        from . import constants as _const
        # Default database path
        if path is None:
            db_path = Path(__file__).resolve().parent.parent.parent / 'databases' / _const.DB_FILENAME
        else:
            db_path = _Path(path)

        if not db_path.exists():
            return False

        try:
            df = _pd.read_excel(db_path, sheet_name='Societes', dtype=str)
        except Exception:
            return False

        if 'DEN_STE' not in df.columns:
            # fallback: try to detect any column that looks like a company name
            candidates = [c for c in df.columns if 'DEN' in str(c).upper() or 'STE' in str(c).upper() or 'NAME' in str(c).upper()]
            if not candidates:
                return False
            col = candidates[0]
        else:
            col = 'DEN_STE'

        target = (str(name or '')).strip().lower()
        if not target:
            return False

        # check for exact matches (case-insensitive) or trimmed contains
        for val in df[col].fillna('').astype(str):
            if val.strip().lower() == target:
                return True
        return False
    except Exception:
        logger.exception('societe_exists check failed')
        return False

