import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from docx import Document
import PySimpleGUI as sg 
from docxtpl import DocxTemplate
import openpyxl
from docx2pdf import convert
import pandas as pd


# Define headers for the Excel file
excel_headers = [
    "DEN_STE",
    "FORME_JUR",
    "ICE",
    "CAPITAL",
    "DATE_ICE",
    "CIVIL",
    "PRENOM",
    "NOM",
    "DATE_NAISS",
    "LIEU_NAISS",
    "GERANT_ADRESS",
    "CIN_NUM",
    "CIN_VALIDATY",
    "GERANT_QUALITY",
    "GERANT_PHONE",
    "GERANT_EMAIL",
    "TRIBUNAL",
    "PERIOD_DOMCIL",
    "DOM_DATEDEB",
    "DOM_DATEFIN",
    "DTAE_CONTRAT",
    "PRIX_CONTRAT",
    "PRIX_INTERMEDIARE_CONTRAT"

]

# EXCEL_FILE = 'DataBase_domiciliation.xlsx'
# df = pd.read_excel(EXCEL_FILE)

# pouvoir_sign_path = Path(__file__).parent / "My_Pouvoir_Signature.docx"
# docPouvoir = DocxTemplate(pouvoir_sign_path)

# Path of the file containing the template
annonce_journal_path = Path(__file__).parent / "Models" / "My_Annonce_Journal.docx"
docAnnonceJournal = DocxTemplate(annonce_journal_path)

attest_domicil_path = Path(__file__).parent / "Models" / "My_Attest_domiciliation.docx"
docAttestDomicil = DocxTemplate(attest_domicil_path)

contrat_domicil_path = Path(__file__).parent / "Models" / "My_Contrat_domiciliation.docx"
docContratDomicil = DocxTemplate(contrat_domicil_path)

contrat_domicil_misa_path = Path(__file__).parent / "Models" / "My_Contrat_domiciliation_Misa.docx"
docContratDomicilMisa = DocxTemplate(contrat_domicil_misa_path)

decl_imm_rc_path = Path(__file__).parent / "Models" / "My_Décl_Imm_Rc.docx"
docDeclImmRc = DocxTemplate(decl_imm_rc_path)

depot_legal_rc_path = Path(__file__).parent / "Models" / "My_Dépot_Légal.docx"
docDepotLegalRc = DocxTemplate(depot_legal_rc_path)

statuts_path = Path(__file__).parent / "Models" / "My_Statuts_SARL_AU.docx"
docStatuts = DocxTemplate(statuts_path)

# Identifier Dictionary of input
DenSte = ["ASTRAPIA", "SAOUZ", "F 4", "OLA MOVING", "LOHACOM", "SKY NEST", "SKY MA", "MAROFLEET"]
Civility = ["Monsieur","Madame"]
Formjur = ["SARL AU", "SARL", "SA"]
Nbmois = ["06", "12","15", "24"]
Capital = ["10 000", "50 000", "100 000"]
PartsSocial = ["100", "200", "500", "1000"]
Nationalite = ["Marocaine", "Cameronnie"]
SteAdresse = ["46 BD ZERKTOUNI ETG 2 APPT 6 CASABLANCA", "56  BOULEVARD MOULAY YOUSSEF 3EME ETAGE APPT 14, CASABLANCA", "96 BD D'ANFA ETAGE 9 APPT N° 91 RESIDENCE LE PRINTEMPS D'ANFA", "61 Av Lalla Yacout Angle Mustapha El Maani 2eme Etage N°62 Centre Riad",  "Bd Zerktouni Et Angle Rue Ibn Al Moualim N° 4, Etage 2, Appt N°10"]
Tribunnaux = ["Casablanca", "Berrechid", "Mohammedia"]
QualityGerant = ["Associé Gérant", "Associé Unique Gérant", "Associé"]
Rcn = ["495551", "516421", "516075", "470819", "596755"]
Ifn = ["50136363", "50551160","50537219", "45920770"]
Icen = ["002737787000056", "002896698000060", "002896781000050", "002590934000018"]
Activities = ["Travaux Divers ou de Construction", "Marchand effectuant Import Export", "Négociant", "Conseil de Gestion"]
PrenGerant = ["Abdeljalil"]
NomGerant = ["RIHANA"]
CinGerant = ["BH584028"]
PrenSign = ["Mohamed Amine"]
NomSign = ["EL HOUMRI"]
DateNaissSign = ["28-06-1985"]
LieuNaissSingn = ["Casablanca"]
AdressSign = ["DR CHHOUBA N°225 OULFA Casablanca"]
CinSign = ["BE785182"]

# get the date of today
today = datetime.today()
# for the name of file like 2023-08-29
date_inverser = today.strftime("%Y_%m_%d")
dateActe = today.strftime("%d-%m-%Y")
Year1 = int(today.year) + 1
Month1 = int(today.month)
Day1 = int(today.day) - 1
datefinContrat = f"0{Day1}-0{Month1}-{Year1}"

    # [sg.Text("IF N"), sg.DD(Ifn,key="IF", size=(43,8))],
    # [sg.Text("TP N"), sg.Input(key="TP", do_not_clear=False)],
    # [sg.Text("RC N"), sg.DD(Rcn, key="NUM_RC", size=(43,8))],
    # [sg.Text("Numero Dépot"), sg.DD(Rcn, key="NUM_DEPOT_RC", size=(43,8))],
    # [sg.Text("Date Dépot"), sg.DD(Rcn, key="DATE_DEPOT_RC", size=(43,8))],

# Define the layout of the window, the layout is divided into three parts (InfosSte, InfosGerant, InfosContrat)
InfosSte = [
    [sg.Text("Dénomination Société")], 
    [sg.DD(DenSte, key="DEN_STE", size=(43,8))],
    [sg.Text("Forme Juridique")], 
    [sg.DD(Formjur, key="FORME_JUR", size=(43,8))],
    [sg.Text("Ice")], 
    [sg.Input(key="ICE", size=(45,8))],
    [sg.CalendarButton("Entrer Date Ice", target="DATE_ICE", format="%d-%m-%Y", size=(39, 1))],
    [sg.Input(key="DATE_ICE", size=(45,8))],   
    [sg.Text("Capital")], 
    [sg.DD(Capital, key="CAPITAL", size=(43,8))],
    [sg.Text("Part Social")], 
    [sg.DD(PartsSocial, key="PART_SOCIAL", size=(43,8))],
    [sg.Text("Adresse Société")], 
    [sg.DD(SteAdresse, key="STE_ADRESS", size=(43,8))],
    [sg.Text("Trubinal")], 
    [sg.DD(Tribunnaux, key="TRIBUNAL", size=(43,8))],
    [sg.Text("Activité 1")], 
    [sg.DD(Activities, key="ACTIVITY1", size=(43,8))],
    [sg.Text("Activité 2")], 
    [sg.DD(Activities, key="ACTIVITY2", size=(43,8))],
    [sg.Text("Activité 3")], 
    [sg.DropDown(Activities, key="ACTIVITY3", size=(43,8))],
    [sg.Text("Activité 4")], 
    [sg.DD(Activities, key="ACTIVITY4", size=(43,8))],
    [sg.Text("Activité 5")], 
    [sg.DD(Activities, key="ACTIVITY5", size=(43,8))],
    [sg.Text("Activité 6")], 
    [sg.DD(Activities, key="ACTIVITY6", size=(43,8))],
    ]

InfosGerant = [
    [sg.Text("Civilité Gérant")], 
    [sg.DD(Civility, key="CIVIL", size=(43,8))],
    [sg.Text("Prénom Gérant")], 
    [sg.Input(key="PRENOM", size=(43,8))],
    [sg.Text("Nom Gérant")], 
    [sg.Input(key="NOM", size=(43,8))],
    [sg.Text("N CIN Gérant")], 
    [sg.Input(key="CIN_NUM", size=(43,8))],
    [sg.CalendarButton("Entrer Validité du CIN", target="CIN_VALIDATY", format="%d-%m-%Y", size=(39, 1))],
    [sg.Input(key="CIN_VALIDATY", size=(43,8))],
    [sg.Text("Qualité du Gérant")], 
    [sg.DD(QualityGerant, key="GERANT_QUALITY", size=(43,8))],
    [sg.CalendarButton("Entrer Date Naiss", target="DATE_NAISS", format="%d-%m-%Y", size=(39, 1))],
    [sg.Input(key="DATE_NAISS", size=(43,8))],
    [sg.Text("Lieux Naiss Gérant")], 
    [sg.Input(key="LIEU_NAISS", size=(43,8))],
    [sg.Text("Adresse Gérant")], 
    [sg.Input(key="GERANT_ADRESS", size=(43,8))],
    [sg.Text("Nationalité Gérant")], 
    [sg.DD(Nationalite, key="NATIONALITY", size=(43,8))],
    [sg.Text("Téléphone Gérant")], 
    [sg.Input(key="GERANT_PHONE", size=(43,8))],
    [sg.Text("Email Gérant")], 
    [sg.Input(key="GERANT_EMAIL", size=(43,8))],
    ]

InfosContrat = [
    [sg.CalendarButton("Entrer Date Contrat", target="DTAE_CONTRAT", format="%d-%m-%Y", size=(39, 1))],
    [sg.InputText(key="DTAE_CONTRAT")],
    [sg.Text("Période de Contrat")],
    [sg.DD(Nbmois, key="PERIOD_DOMCIL")],
    [sg.Text("Prix mensuel de Contrat"), sg.Input(key="PRIX_CONTRAT", size=(23, 1))],
    [sg.Text("Prix intermediare Contrat"), sg.Input(key="PRIX_INTERMEDIARE_CONTRAT", size=(23, 1))],
    # [sg.Input(key="PRIX_CONTRAT")],    
    [sg.CalendarButton("Entrer Débit de Contrat", target="DOM_DATEDEB", format="%d-%m-%Y", size=(39, 1))],
    [sg.InputText(key="DOM_DATEDEB")],
    [sg.CalendarButton("Entrer Fin de Contrat", target="DOM_DATEFIN", format="%d-%m-%Y", size=(39, 1))],
    [sg.InputText(key="DOM_DATEFIN")],
    ]

# Col_1 = [sg.Frame("Infos Ste", InfosSte, element_justification="left")]
# Col_2 = [sg.Frame("Infos Gérant", InfosGerant, element_justification="left")]
# Col_3 = [sg.Frame("Infos Contrat", InfosContrat, element_justification="left")]

layout = [
    [sg.Column(InfosSte), sg.Column(InfosGerant), sg.Column(InfosContrat)],
    # Col_1,Col_2,Col_3,
    [sg.Button("Générer Les Docx Lemedaj"), sg.Button("Save As Pdf"), sg.Exit()]
    ]

window = sg.Window("Générer Mes Docs", layout, element_justification="right")

# Function to insert data into Excel file
def insert_data_to_excel(values):
    try:
        # Open the Excel file or create if it doesn't exist
        excel_file_path = 'DataBase_domiciliation.xlsx'
        if not os.path.exists(excel_file_path):
            workbook = openpyxl.Workbook()
            workbook.save(excel_file_path)

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Check if the "Data" sheet exists, create it if not
        if "DataBaseDom" not in workbook.sheetnames:
            workbook.create_sheet("DataBaseDom")

        # Select the "Data" sheet
        sheet = workbook["DataBaseDom"]

        # Add headers if the file is empty
        if sheet.max_row == 1:
            sheet.append(excel_headers)

        # Append data to the Excel file
        row_data = [
            values["DEN_STE"],
            values["FORME_JUR"],
            values["ICE"],
            values["CAPITAL"],
            values["PART_SOCIAL"],
            values["NATIONALITY"],
            values["STE_ADRESS"],
            values["DATE_ICE"],
            values["CIVIL"],
            values["PRENOM"],
            values["NOM"],
            values["DATE_NAISS"],
            values["LIEU_NAISS"],
            values["GERANT_ADRESS"],
            values["CIN_NUM"],
            values["CIN_VALIDATY"],
            values["GERANT_QUALITY"],
            values["GERANT_PHONE"],
            values["GERANT_EMAIL"],
            values["TRIBUNAL"],
            values["PERIOD_DOMCIL"],
            values["DOM_DATEDEB"],
            values["DOM_DATEFIN"],
            values["DTAE_CONTRAT"],
            values["PRIX_CONTRAT"],
            values["PRIX_INTERMEDIARE_CONTRAT"]
        ]
        sheet.append(row_data)

        # Save changes to the Excel file
        workbook.save(excel_file_path)
    except Exception as e:
        sg.popup_error(f"Une erreur s'est produite lors de l'insertion des données dans le fichier Excel : {str(e)}")

# Event loop
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == "Exit":
        break
    # Event loop for Gébérer Mes Docs 
    elif event == "Générer Les Docx Lemedaj":

        folder_name = f"{date_inverser}_DCons_{values['DEN_STE']}"
        if folder_name.strip() != '':
            folder_path = os.path.join(os.getcwd(), folder_name)
            try:
                os.mkdir(folder_path)
                sg.popup(f"Le dossier '{folder_name}' a été créé avec succès!")

                # Insert data into the Excel file
                insert_data_to_excel(values)

                
                sg.popup(f"Les Fichiers DOCX créés avec succès dans '{folder_name}' !")
                
                # Empty the InputText fields
                for key in values.keys():
                    window[key].update('')
            except FileExistsError:
                sg.popup_error(f"Le dossier '{folder_name}' existe déjà !")
        else:
            sg.popup_error("Veuillez entrer un nom de dossier !")


        # Add calculate fields to our dict
        # values["DTAE_CONTRAT"] = today.strftime("%d-%m-%Y")

        docAnnonceJournal.render(values)
        outputAnnonce_path = os.path.join(folder_path, f"{date_inverser}_Annonce_légale_{values['DEN_STE']}.docx")
        docAnnonceJournal.save(outputAnnonce_path)

        docAttestDomicil.render(values)
        outputattest_domicil_path = os.path.join(folder_path, f"{date_inverser}_Attestation_Domiciliation_{values['DEN_STE']}.docx")
        docAttestDomicil.save(outputattest_domicil_path)

        docContratDomicil.render(values)
        outputcontrat_domicil_path = os.path.join(folder_path, f"{date_inverser}_Contrat_Domiciliation_{values['DEN_STE']}.docx")
        docContratDomicil.save(outputcontrat_domicil_path)

        docDeclImmRc.render(values)
        outputdecl_imm_rc_path = os.path.join(folder_path, f"{date_inverser}_Déclaration_Imm_Rc_{values['DEN_STE']}.docx")
        docDeclImmRc.save(outputdecl_imm_rc_path)

        docDepotLegalRc.render(values)
        outputdepot_legal_rc_path = os.path.join(folder_path, f"{date_inverser}_Dépot_légal_{values['DEN_STE']}.docx")
        docDepotLegalRc.save(outputdepot_legal_rc_path)

        docStatuts.render(values)
        outuputstatuts_path = os.path.join(folder_path, f"{date_inverser}_Statuts_{values['DEN_STE']}.docx")
        docStatuts.save(outuputstatuts_path)

        # Save Docs As Pdf files
        # outputAnnonce_pdfpath = os.path.join(folder_path, f"{date_inverser}_Annonce_légale_{values['DEN_STE']}.pdf")
        # convert(outputAnnonce_path, outputAnnonce_pdfpath)

        # outputattest_domicil_pdfpath = os.path.join(folder_path, f"{date_inverser}_Attestation_Domiciliation_{values['DEN_STE']}.pdf")
        # convert(outputattest_domicil_path, outputattest_domicil_pdfpath)

        outputcontrat_domicil_pdfpath = os.path.join(folder_path, f"{date_inverser}_Contrat_Domiciliation_{values['DEN_STE']}.pdf")
        convert(outputcontrat_domicil_path, outputcontrat_domicil_pdfpath)

        # outputdecl_imm_rc_pdfpath = os.path.join(folder_path, f"{date_inverser}_Déclaration_Imm_Rc_{values['DEN_STE']}.pdf")
        # convert(outputdecl_imm_rc_path, outputdecl_imm_rc_pdfpath)

        # outputdepot_legal_rc_pdfpath = os.path.join(folder_path, f"{date_inverser}_Dépot_légal_{values['DEN_STE']}.pdf")
        # convert(outputdepot_legal_rc_path, outputdepot_legal_rc_pdfpath)

        # outuputstatuts_pdfpath = os.path.join(folder_path, f"{date_inverser}_Statuts_{values['DEN_STE']}.pdf")
        # convert(outuputstatuts_path, outuputstatuts_pdfpath)


        sg.popup("Les Docs Sont bien sauvgarder", f"Les Fichiers Sont Sauvegarder Ici : {folder_name}")

window.close()