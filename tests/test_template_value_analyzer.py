from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import unittest

from src.utils.template_value_analyzer import (
    analyze_templates,
    export_analysis_rows,
    extract_template_variables,
    filter_analysis_rows,
)


def _write_fake_docx(path: Path, body_text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body><w:p><w:r><w:t>"
        f"{body_text}"
        "</w:t></w:r></w:p></w:body></w:document>"
    )
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "")
        zf.writestr("word/document.xml", document_xml)


class TestTemplateValueAnalyzer(unittest.TestCase):
    def test_extract_template_variables_counts_occurrences(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / "Models" / "tpl_a.docx"
            _write_fake_docx(template, "{{ DEN_STE }} - {{ DEN_STE }} - {{ capital|default('') }}")
            counts = extract_template_variables(template)
            self.assertEqual(2, counts["DEN_STE"])
            self.assertEqual(1, counts["capital"])

    def test_analyze_templates_builds_global_and_detail_stats(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            models_dir = Path(tmp) / "Models"
            _write_fake_docx(models_dir / "tpl_a.docx", "{{ DEN_STE }} {{ DEN_STE }} {{ capital }}")
            _write_fake_docx(models_dir / "sous" / "tpl_b.docx", "{{ DATE_CONTRAT }} {{ unknown_var }}")

            data = analyze_templates(models_dir, context_keys={"DEN_STE", "capital", "DATE_CONTRAT"})
            summary = data["summary"]

            self.assertEqual(2, summary["total_templates"])
            self.assertEqual(5, summary["total_variable_occurrences"])
            self.assertEqual(4, summary["total_distinct_variables"])
            self.assertEqual(3, summary["covered_variables"])
            self.assertEqual(1, summary["uncovered_variables"])

            rows = {row["variable"]: row for row in data["variables"]}
            self.assertEqual(2, rows["DEN_STE"]["occurrences"])
            self.assertEqual("couvert", rows["DEN_STE"]["coverage"])
            self.assertEqual("contrat", rows["DATE_CONTRAT"]["section"])
            self.assertEqual("non couvert", rows["unknown_var"]["coverage"])
            self.assertEqual({"tpl_a.docx", "tpl_b.docx"}, {row["template"] for row in data["details"]})

    def test_filter_analysis_rows_supports_text_template_section_and_coverage(self):
        rows = [
            {
                "template": "a.docx",
                "variable": "DEN_STE",
                "occurrences": 2,
                "section": "societe",
                "coverage": "couvert",
                "templates": ["a.docx"],
            },
            {
                "template": "b.docx",
                "variable": "unknown_var",
                "occurrences": 1,
                "section": "autre",
                "coverage": "non couvert",
                "templates": ["b.docx"],
            },
        ]

        filtered = filter_analysis_rows(rows, search_text="den")
        self.assertEqual(1, len(filtered))
        self.assertEqual("DEN_STE", filtered[0]["variable"])

        filtered = filter_analysis_rows(rows, template_name="b.docx")
        self.assertEqual(1, len(filtered))
        self.assertEqual("b.docx", filtered[0]["template"])

        filtered = filter_analysis_rows(rows, section="societe")
        self.assertEqual(1, len(filtered))
        self.assertEqual("societe", filtered[0]["section"])

        filtered = filter_analysis_rows(rows, coverage="non couvert")
        self.assertEqual(1, len(filtered))
        self.assertEqual("non couvert", filtered[0]["coverage"])

    def test_export_analysis_rows_to_csv_and_xlsx(self):
        import tempfile
        rows = [
            {
                "template": "tpl_a.docx",
                "template_path": "C:/tmp/Models/tpl_a.docx",
                "variable": "DEN_STE",
                "occurrences": 2,
                "section": "societe",
                "coverage": "couvert",
                "templates": ["tpl_a.docx"],
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            out_csv = Path(tmp) / "export.csv"
            out_xlsx = Path(tmp) / "export.xlsx"

            export_analysis_rows(rows, out_csv, columns=("template", "variable", "occurrences", "templates"))
            self.assertTrue(out_csv.exists())
            csv_text = out_csv.read_text(encoding="utf-8-sig")
            self.assertIn("template,variable,occurrences,templates", csv_text)
            self.assertIn("tpl_a.docx,DEN_STE,2,tpl_a.docx", csv_text)

            export_analysis_rows(rows, out_xlsx, columns=("template", "variable", "occurrences"))
            self.assertTrue(out_xlsx.exists())

            from openpyxl import load_workbook

            wb = load_workbook(out_xlsx)
            ws = wb.active
            self.assertEqual("template", ws["A1"].value)
            self.assertEqual("DEN_STE", ws["B2"].value)


if __name__ == "__main__":
    unittest.main()
