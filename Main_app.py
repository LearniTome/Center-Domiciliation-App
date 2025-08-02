import os
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import Calendar
from docxtpl import DocxTemplate
import openpyxl
from docx2pdf import convert
import pandas as pd

from utils import PathManager
from constants import (
    societe_headers, associe_headers, contrat_headers,
    excel_sheets, DenSte, Civility, Formjur, Nbmois,
    Capital, PartsSocial, Nationalite, SteAdresse,
    Tribunnaux, QualityGerant, Activities
)

# Path of the template files
PathManager.ensure_directories()

docAnnonceJournal = DocxTemplate(PathManager.get_model_path("My_Annonce_Journal.docx"))
docAttestDomicil = DocxTemplate(PathManager.get_model_path("My_Attest_domiciliation.docx"))
docContratDomicil = DocxTemplate(PathManager.get_model_path("My_Contrat_domiciliation.docx"))
docDeclImmRc = DocxTemplate(PathManager.get_model_path("My_Décl_Imm_Rc.docx"))
docDepotLegalRc = DocxTemplate(PathManager.get_model_path("My_Dépot_Légal.docx"))
docStatuts = DocxTemplate(PathManager.get_model_path("My_Statuts_SARL_AU.docx"))

# get the date of today
today = datetime.today()
date_inverser = today.strftime("%Y_%m_%d")

from utils import ThemeManager, WidgetFactory, ToolTip, apply_style

class DomiciliationApp:
    def __init__(self, root):
        self.root = root
        self.values = {}

        # Configuration du thème et de la fenêtre
        if isinstance(root, tk.Tk):
            self.root.title("Genérateurs Docs Juridiques")
            root_window = self.root
        else:
            root_window = root.winfo_toplevel()

        # Initialiser le gestionnaire de thème
        self.theme_manager = ThemeManager(root_window)
        self.style = self.theme_manager.style

        # Configuration des styles personnalisés
        self.setup_styles()

        # Helper pour appliquer les styles
        self.style_widget = lambda widget: apply_style(widget, self.theme_manager)

        # Liste pour stocker les informations des associés
        self.associes_list = []

        self.setup_gui()

    def setup_styles(self):
        """Configure les styles personnalisés pour l'application"""
        # Style des onglets
        self.style.configure('TNotebook.Tab',
                           padding=(15, 5),
                           font=('Segoe UI', 10, 'bold'))

        # Style des sections
        self.style.configure('Section.TFrame',
                           padding=10,
                           relief='groove')

        # Style des labels de section
        self.style.configure('SectionTitle.TLabel',
                           font=('Segoe UI', 11, 'bold'),
                           padding=(0, 10))

        # Style des champs de formulaire
        self.style.configure('Field.TFrame',
                           padding=5)

        # Style des boutons d'action
        self.style.configure('Action.TButton',
                           padding=(10, 5),
                           font=('Segoe UI', 9))

    def setup_gui(self):
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True)

        # Create frames for each section
        self.frame_ste = ttk.Frame(self.notebook)
        self.frame_associes = ttk.Frame(self.notebook)
        self.frame_contrat = ttk.Frame(self.notebook)

        # Add frames to notebook
        self.notebook.add(self.frame_ste, text='Infos Société')
        self.notebook.add(self.frame_associes, text='Associés')
        self.notebook.add(self.frame_contrat, text='Infos Contrat')

        # Initialize all frames
        self.initialize_ste_section()  # Initialize société section
        self.initialize_associes_frame() # Initialize associés section
        self.initialize_contrat_frame() # Initialize contrat section

        # Create buttons frame at the bottom
        self.setup_buttons()

    def setup_buttons(self):
        # Conteneur principal pour les boutons
        self.buttons_frame = ttk.Frame(self.root)
        self.buttons_frame.pack(pady=15, side=tk.BOTTOM, fill=tk.X, padx=20)

        # Configuration de la grille pour l'alignement
        self.buttons_frame.grid_columnconfigure(2, weight=1)  # Espace flexible au milieu

        # Boutons principaux (à gauche)
        WidgetFactory.create_button(
            self.buttons_frame,
            text="📄 Documents Word",
            command=self.generer_docs,
            style='Action.TButton',
            tooltip="Générer tous les documents au format Word"
        ).grid(row=0, column=0, padx=5)

        WidgetFactory.create_button(
            self.buttons_frame,
            text="📑 Word et PDF",
            command=self.save_as_pdf,
            style='Action.TButton',
            tooltip="Générer tous les documents en Word et PDF"
        ).grid(row=0, column=1, padx=5)

        # Boutons de contrôle (à droite)
        WidgetFactory.create_button(
            self.buttons_frame,
            text="🆕 Nouvelle",
            command=self.clear_form,
            style='Secondary.TButton',
            tooltip="Effacer tous les champs pour une nouvelle société"
        ).grid(row=0, column=3, padx=5)

        WidgetFactory.create_button(
            self.buttons_frame,
            text="💾 Sauvegarder",
            command=self.save_to_database,
            style='Secondary.TButton',
            tooltip="Sauvegarder les informations dans la base de données"
        ).grid(row=0, column=4, padx=5)

        WidgetFactory.create_button(
            self.buttons_frame,
            text="❌ Quitter",
            command=self.root.quit,
            style='Secondary.TButton',
            tooltip="Fermer l'application"
        ).grid(row=0, column=5, padx=5)

    def initialize_ste_section(self):
        row = 0
        ttk.Label(self.frame_ste, text="Dénomination Société").grid(row=row, column=0, padx=5, pady=5)
        self.den_ste_var = tk.StringVar()
        self.den_ste_combo = ttk.Combobox(self.frame_ste, textvariable=self.den_ste_var, values=DenSte, width=40)
        self.den_ste_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        ttk.Label(self.frame_ste, text="Forme Juridique").grid(row=row, column=0, padx=5, pady=5)
        self.forme_jur_var = tk.StringVar()
        self.forme_jur_combo = ttk.Combobox(self.frame_ste, textvariable=self.forme_jur_var, values=Formjur, width=40)
        self.forme_jur_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        ttk.Label(self.frame_ste, text="Ice").grid(row=row, column=0, padx=5, pady=5)
        self.ice_var = tk.StringVar()
        ttk.Entry(self.frame_ste, textvariable=self.ice_var, width=42).grid(row=row, column=1, padx=5, pady=5)
        row += 1

        ttk.Label(self.frame_ste, text="Date Ice").grid(row=row, column=0, padx=5, pady=5)
        self.date_ice_var = tk.StringVar()
        self.date_ice_entry = ttk.Entry(self.frame_ste, textvariable=self.date_ice_var, width=42)
        self.date_ice_entry.grid(row=row, column=1, padx=5, pady=5)
        ttk.Button(self.frame_ste, text="Choisir Date",
                  command=lambda: self.show_calendar(self.date_ice_var)).grid(row=row, column=2, padx=5, pady=5)
        row += 1

        # Capital
        ttk.Label(self.frame_ste, text="Capital").grid(row=row, column=0, padx=5, pady=5)
        self.capital_var = tk.StringVar()
        self.capital_combo = ttk.Combobox(self.frame_ste, textvariable=self.capital_var, values=Capital, width=40)
        self.capital_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Part Social
        ttk.Label(self.frame_ste, text="Part Social").grid(row=row, column=0, padx=5, pady=5)
        self.parts_social_var = tk.StringVar()
        self.parts_social_combo = ttk.Combobox(self.frame_ste, textvariable=self.parts_social_var,
                                             values=PartsSocial, width=40)
        self.parts_social_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Adresse Société
        ttk.Label(self.frame_ste, text="Adresse Société").grid(row=row, column=0, padx=5, pady=5)
        self.ste_adress_var = tk.StringVar()
        self.ste_adress_combo = ttk.Combobox(self.frame_ste, textvariable=self.ste_adress_var,
                                           values=SteAdresse, width=40)
        self.ste_adress_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Tribunal
        ttk.Label(self.frame_ste, text="Tribunal").grid(row=row, column=0, padx=5, pady=5)
        self.tribunal_var = tk.StringVar()
        self.tribunal_combo = ttk.Combobox(self.frame_ste, textvariable=self.tribunal_var,
                                        values=Tribunnaux, width=40)
        self.tribunal_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Activities Section avec titre stylisé
        activities_label = ttk.Label(self.frame_ste, text="🏢 Activités de la Société",
                                   font=('Segoe UI', 10, 'bold'))
        activities_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(15,5))

        # Frame décoratif pour les activités
        activities_container = ttk.Frame(self.frame_ste, style='Card.TFrame')
        activities_container.grid(row=row+1, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

        # Frame pour contenir la liste des activités
        self.activities_frame = ttk.Frame(activities_container)
        self.activities_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Canvas et scrollbar pour la liste des activités
        self.activities_canvas = tk.Canvas(self.activities_frame, height=150,
                                         background=self.theme_manager.colors['bg'])
        scrollbar = ttk.Scrollbar(self.activities_frame, orient="vertical",
                                command=self.activities_canvas.yview)
        self.activities_canvas.configure(yscrollcommand=scrollbar.set)

        # Frame intérieur pour les activités
        self.activities_inner_frame = ttk.Frame(self.activities_canvas)
        self.activities_inner_frame.bind("<Configure>",
            lambda e: self.activities_canvas.configure(scrollregion=self.activities_canvas.bbox("all")))

        # Placer les widgets
        scrollbar.pack(side="right", fill="y")
        self.activities_canvas.pack(side="left", fill="both", expand=True)
        self.activities_canvas.create_window((0, 0), window=self.activities_inner_frame, anchor="nw")

        # Liste pour stocker les variables des activités
        self.activity_vars = []

        # Boutons pour gérer les activités avec style amélioré
        buttons_frame = ttk.Frame(activities_container)
        buttons_frame.pack(fill="x", padx=5, pady=5)

        info_label = ttk.Label(buttons_frame,
                             text="Ajoutez jusqu'à 6 activités principales",
                             font=('Segoe UI', 9))
        info_label.pack(side="left", padx=5)

        add_button = WidgetFactory.create_button(
            buttons_frame,
            text="➕ Ajouter une activité",
            command=self.add_activity_field,
            style='Action.TButton',
            tooltip="Ajouter une nouvelle activité à la société"
        )
        add_button.pack(side="right", padx=5)

        # Ajouter une première activité par défaut
        self.add_activity_field()

        row += 2

    def show_calendar(self, var):
        top = tk.Toplevel(self.root)
        cal = Calendar(top, selectmode='day', date_pattern='dd-mm-yyyy')
        cal.pack(padx=5, pady=5)

        def set_date():
            var.set(cal.get_date())
            top.destroy()

        ttk.Button(top, text="OK", command=set_date).pack(pady=5)

    # Removed initialize_gerant_frame as it's no longer needed in the new structure
        row += 1

    def initialize_associes_frame(self):
        # Titre stylisé pour la section des associés
        associes_label = ttk.Label(self.frame_associes, text="👥 Liste des Associés",
                                 font=('Segoe UI', 12, 'bold'))
        associes_label.pack(pady=(15,10), padx=5)

        # Frame pour contenir la liste des associés
        self.associes_container = ttk.Frame(self.frame_associes)
        self.associes_container.pack(fill="both", expand=True, padx=10, pady=5)

        # Canvas et scrollbar pour la liste des associés
        self.associes_canvas = tk.Canvas(self.associes_container, height=500,
                                       background=self.theme_manager.colors['bg'])
        scrollbar = ttk.Scrollbar(self.associes_container, orient="vertical",
                                command=self.associes_canvas.yview)
        self.associes_canvas.configure(yscrollcommand=scrollbar.set)

        # Frame intérieur pour les associés
        self.associes_inner_frame = ttk.Frame(self.associes_canvas)
        self.associes_inner_frame.bind("<Configure>",
            lambda e: self.associes_canvas.configure(scrollregion=self.associes_canvas.bbox("all")))

        # Placer les widgets
        scrollbar.pack(side="right", fill="y")
        self.associes_canvas.pack(side="left", fill="both", expand=True)
        self.associes_canvas.create_window((0, 0), window=self.associes_inner_frame, anchor="nw")

        # Boutons pour gérer les associés
        buttons_frame = ttk.Frame(self.frame_associes)
        buttons_frame.pack(fill="x", padx=15, pady=(5, 15))

        ttk.Label(buttons_frame,
                text="Ajoutez les associés de la société",
                font=('Segoe UI', 9, 'bold')).pack(side="left", padx=5)

        add_button = WidgetFactory.create_button(
            buttons_frame,
            text="➕ Ajouter un associé",
            command=self.add_associe,
            style='Action.TButton',
            tooltip="Ajouter un nouvel associé à la société"
        )
        add_button.pack(side="right", padx=15)

        # Ajouter toujours un premier associé par défaut
        self.root.after(100, self.add_first_associe)

    def add_first_associe(self):
        """Ajoute toujours un premier associé par défaut"""
        if not self.associes_list:
            self.add_associe()

    def add_associe(self):
        """Ajoute un nouvel associé"""
        # Frame principale avec style card et effet d'ombre
        associe_frame = ttk.Frame(self.associes_inner_frame, style='Card.TFrame')
        associe_frame.pack(fill="x", padx=20, pady=15)

        # Ajouter un effet visuel pour mieux séparer les associés
        separator = ttk.Frame(associe_frame, height=2, style='Separator.TFrame')
        separator.pack(fill="x", padx=15, pady=(0, 10))

        # Variables pour l'associé
        associe_vars = {
            'civil': tk.StringVar(),
            'prenom': tk.StringVar(),
            'nom': tk.StringVar(),
            'parts': tk.StringVar(),
            'nationality': tk.StringVar(),
            'cin_num': tk.StringVar(),
            'cin_validaty': tk.StringVar(),
            'adresse': tk.StringVar(),
            'date_naiss': tk.StringVar(),
            'lieu_naiss': tk.StringVar(),
            'phone': tk.StringVar(),
            'email': tk.StringVar(),
            'is_gerant': tk.BooleanVar(),  # Nouvelle variable pour indiquer si l'associé est gérant
            'quality': tk.StringVar()  # Pour la qualité de l'associé (si gérant)
        }
        self.associes_list.append(associe_vars)

        # En-tête stylisé de l'associé
        num = len(self.associes_list)
        header_frame = ttk.Frame(associe_frame, style='Header.TFrame')
        header_frame.pack(fill="x", padx=10, pady=(5, 15))

        # Titre avec icône et numéro
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side="left")

        ttk.Label(title_frame,
                text=f"👤 Associé {num}",
                font=('Segoe UI', 11, 'bold')).pack(side="left", padx=(15, 5))

        # Indicateur de rôle (optionnel)
        role_label = ttk.Label(title_frame,
                text="Principal" if num == 1 else "Secondaire",
                font=('Segoe UI', 9),
                foreground='#666666')
        role_label.pack(side="left", padx=(0, 15))

        # Grille principale avec meilleur espacement
        grid_frame = ttk.Frame(associe_frame)
        grid_frame.pack(fill="x", padx=20, pady=15)

        # Configuration des colonnes pour un alignement optimal
        for i in range(4):
            grid_frame.grid_columnconfigure(i, weight=1)

        # Section Informations de base avec titre en gras
        info_frame = ttk.LabelFrame(grid_frame)
        info_frame.grid(row=0, column=0, columnspan=4, sticky="nsew", padx=10, pady=10)

        # En-tête de section avec style amélioré
        title_frame = ttk.Frame(info_frame, style='SectionTitle.TFrame')
        title_frame.grid(row=0, column=0, columnspan=4, sticky="ew", padx=10, pady=(8,12))

        # Icône et titre avec description
        icon_label = ttk.Label(title_frame, text="📝", font=('Segoe UI', 12))
        icon_label.pack(side="left", padx=(0, 5))

        title_container = ttk.Frame(title_frame)
        title_container.pack(side="left", fill="x")

        ttk.Label(title_container,
                 text="Informations de base",
                 font=('Segoe UI', 10, 'bold')).pack(anchor="w")
        ttk.Label(title_container,
                 text="Informations principales de l'associé",
                 font=('Segoe UI', 8),
                 foreground='#666666').pack(anchor="w")

        # Première ligne
        row = 1
        # Configuration des colonnes pour un meilleur alignement
        for i in range(4):
            info_frame.grid_columnconfigure(i, weight=1)

        # Champ Civilité avec style amélioré
        field_frame = ttk.Frame(info_frame)
        field_frame.grid(row=row, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

        ttk.Label(field_frame, text="Civilité",
                font=('Segoe UI', 9)).pack(side="left", padx=(0, 10))
        civil_combo = ttk.Combobox(field_frame, textvariable=associe_vars['civil'],
                    values=Civility, width=20)
        civil_combo.pack(side="left", fill="x", expand=True)
        civil_combo.bind('<FocusIn>', lambda e: civil_combo.selection_clear())

        # Champ Prénom avec style amélioré
        field_frame = ttk.Frame(info_frame)
        field_frame.grid(row=row, column=2, columnspan=2, padx=10, pady=5, sticky="ew")

        ttk.Label(field_frame, text="Prénom",
                font=('Segoe UI', 9)).pack(side="left", padx=(0, 10))
        prenom_entry = ttk.Entry(field_frame, textvariable=associe_vars['prenom'],
                width=25)
        prenom_entry.pack(side="left", fill="x", expand=True)

        # Ajouter un tooltip
        ToolTip(prenom_entry, "Entrez le prénom de l'associé")

        # Deuxième ligne
        row += 1
        ttk.Label(info_frame, text="Nom").grid(row=row, column=0, padx=5, pady=2, sticky="e")
        ttk.Entry(info_frame, textvariable=associe_vars['nom'],
                width=25).grid(row=row, column=1, padx=(5,15), pady=2, sticky="w")

        ttk.Label(info_frame, text="Parts (%)").grid(row=row, column=2, padx=5, pady=2, sticky="e")
        ttk.Entry(info_frame, textvariable=associe_vars['parts'],
                width=25).grid(row=row, column=3, padx=(5,15), pady=(2,10), sticky="w")

        # Section Identité
        identity_frame = ttk.LabelFrame(grid_frame)
        identity_frame.grid(row=1, column=0, columnspan=4, sticky="nsew", padx=5, pady=2)

        # Titre compact
        ttk.Label(identity_frame, text="🪪 Identité",
                 font=('Segoe UI', 9, 'bold')).grid(row=0, column=0, columnspan=4, sticky="w", padx=5, pady=2)

        # Configuration de la grille uniforme
        for i in range(4):
            identity_frame.grid_columnconfigure(i, weight=1, uniform="identity_col")

        row = 1
        ttk.Label(identity_frame, text="Nationalité").grid(row=row, column=0, padx=5, pady=2, sticky="e")
        nationality_combo = ttk.Combobox(identity_frame, textvariable=associe_vars['nationality'],
                    values=Nationalite, width=25)
        nationality_combo.grid(row=row, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(identity_frame, text="N° CIN").grid(row=row, column=2, padx=5, pady=2, sticky="e")
        cin_entry = ttk.Entry(identity_frame, textvariable=associe_vars['cin_num'],
                width=25)
        cin_entry.grid(row=row, column=3, padx=5, pady=2, sticky="ew")

        row += 1
        ttk.Label(identity_frame, text="Validité CIN").grid(row=row, column=0, padx=5, pady=2, sticky="e")
        cin_frame = ttk.Frame(identity_frame)
        cin_frame.grid(row=row, column=1, sticky="w", padx=(5,20), pady=5)

        ttk.Entry(cin_frame, textvariable=associe_vars['cin_validaty'],
                width=25).pack(side="left", padx=(0,8))
        ttk.Button(cin_frame, text="📅",
                  command=lambda: self.show_calendar(associe_vars['cin_validaty']),
                  style='Calendar.TButton').pack(side="left")

        # Section Naissance
        birth_frame = ttk.LabelFrame(grid_frame)
        birth_frame.grid(row=2, column=0, columnspan=4, sticky="nsew", padx=5, pady=2)

        # Titre compact
        ttk.Label(birth_frame, text="👶 Naissance",
                 font=('Segoe UI', 9, 'bold')).grid(row=0, column=0, columnspan=4, sticky="w", padx=5, pady=2)

        # Configuration de la grille uniforme
        for i in range(4):
            birth_frame.grid_columnconfigure(i, weight=1, uniform="birth_col")

        row = 1
        ttk.Label(birth_frame, text="Date de naissance").grid(row=row, column=0, padx=5, pady=2, sticky="e")
        birth_date_frame = ttk.Frame(birth_frame)
        birth_date_frame.grid(row=row, column=1, sticky="ew", padx=5, pady=2)

        ttk.Entry(birth_date_frame, textvariable=associe_vars['date_naiss'],
                width=22).pack(side="left", padx=(0,2))
        ttk.Button(birth_date_frame, text="📅",
                  command=lambda: self.show_calendar(associe_vars['date_naiss']),
                  width=3).pack(side="left")

        ttk.Label(birth_frame, text="Lieu de naissance").grid(row=row, column=2, padx=5, pady=2, sticky="e")
        ttk.Entry(birth_frame, textvariable=associe_vars['lieu_naiss'],
                width=25).grid(row=row, column=3, padx=(5,15), pady=(2,10), sticky="w")

        # Section Contact
        contact_frame = ttk.LabelFrame(grid_frame)
        contact_frame.grid(row=3, column=0, columnspan=4, sticky="nsew", padx=5, pady=2)

        # Titre compact
        ttk.Label(contact_frame, text="📞 Contact",
                 font=('Segoe UI', 9, 'bold')).grid(row=0, column=0, columnspan=4, sticky="w", padx=5, pady=2)

        # Configuration de la grille uniforme
        for i in range(4):
            contact_frame.grid_columnconfigure(i, weight=1, uniform="contact_col")

        row = 1
        ttk.Label(contact_frame, text="Adresse").grid(row=row, column=0, padx=5, pady=2, sticky="e")
        ttk.Entry(contact_frame, textvariable=associe_vars['adresse'],
                width=25).grid(row=row, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(contact_frame, text="Téléphone").grid(row=row, column=2, padx=5, pady=2, sticky="e")
        ttk.Entry(contact_frame, textvariable=associe_vars['phone'],
                width=25).grid(row=row, column=3, padx=5, pady=2, sticky="ew")

        row += 1
        ttk.Label(contact_frame, text="Email").grid(row=row, column=0, padx=5, pady=2, sticky="e")
        ttk.Entry(contact_frame, textvariable=associe_vars['email'],
                width=25).grid(row=row, column=1, padx=(5,15), pady=(2,10), sticky="w")

        # Section Gérant avec titre en gras
        gerant_frame = ttk.LabelFrame(grid_frame)
        gerant_frame.grid(row=4, column=0, columnspan=4, sticky="nsew", padx=10, pady=10)

        # Titre stylisé
        title_frame = ttk.Frame(gerant_frame)
        title_frame.grid(row=0, column=0, columnspan=4, sticky="ew", padx=10, pady=(8,12))
        ttk.Label(title_frame, text="👔 Statut de Gérant",
                 font=('Segoe UI', 10, 'bold')).pack(anchor="w")

        def toggle_gerant_fields(*args):
            state = 'normal' if associe_vars['is_gerant'].get() else 'disabled'
            quality_combo['state'] = state

        row = 0
        gerant_check = ttk.Checkbutton(gerant_frame, text="Est Gérant",
                                     variable=associe_vars['is_gerant'])
        gerant_check.grid(row=row, column=0, padx=5, pady=2)

        ttk.Label(gerant_frame, text="Qualité").grid(row=row, column=1, padx=5, pady=2)
        quality_combo = ttk.Combobox(gerant_frame, textvariable=associe_vars['quality'],
                                   values=QualityGerant, width=20, state='disabled')
        quality_combo.grid(row=row, column=2, padx=5, pady=2)

        # Lier la case à cocher avec l'état du combobox
        associe_vars['is_gerant'].trace_add('write', toggle_gerant_fields)

        # Bouton de suppression
        def remove_associe():
            if len(self.associes_list) > 1:  # Garder au moins un associé pour une SARL
                if messagebox.askyesno("Confirmation",
                                    "Voulez-vous vraiment supprimer cet associé ?"):
                    self.associes_list.remove(associe_vars)
                    associe_frame.destroy()
                    self.update_associes_numbers()
            else:
                messagebox.showwarning("Attention",
                                    "Une SARL doit avoir au moins un associé !")

        # Boutons d'action
        actions_frame = ttk.Frame(header_frame)
        actions_frame.pack(side="right", padx=10)

        # Bouton de suppression avec confirmation visuelle
        delete_btn = WidgetFactory.create_button(
            actions_frame,
            text="🗑️ Supprimer",
            command=remove_associe,
            style='Danger.TButton',
            tooltip="Supprimer cet associé"
        )
        delete_btn.pack(side="right", padx=(5, 0))

        # Indicateur de statut (optionnel)
        if associe_vars['is_gerant'].get():
            ttk.Label(actions_frame,
                    text="👑 Gérant",
                    font=('Segoe UI', 9),
                    foreground='#228B22').pack(side="right", padx=10)

    def update_associes_numbers(self):
        """Met à jour les numéros des associés"""
        for i, frame in enumerate(self.associes_inner_frame.winfo_children(), 1):
            header = frame.winfo_children()[0]  # Le premier enfant est le header_frame
            for widget in header.winfo_children():
                if isinstance(widget, ttk.Label):
                    widget.config(text=f"Associé {i}")

    def initialize_contrat_frame(self):
        row = 0
        # Date Contrat
        ttk.Label(self.frame_contrat, text="Date Contrat").grid(row=row, column=0, padx=5, pady=5)
        self.date_contrat_var = tk.StringVar()
        self.date_contrat_entry = ttk.Entry(self.frame_contrat, textvariable=self.date_contrat_var, width=42)
        self.date_contrat_entry.grid(row=row, column=1, padx=5, pady=5)
        ttk.Button(self.frame_contrat, text="Choisir Date",
                  command=lambda: self.show_calendar(self.date_contrat_var)).grid(row=row, column=2, padx=5, pady=5)
        row += 1

        # Période de Contrat
        ttk.Label(self.frame_contrat, text="Période de Contrat").grid(row=row, column=0, padx=5, pady=5)
        self.period_var = tk.StringVar()
        self.period_combo = ttk.Combobox(self.frame_contrat, textvariable=self.period_var, values=Nbmois, width=40)
        self.period_combo.grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Prix mensuel
        ttk.Label(self.frame_contrat, text="Prix mensuel de Contrat").grid(row=row, column=0, padx=5, pady=5)
        self.prix_mensuel_var = tk.StringVar()
        ttk.Entry(self.frame_contrat, textvariable=self.prix_mensuel_var, width=42).grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Prix intermediaire
        ttk.Label(self.frame_contrat, text="Prix intermediaire Contrat").grid(row=row, column=0, padx=5, pady=5)
        self.prix_inter_var = tk.StringVar()
        ttk.Entry(self.frame_contrat, textvariable=self.prix_inter_var, width=42).grid(row=row, column=1, padx=5, pady=5)
        row += 1

        # Date Début Contrat
        ttk.Label(self.frame_contrat, text="Date Début Contrat").grid(row=row, column=0, padx=5, pady=5)
        self.date_debut_var = tk.StringVar()
        self.date_debut_entry = ttk.Entry(self.frame_contrat, textvariable=self.date_debut_var, width=42)
        self.date_debut_entry.grid(row=row, column=1, padx=5, pady=5)
        ttk.Button(self.frame_contrat, text="Choisir Date",
                  command=lambda: self.show_calendar(self.date_debut_var)).grid(row=row, column=2, padx=5, pady=5)
        row += 1

        # Date Fin Contrat
        ttk.Label(self.frame_contrat, text="Date Fin Contrat").grid(row=row, column=0, padx=5, pady=5)
        self.date_fin_var = tk.StringVar()
        self.date_fin_entry = ttk.Entry(self.frame_contrat, textvariable=self.date_fin_var, width=42)
        self.date_fin_entry.grid(row=row, column=1, padx=5, pady=5)
        ttk.Button(self.frame_contrat, text="Choisir Date",
                  command=lambda: self.show_calendar(self.date_fin_var)).grid(row=row, column=2, padx=5, pady=5)
        row += 1

    def add_activity_field(self):
        """Ajoute un nouveau champ d'activité"""
        # Vérifier le nombre maximum d'activités
        if len(self.activity_vars) >= 6:
            messagebox.showwarning("Limite atteinte",
                                 "Vous ne pouvez pas ajouter plus de 6 activités.")
            return

        # Créer un cadre stylisé pour l'activité
        frame = ttk.Frame(self.activities_inner_frame, style='Card.TFrame')
        frame.pack(fill="x", padx=5, pady=3)

        # Créer une nouvelle variable pour l'activité
        var = tk.StringVar()
        self.activity_vars.append(var)

        # Label numéroté
        num_label = ttk.Label(frame,
                             text=f"Activité {len(self.activity_vars)}",
                             font=('Segoe UI', 9))
        num_label.pack(side="left", padx=(5, 10))

        # Créer le combobox avec style amélioré
        combo = ttk.Combobox(frame, textvariable=var,
                            values=Activities, width=35,
                            font=('Segoe UI', 9))
        combo.pack(side="left", padx=(0, 5), pady=2)
        combo.bind('<FocusIn>', lambda e: combo.selection_clear())

        # Bouton de suppression stylisé
        def remove_activity():
            if messagebox.askyesno("Confirmation",
                                 "Voulez-vous vraiment supprimer cette activité ?"):
                self.activity_vars.remove(var)
                frame.destroy()
                # Mettre à jour les numéros des activités
                for i, child in enumerate(self.activities_inner_frame.winfo_children(), 1):
                    for widget in child.winfo_children():
                        if isinstance(widget, ttk.Label):
                            widget.config(text=f"Activité {i}")

        delete_btn = WidgetFactory.create_button(
            frame,
            text="🗑️",
            command=remove_activity,
            style='Danger.TButton',
            tooltip="Supprimer cette activité"
        )
        delete_btn.pack(side="right", padx=5, pady=2)

        # Mettre à jour le scrollregion
        self.activities_canvas.update_idletasks()
        self.activities_canvas.configure(scrollregion=self.activities_canvas.bbox("all"))

    def collect_values(self):
        """Collect all values from the form fields"""
        self.values = {
            # Infos Société
            "DEN_STE": self.den_ste_var.get(),
            "FORME_JUR": self.forme_jur_var.get(),
            "ICE": self.ice_var.get(),
            "DATE_ICE": self.date_ice_var.get(),
            "CAPITAL": self.capital_var.get(),
            "PART_SOCIAL": self.parts_social_var.get(),
            "STE_ADRESS": self.ste_adress_var.get(),
            "TRIBUNAL": self.tribunal_var.get(),

            # Infos Gérant
            "CIVIL": self.civil_var.get(),
            "PRENOM": self.prenom_var.get(),
            "NOM": self.nom_var.get(),
            "CIN_NUM": self.cin_num_var.get(),
            "CIN_VALIDATY": self.cin_validaty_var.get(),
            "GERANT_QUALITY": self.gerant_quality_var.get(),
            "DATE_NAISS": self.date_naiss_var.get(),
            "LIEU_NAISS": self.lieu_naiss_var.get(),
            "GERANT_ADRESS": self.gerant_adress_var.get(),
            "NATIONALITY": self.nationality_var.get(),
            "GERANT_PHONE": self.gerant_phone_var.get(),
            "GERANT_EMAIL": self.gerant_email_var.get(),

            # Infos Contrat
            "DATE_CONTRAT": self.date_contrat_var.get(),
            "PERIOD_DOMCIL": self.period_var.get(),
            "PRIX_CONTRAT": self.prix_mensuel_var.get(),
            "PRIX_INTERMEDIARE_CONTRAT": self.prix_inter_var.get(),
            "DOM_DATEDEB": self.date_debut_var.get(),
            "DOM_DATEFIN": self.date_fin_var.get()
        }

        # Add activities
        for i, var in enumerate(self.activity_vars, 1):
            self.values[f"ACTIVITY{i}"] = var.get()
            if i >= 6:  # Limiter à 6 activités maximum pour la compatibilité
                break

        # Add associés
        for i, associe in enumerate(self.associes_list, 1):
            prefix = f"ASSOCIE{i}_"
            self.values[f"{prefix}CIVIL"] = associe['civil'].get()
            self.values[f"{prefix}PRENOM"] = associe['prenom'].get()
            self.values[f"{prefix}NOM"] = associe['nom'].get()
            self.values[f"{prefix}PARTS"] = associe['parts'].get()
            self.values[f"{prefix}NATIONALITY"] = associe['nationality'].get()
            self.values[f"{prefix}CIN_NUM"] = associe['cin_num'].get()
            self.values[f"{prefix}CIN_VALIDATY"] = associe['cin_validaty'].get()
            self.values[f"{prefix}ADRESSE"] = associe['adresse'].get()
            self.values[f"{prefix}DATE_NAISS"] = associe['date_naiss'].get()
            self.values[f"{prefix}LIEU_NAISS"] = associe['lieu_naiss'].get()
            self.values[f"{prefix}PHONE"] = associe['phone'].get()
            self.values[f"{prefix}EMAIL"] = associe['email'].get()
            self.values[f"{prefix}IS_GERANT"] = "Oui" if associe['is_gerant'].get() else "Non"
            self.values[f"{prefix}QUALITY"] = associe['quality'].get() if associe['is_gerant'].get() else ""

    def get_next_id(self, workbook, sheet_name):
        """Get the next available ID for a given sheet"""
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.append(excel_sheets[sheet_name])  # Ajouter les en-têtes
            return 1

        sheet = workbook[sheet_name]
        if sheet.max_row == 1:  # Seulement les en-têtes
            return 1
        # Trouver le plus grand ID et ajouter 1
        id_column = sheet['A']  # La colonne ID est toujours la première
        max_id = max(int(cell.value or 0) for cell in id_column[1:])  # Ignorer l'en-tête
        return max_id + 1

    def save_to_sheet(self, workbook, sheet_name, data, headers):
        """Save data to a specific sheet"""
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.append(headers)
        else:
            sheet = workbook[sheet_name]
            if sheet.max_row == 0:
                sheet.append(headers)

        # Préparer les données dans l'ordre des en-têtes
        row_data = [data.get(header, '') for header in headers]
        sheet.append(row_data)

    def insert_data_to_excel(self):
        """Insert form data into Excel file with multiple sheets"""
        try:
            excel_file_path = os.path.join(os.path.dirname(__file__), 'databases', 'DataBase_domiciliation.xlsx')
            os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)

            if not os.path.exists(excel_file_path):
                workbook = openpyxl.Workbook()
                workbook.save(excel_file_path)

            workbook = openpyxl.load_workbook(excel_file_path)

            # Créer ou obtenir l'ID de la société
            societe_id = self.get_next_id(workbook, "Societes")

            # Données de la société
            societe_data = {
                "ID_SOCIETE": societe_id,
                "DEN_STE": self.values.get("DEN_STE"),
                "FORME_JUR": self.values.get("FORME_JUR"),
                "ICE": self.values.get("ICE"),
                "DATE_ICE": self.values.get("DATE_ICE"),
                "CAPITAL": self.values.get("CAPITAL"),
                "PART_SOCIAL": self.values.get("PART_SOCIAL"),
                "STE_ADRESS": self.values.get("STE_ADRESS"),
                "TRIBUNAL": self.values.get("TRIBUNAL")
            }

            # Sauvegarder les données de la société
            self.save_to_sheet(workbook, "Societes", societe_data, societe_headers)

            # Sauvegarder les données des associés
            for i, associe in enumerate(self.associes_list, 1):
                associe_id = self.get_next_id(workbook, "Associes")
                prefix = f"ASSOCIE{i}_"
                associe_data = {
                    "ID_ASSOCIE": associe_id,
                    "ID_SOCIETE": societe_id,
                    "CIVIL": self.values.get(f"{prefix}CIVIL"),
                    "PRENOM": self.values.get(f"{prefix}PRENOM"),
                    "NOM": self.values.get(f"{prefix}NOM"),
                    "NATIONALITY": self.values.get(f"{prefix}NATIONALITY"),
                    "CIN_NUM": self.values.get(f"{prefix}CIN_NUM"),
                    "CIN_VALIDATY": self.values.get(f"{prefix}CIN_VALIDATY"),
                    "DATE_NAISS": self.values.get(f"{prefix}DATE_NAISS"),
                    "LIEU_NAISS": self.values.get(f"{prefix}LIEU_NAISS"),
                    "ADRESSE": self.values.get(f"{prefix}ADRESSE"),
                    "PHONE": self.values.get(f"{prefix}PHONE"),
                    "EMAIL": self.values.get(f"{prefix}EMAIL"),
                    "PARTS": self.values.get(f"{prefix}PARTS"),
                    "IS_GERANT": self.values.get(f"{prefix}IS_GERANT"),
                    "QUALITY": self.values.get(f"{prefix}QUALITY")
                }
                self.save_to_sheet(workbook, "Associes", associe_data, associe_headers)

            # Sauvegarder les données du contrat
            contrat_id = self.get_next_id(workbook, "Contrats")
            contrat_data = {
                "ID_CONTRAT": contrat_id,
                "ID_SOCIETE": societe_id,
                "DATE_CONTRAT": self.values.get("DATE_CONTRAT"),
                "PERIOD_DOMCIL": self.values.get("PERIOD_DOMCIL"),
                "PRIX_CONTRAT": self.values.get("PRIX_CONTRAT"),
                "PRIX_INTERMEDIARE_CONTRAT": self.values.get("PRIX_INTERMEDIARE_CONTRAT"),
                "DOM_DATEDEB": self.values.get("DOM_DATEDEB"),
                "DOM_DATEFIN": self.values.get("DOM_DATEFIN")
            }
            self.save_to_sheet(workbook, "Contrats", contrat_data, contrat_headers)

            workbook.save(excel_file_path)

        except Exception as e:
            messagebox.showerror("Erreur",
                f"Erreur lors de l'insertion dans Excel : {str(e)}")

    def generer_docs(self):
        """Handle document generation with custom save location"""
        self.collect_values()

        try:
            if not self.values['DEN_STE']:
                messagebox.showerror("Erreur", "Veuillez entrer un nom de société!")
                return

            # Ask user for save location
            folder_name = f"{date_inverser}_DCons_{self.values['DEN_STE']}"
            folder_path = tk.filedialog.askdirectory(
                title="Choisir l'emplacement pour sauvegarder les documents Word"
            )

            if not folder_path:  # User cancelled
                return

            folder_path = os.path.join(folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Generate DOCX files
            documents = {
                'Annonce_légale': docAnnonceJournal,
                'Attestation_Domiciliation': docAttestDomicil,
                'Contrat_Domiciliation': docContratDomicil,
                'Déclaration_Imm_Rc': docDeclImmRc,
                'Dépot_légal': docDepotLegalRc,
                'Statuts': docStatuts
            }

            for doc_name, template in documents.items():
                template.render(self.values)
                output_path = os.path.join(folder_path,
                    f"{date_inverser}_{doc_name}_{self.values['DEN_STE']}.docx")
                template.save(output_path)

            messagebox.showinfo("Succès",
                f"Documents créés avec succès dans le dossier:\n{folder_path}")

        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def save_as_pdf(self):
        """Generate or convert documents to PDF with custom save location"""
        try:
            self.collect_values()

            if not self.values['DEN_STE']:
                messagebox.showerror("Erreur", "Veuillez entrer un nom de société!")
                return

            # Ask user for save location
            folder_name = f"{date_inverser}_DCons_{self.values['DEN_STE']}"
            folder_path = tk.filedialog.askdirectory(
                title="Choisir l'emplacement pour sauvegarder les documents PDF"
            )

            if not folder_path:  # User cancelled
                return

            folder_path = os.path.join(folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # First generate DOCX files if they don't exist
            documents = {
                'Annonce_légale': docAnnonceJournal,
                'Attestation_Domiciliation': docAttestDomicil,
                'Contrat_Domiciliation': docContratDomicil,
                'Déclaration_Imm_Rc': docDeclImmRc,
                'Dépot_légal': docDepotLegalRc,
                'Statuts': docStatuts
            }

            for doc_name, template in documents.items():
                # Create DOCX
                template.render(self.values)
                docx_path = os.path.join(folder_path,
                    f"{date_inverser}_{doc_name}_{self.values['DEN_STE']}.docx")
                template.save(docx_path)

                # Convert to PDF
                pdf_path = docx_path.replace('.docx', '.pdf')
                convert(docx_path, pdf_path)

            messagebox.showinfo("Succès",
                f"Documents PDF créés avec succès dans le dossier:\n{folder_path}")

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération des PDF : {str(e)}")

    def save_to_database(self):
        """Save current form data to database"""
        try:
            self.collect_values()

            if not self.values['DEN_STE']:
                messagebox.showerror("Erreur", "Veuillez entrer un nom de société!")
                return

            # Vérifier si une société avec le même nom existe déjà
            excel_file_path = os.path.join(os.path.dirname(__file__), 'databases', 'DataBase_domiciliation.xlsx')
            if os.path.exists(excel_file_path):
                existing_df = pd.read_excel(excel_file_path, sheet_name="DataBaseDom")
                if not existing_df.empty and self.values['DEN_STE'] in existing_df['DEN_STE'].values:
                    messagebox.showerror("Erreur", "Une société avec ce nom existe déjà dans la base de données!")
                    return

            self.insert_data_to_excel()
            messagebox.showinfo("Succès", "Données sauvegardées avec succès dans la base de données!")

        except Exception as e:
            messagebox.showerror("Erreur",
                f"Erreur lors de la sauvegarde dans la base de données : {str(e)}")

    def clear_form(self):
        """Clear all form fields"""
        # Clear main fields
        for var_name in ['den_ste_var', 'forme_jur_var', 'ice_var', 'date_ice_var',
                        'capital_var', 'part_social_var', 'ste_adress_var', 'tribunal_var',
                        'date_contrat_var', 'period_var', 'prix_contrat_var', 'prix_intermediaire_var',
                        'date_debut_var', 'date_fin_var']:
            if hasattr(self, var_name):
                getattr(self, var_name).set('')

        # Clear activities
        for frame in self.activities_inner_frame.winfo_children():
            frame.destroy()
        self.activity_vars.clear()
        # Ajouter une nouvelle activité vide
        self.add_activity_field()

        # Clear associés
        for frame in self.associes_inner_frame.winfo_children():
            frame.destroy()
        self.associes_list.clear()
        # Ajouter un nouvel associé vide si c'est une SARL
        if self.forme_jur_var.get() == "SARL":
            self.add_associe()

# Create and run the application
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1200x800")
    app = DomiciliationApp(root)
    root.mainloop()
